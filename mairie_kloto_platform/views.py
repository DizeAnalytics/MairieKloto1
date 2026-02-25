from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDay
from django.utils import timezone
from datetime import timedelta, datetime, date
from decimal import Decimal, InvalidOperation
import os
import json
from django.core.serializers.json import DjangoJSONEncoder
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas as pdfcanvas
from mairie.models import (
    ConfigurationMairie,
    VisiteSite,
    CampagnePublicitaire,
    Publicite,
    Suggestion,
    NewsletterSubscription,
    AgentCollecteur,
    Contribuable,
    BoutiqueMagasin,
    CotisationAnnuelle,
    PaiementCotisation,
    TicketMarche,
    EmplacementMarche,
    CotisationAnnuelleActeur,
    CotisationAnnuelleInstitution,
    PaiementCotisationActeur,
    PaiementCotisationInstitution,
    DirectionMairie,
    DivisionDirection,
    SectionDirection,
    PersonnelSection,
    ServiceSection,
    CartographieCommune,
    InfrastructureCommune,
)

from acteurs.models import ActeurEconomique, InstitutionFinanciere
from emploi.models import ProfilEmploi
from mairie.models import Candidature, AppelOffre
from mairie.forms import (
    DirectionMairieForm,
    DivisionDirectionForm,
    SectionDirectionForm,
    PersonnelSectionForm,
    ServiceSectionForm,
)
from comptes.models import Notification
from diaspora.models import MembreDiaspora
from osc.models import OrganisationSocieteCivile, OSC_TYPE_CHOICES, get_osc_type_display
from django.utils.html import escape
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class NumberedCanvas(pdfcanvas.Canvas):
    def __init__(self, *args, **kwargs):
        pdfcanvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            pdfcanvas.Canvas.showPage(self)
        pdfcanvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        width, height = self._pagesize
        self.drawCentredString(width / 2, 30, f"Page {self._pageNumber} / {page_count}")


# Hauteur réservée à l'en-tête PDF (logo + titres + contacts + trait jaune + marge)
# Le contenu (ex. "Fiche Institution Financière...") commence en dessous.
PDF_HEADER_HEIGHT_CM = 5.5


def _draw_pdf_header(c, d, conf=None):
    """
    Dessine l'en-tête standard pour tous les PDF : logo centré, nom de la mairie,
    contacts (adresse, téléphone, email) et un trait jaune horizontal séparant l'en-tête du contenu.
    """
    if conf is None:
        conf = ConfigurationMairie.objects.filter(est_active=True).first()

    width, height = d.pagesize
    y = height - 35

    c.saveState()

    # Logo centré en haut
    if conf and getattr(conf, "logo", None) and getattr(conf.logo, "path", None):
        try:
            logo_w = 50
            logo_h = 50
            c.drawImage(
                conf.logo.path,
                (width - logo_w) / 2,
                y - logo_h,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto",
            )
            y -= logo_h + 12  # Espacement après le logo
        except Exception:
            pass

    # Nom de la mairie / entête république
    commune = getattr(conf, "nom_commune", None) or "Mairie de Kloto 1"
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y, f"République Togolaise – {commune}")
    y -= 20  # Espacement après le titre

    # Ligne de contacts (adresse, téléphone, email de la mairie)
    if conf:
        contact_parts = []
        if getattr(conf, "adresse", None):
            contact_parts.append(conf.adresse)
        if getattr(conf, "telephone", None):
            contact_parts.append(f"Tél: {conf.telephone}")
        if getattr(conf, "email", None):
            contact_parts.append(f"Email: {conf.email}")
        if contact_parts:
            c.setFont("Helvetica", 9)
            c.drawCentredString(width / 2, y, " | ".join(contact_parts))
            y -= 20  # Espacement après les contacts

    # Trait jaune horizontal juste après la ligne de contacts
    line_y = y - 12  # Espace avant le trait
    c.setStrokeColorRGB(1, 0.85, 0)  # Jaune doré
    c.setLineWidth(2)
    c.line(40, line_y, width - 40, line_y)

    # Le contenu (titre du document, ex. "Fiche Institution Financière...") commence
    # en dessous, grâce à topMargin = PDF_HEADER_HEIGHT_CM

    # Date d'édition en bas à gauche
    c.setFont("Helvetica", 8)
    c.drawString(40, 20, timezone.now().strftime("Édité le %d/%m/%Y %H:%M"))

    c.restoreState()

def _format_pdf_value(value):
    if value in (None, "", [], ()):
        return "Non renseigné"
    if isinstance(value, bool):
        text = "Oui" if value else "Non"
    elif isinstance(value, datetime):
        text = value.strftime("%d/%m/%Y %H:%M")
    elif isinstance(value, date):
        text = value.strftime("%d/%m/%Y")
    elif isinstance(value, (list, tuple, set)):
        flattened = ", ".join(str(item) for item in value if item)
        text = flattened or "Non renseigné"
    else:
        text = str(value)
    # Escape HTML but preserve line breaks
    escaped = escape(text)
    return escaped.replace("\n", "<br/>")


def _make_pdf_filename(prefix, label):
    base = slugify(label) if label else ""
    if not base:
        base = slugify(prefix) or "detail"
    return f"{prefix}_{base}.pdf"


def _build_detail_pdf(filename, title, sections):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    # Top margin augmentée pour laisser la place à l'en-tête (logo + texte)
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DetailTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=18,
    )
    section_style = ParagraphStyle(
        "DetailSection",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#004d28"),
        spaceBefore=12,
        spaceAfter=8,
    )
    label_style = ParagraphStyle(
        "DetailLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#004d28"),
    )
    value_style = ParagraphStyle(
        "DetailValue",
        parent=styles["Normal"],
        leading=14,
    )
    story = [Paragraph(title, title_style)]

    for section_title, items in sections:
        if not items:
            continue
        story.append(Paragraph(section_title, section_style))
        table_data = []
        for label, value in items:
            table_data.append(
                [
                    Paragraph(str(escape(label)), label_style),
                    Paragraph(_format_pdf_value(value), value_style),
                ]
            )
        table = Table(table_data, colWidths=[6 * cm, 10.5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F5E9")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C8E6C9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.25 * cm))

    if not sections:
        story.append(Paragraph("Aucune information disponible.", styles["Normal"]))

    conf = ConfigurationMairie.objects.filter(est_active=True).first()

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


def home(request):
    """Page d'accueil de la plateforme (page Enregistrement)."""

    context = {}
    return render(request, "mairie-kloto-platform.html", context)


def fake_admin(request):
    """Fausse route admin pour sécuriser l'accès à l'administration Django."""
    return render(request, "admin_fake.html", status=404)


def politique_cookies(request):
    """
    Page d'information sur les cookies (conformité / transparence).
    Le consentement est géré côté client via le cookie 'cookie_consent'.
    """
    return render(request, "legal/politique_cookies.html", {})


@require_POST
def newsletter_subscribe(request):
    """
    Inscription à la newsletter via le champ email du popup de publicité.
    - Enregistre l'email s'il n'existe pas déjà.
    - Marque l'abonné comme actif.
    - Redirige vers la page précédente (ou l'accueil) avec un message flash.
    """
    email = (request.POST.get("email") or "").strip()
    referer = request.META.get("HTTP_REFERER") or "/"

    if not email:
        messages.error(
            request,
            "Veuillez saisir une adresse email valide pour la newsletter.",
        )
        return redirect(referer)

    cookie_should_be_set = False

    try:
        abonnee, created = NewsletterSubscription.objects.get_or_create(
            email__iexact=email,
            defaults={"email": email, "source": "popup"},
        )
        if not created:
            if not abonnee.est_actif:
                abonnee.est_actif = True
                abonnee.source = abonnee.source or "popup"
                abonnee.save(update_fields=["est_actif", "source"])
            cookie_should_be_set = True
            messages.info(
                request,
                "Cette adresse email est déjà inscrite à la newsletter de la mairie."
            )
        else:
            cookie_should_be_set = True
            messages.success(
                request,
                "Merci ! Votre adresse email a bien été enregistrée pour recevoir les actualités de la mairie."
            )
    except Exception:
        messages.error(
            request,
            "Une erreur est survenue lors de votre inscription à la newsletter. Merci de réessayer plus tard."
        )
        return redirect(referer)

    response = redirect(referer)

    # Marquer côté navigateur que la newsletter est déjà souscrite
    if cookie_should_be_set:
        # 1 an
        max_age = 365 * 24 * 60 * 60
        response.set_cookie(
            "newsletter_subscribed",
            "1",
            max_age=max_age,
            samesite="Lax",
        )

    return response


def is_staff_user(user):
    """Vérifie si l'utilisateur est staff ou superuser."""
    return user.is_authenticated and (user.is_staff or user.is_superuser)


@login_required
@user_passes_test(is_staff_user)
def newsletters_admin(request):
    """
    Liste des inscriptions à la newsletter pour le tableau de bord.
    Affiche les emails dans un tableau et propose un bouton 'Envoyer une newsletter'
    qui ouvre la boîte mail de la mairie avec tous les emails en copie cachée.
    """
    abonnements = NewsletterSubscription.objects.order_by("-date_inscription")
    emails_actifs = [a.email for a in abonnements if a.est_actif]

    mairie_config = ConfigurationMairie.objects.filter(est_active=True).first()
    mairie_email = getattr(mairie_config, "email", "") or "contact@mairiekloto1.tg"

    # Chaîne d'emails séparés par des virgules pour le BCC du mailto
    bcc_emails = ",".join(emails_actifs)

    context = {
        "abonnements": abonnements,
        "emails_actifs": emails_actifs,
        "bcc_emails": bcc_emails,
        "mairie_email": mairie_email,
    }
    return render(request, "admin/newsletters.html", context)


@login_required
@user_passes_test(is_staff_user)
def tableau_bord(request):
    """Tableau de bord administrateur."""
    
    # Statistiques générales
    stats = {
        'acteurs_economiques': ActeurEconomique.objects.count(),
        'institutions_financieres': InstitutionFinanciere.objects.count(),
        'jeunes': ProfilEmploi.objects.filter(type_profil='jeune').count(),
        'retraites': ProfilEmploi.objects.filter(type_profil='retraite').count(),
        'diaspora': MembreDiaspora.objects.count(),
        'osc': OrganisationSocieteCivile.objects.count(),
        'candidatures': Candidature.objects.count(),
        'suggestions': Suggestion.objects.count(),
        'agents_collecteurs': AgentCollecteur.objects.count(),
        'contribuables': Contribuable.objects.count(),
        'boutiques_magasins': BoutiqueMagasin.objects.count(),
        'cotisations_annuelles': CotisationAnnuelle.objects.count(),
        'paiements_cotisations': PaiementCotisation.objects.count(),
        'tickets_marche': TicketMarche.objects.count(),
        'directions_mairie': DirectionMairie.objects.count(),
        'sections_mairie': SectionDirection.objects.count(),
        'personnels_sections': PersonnelSection.objects.count(),
        'infrastructures_commune': 0,
        'total_inscriptions': (
            ActeurEconomique.objects.count() +
            InstitutionFinanciere.objects.count() +
            ProfilEmploi.objects.count() +
            MembreDiaspora.objects.count() +
            OrganisationSocieteCivile.objects.count()
        ),
    }

    # Données pour le graphique (30 derniers jours)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    # Génération des dates
    dates = [(start_date + timedelta(days=i)).date() for i in range(31)]
    labels = [d.strftime('%d/%m') for d in dates]
    
    def get_counts(queryset, date_field):
        """
        Retourne une liste de 31 valeurs (une par jour) correspondant au nombre
        d'objets du queryset par jour entre start_date et end_date (inclus).
        Utilise un alias 'day' pour éviter les conflits avec d'éventuels champs 'date'.
        """
        data = (
            queryset.filter(**{f"{date_field}__gte": start_date})
            .annotate(day=TruncDay(date_field))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )

        count_dict = {item["day"].date(): item["count"] for item in data if item["day"]}
        return [count_dict.get(d, 0) for d in dates]

    chart_data = {
        'labels': labels,
        'acteurs': get_counts(ActeurEconomique.objects.all(), 'date_enregistrement'),
        'institutions': get_counts(InstitutionFinanciere.objects.all(), 'date_enregistrement'),
        'jeunes': get_counts(ProfilEmploi.objects.filter(type_profil='jeune'), 'date_inscription'),
        'retraites': get_counts(ProfilEmploi.objects.filter(type_profil='retraite'), 'date_inscription'),
        'diaspora': get_counts(MembreDiaspora.objects.all(), 'date_inscription'),
        'osc': get_counts(OrganisationSocieteCivile.objects.all(), 'date_enregistrement'),
        'visites': get_counts(VisiteSite.objects.all(), 'date'),
    }
    
    # Nombre total de visites sur les 30 derniers jours (toutes pages confondues)
    total_visites_30j = VisiteSite.objects.filter(date__gte=start_date, date__lte=end_date).count()

    # Données pour la carte : acteurs économiques et institutions avec géolocalisation
    map_markers = []
    for a in ActeurEconomique.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True):
        try:
            map_markers.append({
                "nom": a.raison_sociale,
                "lat": float(a.latitude),
                "lng": float(a.longitude),
                "type": "acteur",
            })
        except (TypeError, ValueError):
            pass
    for i in InstitutionFinanciere.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True):
        try:
            map_markers.append({
                "nom": i.nom_institution,
                "lat": float(i.latitude),
                "lng": float(i.longitude),
                "type": "institution",
            })
        except (TypeError, ValueError):
            pass

    # Données pour la carte des infrastructures de la commune
    infrastructures_data = []
    infra_center = {"lat": 6.9057, "lng": 0.6287, "zoom": 12}

    config_active = ConfigurationMairie.objects.filter(est_active=True).first()
    cartographie = None
    if config_active:
        try:
            cartographie = CartographieCommune.objects.get(configuration=config_active)
        except CartographieCommune.DoesNotExist:
            cartographie = None

    if cartographie:
        try:
            infra_center["lat"] = float(cartographie.centre_latitude)
            infra_center["lng"] = float(cartographie.centre_longitude)
            infra_center["zoom"] = int(cartographie.zoom_carte or 13)
        except (TypeError, ValueError):
            infra_center["lat"] = 6.9057
            infra_center["lng"] = 0.6287
            infra_center["zoom"] = 13

        infrastructures_qs = InfrastructureCommune.objects.filter(
            cartographie=cartographie
        ).order_by("type_infrastructure", "nom")
        stats["infrastructures_commune"] = infrastructures_qs.count()
        for infra in infrastructures_qs:
            try:
                infrastructures_data.append(
                    {
                        "id": infra.id,
                        "nom": infra.nom,
                        "type": infra.type_infrastructure,
                        "description": infra.description,
                        "adresse": infra.adresse,
                        "lat": float(infra.latitude),
                        "lng": float(infra.longitude),
                        "est_active": infra.est_active,
                    }
                )
            except (TypeError, ValueError):
                continue

    context = {
        'stats': stats,
        'chart_data_json': json.dumps(chart_data, cls=DjangoJSONEncoder),
        'total_visites_30j': total_visites_30j,
        'map_markers_json': json.dumps(map_markers, ensure_ascii=False),
        'infrastructures_json': json.dumps(infrastructures_data, ensure_ascii=False),
        'infrastructures_center_json': json.dumps(infra_center, ensure_ascii=False),
    }
    
    return render(request, "admin/tableau_bord.html", context)


@login_required
@user_passes_test(is_staff_user)
def tableau_bord_organigramme(request):
    """
    Page de gestion de l'organigramme (Directions, Sections, Personnel)
    depuis le tableau de bord, avec un design proche de la page Boutiques.
    """

    message_success = None
    message_error = None

    # Champ de recherche simple (sections / directions / personnel / services)
    q = request.GET.get("q", "").strip()

    # Identifiants/objets éventuels à éditer (mode édition sur la même page)
    editing_direction = None
    editing_section = None
    editing_division = None
    editing_personnel = None
    editing_service = None

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "create_direction":
            direction_form = DirectionMairieForm(request.POST)
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if direction_form.is_valid():
                direction_form.save()
                message_success = "La direction a été créée avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la création de la direction. Veuillez vérifier les informations."

        elif action == "update_direction":
            direction_id = request.POST.get("direction_id")
            editing_direction = get_object_or_404(DirectionMairie, pk=direction_id)
            direction_form = DirectionMairieForm(request.POST, instance=editing_direction)
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if direction_form.is_valid():
                direction_form.save()
                message_success = "La direction a été mise à jour avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la mise à jour de la direction. Veuillez vérifier les informations."

        elif action == "create_division":
            division_form = DivisionDirectionForm(request.POST)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if division_form.is_valid():
                division_form.save()
                message_success = "La division a été créée avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la création de la division. Veuillez vérifier les informations."

        elif action == "update_division":
            division_id = request.POST.get("division_id")
            editing_division = get_object_or_404(DivisionDirection, pk=division_id)
            division_form = DivisionDirectionForm(request.POST, instance=editing_division)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if division_form.is_valid():
                division_form.save()
                message_success = "La division a été mise à jour avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la mise à jour de la division. Veuillez vérifier les informations."

        elif action == "create_section":
            section_form = SectionDirectionForm(request.POST)
            direction_form = DirectionMairieForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if section_form.is_valid():
                section_form.save()
                message_success = "La section a été créée avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la création de la section. Veuillez vérifier les informations."

        elif action == "update_section":
            section_id = request.POST.get("section_id")
            editing_section = get_object_or_404(SectionDirection, pk=section_id)
            section_form = SectionDirectionForm(request.POST, instance=editing_section)
            direction_form = DirectionMairieForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            if section_form.is_valid():
                section_form.save()
                message_success = "La section a été mise à jour avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la mise à jour de la section. Veuillez vérifier les informations."

        elif action == "create_personnel":
            personnel_form = PersonnelSectionForm(request.POST)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            service_form = ServiceSectionForm()
            if personnel_form.is_valid():
                personnel_form.save()
                message_success = "Le membre du personnel a été créé avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la création du membre du personnel. Veuillez vérifier les informations."

        elif action == "update_personnel":
            personnel_id = request.POST.get("personnel_id")
            editing_personnel = get_object_or_404(PersonnelSection, pk=personnel_id)
            personnel_form = PersonnelSectionForm(request.POST, instance=editing_personnel)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            service_form = ServiceSectionForm()
            if personnel_form.is_valid():
                personnel_form.save()
                message_success = "Le membre du personnel a été mis à jour avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la mise à jour du membre du personnel. Veuillez vérifier les informations."

        elif action == "create_service":
            service_form = ServiceSectionForm(request.POST)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            if service_form.is_valid():
                service_form.save()
                message_success = "Le service a été créé avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la création du service. Veuillez vérifier les informations."

        elif action == "update_service":
            service_id = request.POST.get("service_id")
            editing_service = get_object_or_404(ServiceSection, pk=service_id)
            service_form = ServiceSectionForm(request.POST, instance=editing_service)
            direction_form = DirectionMairieForm()
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            if service_form.is_valid():
                service_form.save()
                message_success = "Le service a été mis à jour avec succès."
                return redirect("tableau_bord_organigramme")
            else:
                message_error = "Erreur lors de la mise à jour du service. Veuillez vérifier les informations."

        else:
            # Action inconnue
            direction_form = DirectionMairieForm()
            division_form = DivisionDirectionForm()
            section_form = SectionDirectionForm()
            personnel_form = PersonnelSectionForm()
            service_form = ServiceSectionForm()
            message_error = "Action non reconnue."
    else:
        # Mode édition depuis les paramètres GET (ex: ?edit_direction=1)
        edit_direction_id = request.GET.get("edit_direction")
        edit_division_id = request.GET.get("edit_division")
        edit_section_id = request.GET.get("edit_section")
        edit_personnel_id = request.GET.get("edit_personnel")
        edit_service_id = request.GET.get("edit_service")

        if edit_direction_id:
            editing_direction = get_object_or_404(DirectionMairie, pk=edit_direction_id)
            direction_form = DirectionMairieForm(instance=editing_direction)
        else:
            direction_form = DirectionMairieForm()

        if edit_division_id:
            editing_division = get_object_or_404(DivisionDirection, pk=edit_division_id)
            division_form = DivisionDirectionForm(instance=editing_division)
        else:
            division_form = DivisionDirectionForm()

        if edit_section_id:
            editing_section = get_object_or_404(SectionDirection, pk=edit_section_id)
            section_form = SectionDirectionForm(instance=editing_section)
        else:
            section_form = SectionDirectionForm()

        if edit_personnel_id:
            editing_personnel = get_object_or_404(PersonnelSection, pk=edit_personnel_id)
            personnel_form = PersonnelSectionForm(instance=editing_personnel)
        else:
            personnel_form = PersonnelSectionForm()

        if edit_service_id:
            editing_service = get_object_or_404(ServiceSection, pk=edit_service_id)
            service_form = ServiceSectionForm(instance=editing_service)
        else:
            service_form = ServiceSectionForm()

    directions_qs = (
        DirectionMairie.objects.all()
        .annotate(total_personnels=Count("sections__personnels", distinct=True))
        .prefetch_related("divisions__sections__personnels")
        .order_by("ordre_affichage", "nom")
    )

    divisions_qs = (
        DivisionDirection.objects.select_related("direction")
        .prefetch_related("sections")
        .order_by("direction__ordre_affichage", "ordre_affichage", "nom")
    )

    sections_qs = (
        SectionDirection.objects.select_related("direction", "division")
        .prefetch_related("personnels", "services")
        .order_by(
            "direction__ordre_affichage",
            "division__ordre_affichage",
            "ordre_affichage",
            "nom",
        )
    )

    personnels_qs = (
        PersonnelSection.objects.select_related(
            "section",
            "section__direction",
            "section__division",
        ).order_by(
            "section__direction__ordre_affichage",
            "section__division__ordre_affichage",
            "section__ordre_affichage",
            "ordre_affichage",
            "nom_prenoms",
        )
    )

    if q:
        # Filtrer les directions liées à la recherche (nom, sigle, chef, sections, personnel)
        directions_qs = directions_qs.filter(
            Q(nom__icontains=q)
            | Q(sigle__icontains=q)
            | Q(chef_direction__icontains=q)
            | Q(divisions__nom__icontains=q)
            | Q(divisions__sigle__icontains=q)
            | Q(divisions__chef_division__icontains=q)
            | Q(sections__nom__icontains=q)
            | Q(sections__sigle__icontains=q)
            | Q(sections__chef_section__icontains=q)
            | Q(sections__personnels__nom_prenoms__icontains=q)
            | Q(sections__personnels__fonction__icontains=q)
        ).distinct()

        # Filtrer les divisions (nom, sigle, chef, direction)
        divisions_qs = divisions_qs.filter(
            Q(nom__icontains=q)
            | Q(sigle__icontains=q)
            | Q(chef_division__icontains=q)
            | Q(direction__nom__icontains=q)
            | Q(direction__sigle__icontains=q)
        ).distinct()

        # Filtrer les sections (nom, sigle, chef, direction, personnel, services)
        sections_qs = sections_qs.filter(
            Q(nom__icontains=q)
            | Q(sigle__icontains=q)
            | Q(chef_section__icontains=q)
            | Q(direction__nom__icontains=q)
            | Q(direction__sigle__icontains=q)
            | Q(division__nom__icontains=q)
            | Q(division__sigle__icontains=q)
            | Q(division__chef_division__icontains=q)
            | Q(personnels__nom_prenoms__icontains=q)
            | Q(personnels__fonction__icontains=q)
            | Q(services__titre__icontains=q)
        ).distinct()

        # Filtrer le personnel (nom, fonction, section, direction)
        personnels_qs = personnels_qs.filter(
            Q(nom_prenoms__icontains=q)
            | Q(fonction__icontains=q)
            | Q(section__nom__icontains=q)
            | Q(section__division__nom__icontains=q)
            | Q(section__division__sigle__icontains=q)
            | Q(section__direction__nom__icontains=q)
            | Q(section__direction__sigle__icontains=q)
        ).distinct()

    directions = directions_qs
    divisions = divisions_qs
    sections = sections_qs
    personnels = personnels_qs

    context = {
        "titre": "Organigramme de la Mairie",
        "direction_form": direction_form,
        "division_form": division_form,
        "section_form": section_form,
        "personnel_form": personnel_form,
        "directions": directions,
        "divisions": divisions,
        "sections": sections,
        "personnels": personnels,
        "message_success": message_success,
        "message_error": message_error,
        "editing_direction": editing_direction,
        "editing_division": editing_division,
        "editing_section": editing_section,
        "editing_personnel": editing_personnel,
        "editing_service": editing_service,
        "service_form": service_form,
        "current_filters": {
            "q": q,
        },
    }

    return render(request, "admin/organigramme_mairie.html", context)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_organigramme(request):
    """
    Export PDF de l'organigramme sous forme de tableau.
    """
    q = request.GET.get("q", "").strip()

    sections = (
        SectionDirection.objects.select_related("direction", "division")
        .prefetch_related("personnels")
        .order_by(
            "direction__ordre_affichage",
            "division__ordre_affichage",
            "ordre_affichage",
            "nom",
        )
    )

    if q:
        sections = sections.filter(
            Q(nom__icontains=q)
            | Q(sigle__icontains=q)
            | Q(chef_section__icontains=q)
            | Q(direction__nom__icontains=q)
            | Q(direction__sigle__icontains=q)
            | Q(division__nom__icontains=q)
            | Q(division__sigle__icontains=q)
            | Q(division__chef_division__icontains=q)
            | Q(personnels__nom_prenoms__icontains=q)
            | Q(personnels__fonction__icontains=q)
        ).distinct()

    conf = ConfigurationMairie.objects.filter(est_active=True).first()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="organigramme_mairie.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=16,
    )

    story.append(Paragraph("Organigramme de la Mairie", title_style))

    if q:
        story.append(
            Paragraph(
                f"Filtre de recherche : {q}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=8.7,
        leading=10,
    )

    def p(text: str) -> Paragraph:
        return Paragraph((text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), cell_style)

    data = [
        [
            "Direction",
            "Division",
            "Chef de direction",
            "Section",
            "Chef de section",
            "Personnel",
            "Fonction",
        ]
    ]

    # Lignes : 1 ligne par personnel (et une ligne vide si section sans personnel)
    # Limite de sécurité pour éviter des PDFs trop lourds
    max_rows = 5000
    rows_added = 0

    for s in sections:
        direction_label = s.direction.sigle or s.direction.nom
        chef_direction = getattr(s.direction, "chef_direction", "") or ""

        division_label = ""
        if getattr(s, "division", None):
            division_label = s.division.sigle or s.division.nom

        section_label = s.nom
        if s.sigle:
            section_label += f" ({s.sigle})"

        chef_section = s.chef_section or ""

        personnels_qs = s.personnels.all().order_by("nom_prenoms")
        if personnels_qs.exists():
            for pers in personnels_qs:
                if rows_added >= max_rows:
                    break
                data.append(
                    [
                        p(direction_label),
                        p(division_label),
                        p(chef_direction),
                        p(section_label),
                        p(chef_section),
                        p(pers.nom_prenoms),
                        p(pers.fonction),
                    ]
                )
                rows_added += 1
        else:
            if rows_added >= max_rows:
                break
            data.append(
                [
                    p(direction_label),
                    p(division_label),
                    p(chef_direction),
                    p(section_label),
                    p(chef_section),
                    p(""),
                    p(""),
                ]
            )
            rows_added += 1

        if rows_added >= max_rows:
            break

    if len(data) == 1:
        story.append(
            Paragraph(
                "Aucune direction / division / section ne correspond aux critères sélectionnés.",
                styles["Normal"],
            )
        )
    else:
        col_widths = [
            3.5 * cm,  # direction
            3.5 * cm,  # division
            4.0 * cm,  # chef direction
            4.3 * cm,  # section
            3.3 * cm,  # chef section
            4.3 * cm,  # personnel
            3.8 * cm,  # fonction
        ]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.7),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

        if rows_added >= max_rows:
            story.append(Spacer(1, 0.25 * cm))
            story.append(
                Paragraph(
                    f"Note : export limité à {max_rows} lignes pour éviter un document trop lourd.",
                    styles["Normal"],
                )
            )

    story.append(Spacer(1, 0.5 * cm))
    story.append(
        Paragraph(
            "Document généré automatiquement depuis le tableau de bord de la Mairie de Kloto 1.",
            styles["Normal"],
        )
    )

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_organigramme(request):
    """
    Export Excel de l'organigramme (vue sections).
    """
    q = request.GET.get("q", "").strip()

    sections = (
        SectionDirection.objects.select_related("direction", "division")
        .prefetch_related("personnels", "services")
        .order_by(
            "direction__ordre_affichage",
            "division__ordre_affichage",
            "ordre_affichage",
            "nom",
        )
    )

    if q:
        sections = sections.filter(
            Q(nom__icontains=q)
            | Q(sigle__icontains=q)
            | Q(chef_section__icontains=q)
            | Q(direction__nom__icontains=q)
            | Q(direction__sigle__icontains=q)
            | Q(division__nom__icontains=q)
            | Q(division__sigle__icontains=q)
            | Q(division__chef_division__icontains=q)
            | Q(personnels__nom_prenoms__icontains=q)
            | Q(personnels__fonction__icontains=q)
            | Q(services__titre__icontains=q)
        ).distinct()

    wb = Workbook()

    # Feuille 1 : synthèse par section (avec chef de direction)
    ws_sections = wb.active
    ws_sections.title = "Sections"

    headers_sections = [
        "ID Section",
        "Direction",
        "Sigle direction",
        "Division",
        "Sigle division",
        "Chef de division",
        "Section",
        "Sigle section",
        "Chef de section",
        "Nombre de personnel",
        "Liste du personnel (résumé)",
        "Services (résumé)",
    ]
    ws_sections.append(headers_sections)

    for s in sections:
        personnels_qs = s.personnels.all()
        services_qs = getattr(s, "services", None)

        personnels_labels = ", ".join(
            personnels_qs.values_list("nom_prenoms", flat=True)[:10]
        )
        if personnels_qs.count() > 10:
            personnels_labels += "…"

        services_titles = ""
        if services_qs is not None:
            services_titles = ", ".join(
                services_qs.values_list("titre", flat=True)[:10]
            )
            if services_qs.count() > 10:
                services_titles += "…"

        ws_sections.append(
            [
                s.id,
                s.direction.nom,
                s.direction.sigle or "",
                (s.division.nom if s.division else ""),
                (s.division.sigle if s.division and s.division.sigle else ""),
                (getattr(s.division, "chef_division", "") if s.division else ""),
                s.nom,
                s.sigle or "",
                s.chef_section or "",
                personnels_qs.count(),
                personnels_labels,
                services_titles,
            ]
        )

    # Feuille 2 : directions (avec effectif global)
    ws_dirs = wb.create_sheet(title="Directions")
    headers_dirs = [
        "ID Direction",
        "Nom direction",
        "Sigle",
        "Chef de direction",
        "Nombre de sections",
        "Nombre de personnel",
    ]
    ws_dirs.append(headers_dirs)

    # Regrouper les sections par direction
    directions_map = {}
    for s in sections:
        d = s.direction
        entry = directions_map.setdefault(
            d.pk,
            {
                "direction": d,
                "sections": [],
                "personnels_count": 0,
            },
        )
        entry["sections"].append(s)
        entry["personnels_count"] += s.personnels.count()

    for entry in sorted(
        directions_map.values(),
        key=lambda item: (item["direction"].ordre_affichage, item["direction"].nom),
    ):
        d = entry["direction"]
        ws_dirs.append(
            [
                d.id,
                d.nom,
                d.sigle or "",
                getattr(d, "chef_direction", "") or "",
                len(entry["sections"]),
                entry["personnels_count"],
            ]
        )

    # Feuille 3 : personnel détaillé
    ws_personnel = wb.create_sheet(title="Personnel")
    headers_personnel = [
        "ID Personnel",
        "Nom et prénoms",
        "Fonction",
        "Section",
        "Division",
        "Direction",
        "Chef de direction",
        "Contact",
        "Adresse",
        "Actif",
    ]
    ws_personnel.append(headers_personnel)

    from mairie.models import PersonnelSection  # import local pour éviter les cycles

    personnels = (
        PersonnelSection.objects.select_related("section", "section__direction", "section__division")
        .filter(section__in=sections)
        .order_by(
            "section__direction__ordre_affichage",
            "section__division__ordre_affichage",
            "section__ordre_affichage",
            "nom_prenoms",
        )
    )

    for p in personnels:
        section = p.section
        direction = section.direction
        division = getattr(section, "division", None)
        ws_personnel.append(
            [
                p.id,
                p.nom_prenoms,
                p.fonction,
                section.nom,
                (division.nom if division else ""),
                direction.nom,
                getattr(direction, "chef_direction", "") or "",
                p.contact,
                p.adresse,
                "Oui" if p.est_actif else "Non",
            ]
        )

    # Ajustement simple de la largeur des colonnes pour chaque feuille
    for ws in [ws_sections, ws_dirs, ws_personnel]:
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    continue
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = 'attachment; filename="organigramme_mairie.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def gestion_publicites(request):
    """Gestion des campagnes publicitaires (validation, paiement, activation)."""

    # Filtrage par statut optionnel
    statut = request.GET.get("statut", "")

    campagnes = CampagnePublicitaire.objects.select_related("proprietaire").order_by(
        "-date_demande"
    )
    if statut:
        campagnes = campagnes.filter(statut=statut)

    if request.method == "POST":
        campagne_id = request.POST.get("campagne_id")
        action = request.POST.get("action")
        campagne = get_object_or_404(CampagnePublicitaire, pk=campagne_id)

        if action == "set_statut":
            nouveau_statut = request.POST.get("nouveau_statut")
            if nouveau_statut in dict(CampagnePublicitaire.STATUT_CHOICES):
                ancien_statut = campagne.statut
                campagne.statut = nouveau_statut
                campagne.save()

                # Notification automatique au demandeur quand la mairie accepte la demande
                if ancien_statut != "acceptee" and nouveau_statut == "acceptee":
                    montant_txt = (
                        f"{campagne.montant:,.0f} FCFA".replace(",", " ")
                        if campagne.montant is not None
                        else "le montant indiqué par la mairie"
                    )
                    Notification.objects.create(
                        recipient=campagne.proprietaire,
                        title="Votre campagne publicitaire a été acceptée",
                        message=(
                            f"Bonjour,\n\n"
                            f"Votre demande de campagne publicitaire « {campagne.titre} » a été acceptée par la mairie.\n\n"
                            f"Prochaine étape : paiement des frais de publicité ({montant_txt}).\n"
                            f"Merci de vous rapprocher du service compétent de la mairie pour effectuer le paiement et "
                            f"faire enregistrer votre règlement.\n\n"
                            f"Après enregistrement du paiement, vous pourrez créer vos publicités depuis votre compte."
                        ),
                        type=Notification.TYPE_INFO,
                        created_by=request.user,
                    )
                messages.success(
                    request,
                    f"Le statut de la campagne « {campagne.titre} » a été mis à jour ({campagne.get_statut_display()}).",
                )
        elif action == "update_montant":
            montant = request.POST.get("montant") or ""
            try:
                campagne.montant = float(montant.replace(",", ".") or 0)
                campagne.save()
                messages.success(
                    request,
                    f"Le montant de la campagne « {campagne.titre} » a été mis à jour.",
                )
            except ValueError:
                messages.error(request, "Montant invalide.")
        elif action == "set_dates":
            date_debut = request.POST.get("date_debut") or ""
            date_fin = request.POST.get("date_fin") or ""
            try:
                campagne.date_debut = (
                    datetime.strptime(date_debut, "%Y-%m-%dT%H:%M")
                    if date_debut
                    else None
                )
                campagne.date_fin = (
                    datetime.strptime(date_fin, "%Y-%m-%dT%H:%M") if date_fin else None
                )
                campagne.save()
                messages.success(
                    request,
                    f"Les dates de diffusion de la campagne « {campagne.titre} » ont été mises à jour.",
                )
            except ValueError:
                messages.error(request, "Format de date invalide.")
        elif action == "renew_campaign":
            # Renouveler une campagne terminée : on crée une nouvelle période à partir
            # de maintenant (ou de l'ancienne date de fin si elle est dans le futur)
            maintenant = timezone.now()
            point_depart = campagne.date_fin or maintenant
            if point_depart < maintenant:
                point_depart = maintenant

            # Utilise la durée de la campagne pour recalculer la nouvelle date de fin
            duree = campagne.duree_jours or 30
            campagne.date_debut = point_depart
            campagne.date_fin = point_depart + timedelta(days=duree)
            campagne.statut = "active"
            campagne.save()
            messages.success(
                request,
                (
                    f"La campagne « {campagne.titre} » a été renouvelée. "
                    f"Nouvelle période du {campagne.date_debut.strftime('%d/%m/%Y %H:%M')} "
                    f"au {campagne.date_fin.strftime('%d/%m/%Y %H:%M')}."
                ),
            )

        return redirect("gestion_publicites")

    # Pré-calcul du nombre de publicités par campagne
    campagnes = campagnes.annotate(nb_publicites=Count("publicites"))

    context = {
        "campagnes": campagnes,
        "statut": statut,
    }
    return render(request, "admin/gestion_publicites.html", context)


@login_required
@user_passes_test(is_staff_user)
def detail_campagne_publicite(request, pk: int):
    """Détail d'une demande/campagne publicitaire (vue admin tableau de bord)."""

    campagne = get_object_or_404(
        CampagnePublicitaire.objects.select_related("proprietaire"), pk=pk
    )
    publicites = Publicite.objects.filter(campagne=campagne).order_by("-date_creation")

    # Envoi optionnel d'instructions de paiement personnalisées
    if request.method == "POST":
        titre = request.POST.get("titre", "").strip() or "Paiement de votre campagne publicitaire"
        message_libre = request.POST.get("message", "").strip()
        moyens = []
        if request.POST.get("tmoney"):
            moyens.append("Tmoney")
        if request.POST.get("flooz"):
            moyens.append("Flooz")
        if request.POST.get("carte"):
            moyens.append("Carte bancaire")

        if not message_libre:
            messages.error(request, "Le message d'explication du paiement est obligatoire.")
        else:
            lignes = [
                "Bonjour,",
                "",
                f"Votre campagne publicitaire « {campagne.titre} » est en cours de traitement par la mairie.",
                "",
            ]
            if moyens:
                lignes.append(
                    "Vous pouvez effectuer le paiement de vos frais de publicité via : "
                    + ", ".join(moyens)
                    + "."
                )
                lignes.append("")
            lignes.append(message_libre)
            lignes.append("")
            lignes.append(
                "Après validation de votre paiement par la mairie, vous pourrez créer vos publicités "
                "depuis votre compte sur la plateforme."
            )

            Notification.objects.create(
                recipient=campagne.proprietaire,
                title=titre,
                message="\n".join(lignes),
                type=Notification.TYPE_INFO,
                created_by=request.user,
            )
            messages.success(
                request,
                "Les instructions de paiement ont été envoyées au demandeur dans son espace personnel.",
            )
            return redirect("detail_campagne_publicite", pk=campagne.pk)

    context = {
        "campagne": campagne,
        "publicites": publicites,
    }
    return render(request, "admin/detail_campagne_publicite.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_acteurs_economiques(request):
    """Liste des acteurs économiques enregistrés."""
    
    acteurs = ActeurEconomique.objects.all().order_by('-date_enregistrement')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    type_acteur = request.GET.get('type', '')
    secteur = request.GET.get('secteur', '')
    
    # Application des filtres
    if q:
        acteurs = acteurs.filter(
            Q(raison_sociale__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    
    if type_acteur:
        acteurs = acteurs.filter(type_acteur=type_acteur)
        
    if secteur:
        acteurs = acteurs.filter(secteur_activite=secteur)
    
    context = {
        'acteurs': acteurs,
        'titre': 'Acteurs Économiques',
        'type_choices': ActeurEconomique.TYPE_ACTEUR_CHOICES,
        'secteur_choices': ActeurEconomique.SECTEUR_ACTIVITE_CHOICES,
        'current_filters': {
            'q': q,
            'type': type_acteur,
            'secteur': secteur
        }
    }
    
    return render(request, "admin/liste_inscriptions.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_institutions_financieres(request):
    """Liste des institutions financières enregistrées."""
    
    institutions = InstitutionFinanciere.objects.all().order_by('-date_enregistrement')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    type_inst = request.GET.get('type', '')
    
    # Application des filtres
    if q:
        institutions = institutions.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    
    if type_inst:
        institutions = institutions.filter(type_institution=type_inst)
    
    context = {
        'institutions': institutions,
        'titre': 'Institutions Financières',
        'type_choices': InstitutionFinanciere.TYPE_INSTITUTION_CHOICES,
        'current_filters': {
            'q': q,
            'type': type_inst
        }
    }
    
    return render(request, "admin/liste_institutions.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_jeunes(request):
    """Liste des jeunes demandeurs d'emploi."""
    
    jeunes = ProfilEmploi.objects.filter(type_profil='jeune').order_by('-date_inscription')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    # Application des filtres
    if q:
        jeunes = jeunes.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
        
    if niveau:
        jeunes = jeunes.filter(niveau_etude=niveau)
        
    if dispo:
        jeunes = jeunes.filter(disponibilite=dispo)
    
    context = {
        'profils': jeunes,
        'titre': 'Jeunes Demandeurs d\'Emploi',
        'type_profil': 'jeune',
        'niveau_choices': ProfilEmploi.NIVEAU_ETUDE_CHOICES,
        'dispo_choices': ProfilEmploi.DISPONIBILITE_CHOICES,
        'current_filters': {
            'q': q,
            'niveau': niveau,
            'dispo': dispo
        }
    }
    
    return render(request, "admin/liste_profils_emploi.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_retraites(request):
    """Liste des retraités actifs."""
    
    retraites = ProfilEmploi.objects.filter(type_profil='retraite').order_by('-date_inscription')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    # Application des filtres
    if q:
        retraites = retraites.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
        
    if niveau:
        retraites = retraites.filter(niveau_etude=niveau)
        
    if dispo:
        retraites = retraites.filter(disponibilite=dispo)
    
    context = {
        'profils': retraites,
        'titre': 'Retraités Actifs',
        'type_profil': 'retraite',
        'niveau_choices': ProfilEmploi.NIVEAU_ETUDE_CHOICES,
        'dispo_choices': ProfilEmploi.DISPONIBILITE_CHOICES,
        'current_filters': {
            'q': q,
            'niveau': niveau,
            'dispo': dispo
        }
    }
    
    return render(request, "admin/liste_profils_emploi.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_diaspora_tableau_bord(request):
    """Liste des membres de la diaspora pour le tableau de bord."""
    
    membres = MembreDiaspora.objects.all().order_by('-date_inscription')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    pays = request.GET.get('pays', '')
    secteur = request.GET.get('secteur', '')
    
    # Application des filtres
    if q:
        membres = membres.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_whatsapp__icontains=q) |
            Q(profession_actuelle__icontains=q) |
            Q(domaine_formation__icontains=q)
        )
        
    if pays:
        membres = membres.filter(pays_residence_actuelle__icontains=pays)
        
    if secteur:
        membres = membres.filter(secteur_activite=secteur)
    
    context = {
        'membres': membres,
        'titre': '🌍 Membres de la Diaspora',
        'secteur_choices': MembreDiaspora.SECTEUR_ACTIVITE_CHOICES,
        'current_filters': {
            'q': q,
            'pays': pays,
            'secteur': secteur
        }
    }
    
    return render(request, "admin/liste_diaspora.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_osc_tableau_bord(request):
    """Liste des Organisations de la Société Civile (OSC) pour le tableau de bord."""

    osc_qs = OrganisationSocieteCivile.objects.all().order_by("-date_enregistrement")

    # Filtres simples (recherche texte + type d'OSC)
    q = request.GET.get("q", "") or ""
    type_osc = request.GET.get("type", "") or ""

    if q:
        osc_qs = osc_qs.filter(
            Q(nom_osc__icontains=q)
            | Q(sigle__icontains=q)
            | Q(email__icontains=q)
            | Q(telephone__icontains=q)
        )

    if type_osc:
        osc_qs = osc_qs.filter(type_osc=type_osc)

    # Choix pour le filtre et le PDF : (valeur, libellé) depuis la liste d'inscription
    type_choices = [(v, l) for v, l in OSC_TYPE_CHOICES if v]

    context = {
        "osc_list": osc_qs,
        "titre": "🤝 Organisations de la Société Civile (OSC)",
        "current_filters": {
            "q": q,
            "type": type_osc,
        },
        "type_choices": type_choices,
    }

    return render(request, "admin/liste_osc.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_agents_collecteurs(request):
    """Liste des agents collecteurs de taxes."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    from urllib.parse import urlencode
    from django.urls import reverse
    
    # POST : modifier un agent depuis la modale (y compris ses affectations)
    if request.method == "POST" and request.POST.get("action") == "modifier_agent":
        agent_id = request.POST.get("agent_id")
        matricule = (request.POST.get("matricule") or "").strip()
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        telephone = (request.POST.get("telephone") or "").strip()
        email = (request.POST.get("email") or "").strip()
        statut = request.POST.get("statut") or "actif"
        date_embauche_str = request.POST.get("date_embauche") or ""
        notes = (request.POST.get("notes") or "").strip()
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        emplacement_ids = request.POST.getlist("emplacements")
        acteurs_ids = request.POST.getlist("acteurs")
        institutions_ids = request.POST.getlist("institutions")

        date_embauche = None
        if date_embauche_str:
            try:
                from datetime import datetime as dt
                date_embauche = dt.strptime(date_embauche_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        erreurs = []
        try:
            agent = AgentCollecteur.objects.get(pk=agent_id)
        except (AgentCollecteur.DoesNotExist, ValueError, TypeError):
            agent = None
            erreurs.append("Agent introuvable.")

        if agent:
            if not matricule:
                erreurs.append("Le matricule est obligatoire.")
            elif AgentCollecteur.objects.filter(matricule=matricule).exclude(pk=agent.pk).exists():
                erreurs.append("Un autre agent avec ce matricule existe déjà.")
            if not nom or not prenom:
                erreurs.append("Le nom et le prénom sont obligatoires.")
            if not telephone:
                erreurs.append("Le téléphone est obligatoire.")
            if not username:
                erreurs.append("L'identifiant de connexion (username) est obligatoire.")
            if User.objects.filter(username=username).exclude(pk=agent.user_id).exists():
                erreurs.append("Ce nom d'utilisateur est déjà utilisé.")
            if password and len(password) < 6:
                erreurs.append("Le mot de passe doit contenir au moins 6 caractères.")

            if erreurs:
                for e in erreurs:
                    messages.error(request, e)
            else:
                agent.matricule = matricule
                agent.nom = nom
                agent.prenom = prenom
                agent.telephone = telephone
                agent.email = email or ""
                agent.statut = statut
                agent.notes = notes
                agent.date_embauche = date_embauche
                agent.save()

                # Emplacements (marchés / places) assignés
                if emplacement_ids:
                    agent.emplacements_assignes.set(
                        EmplacementMarche.objects.filter(pk__in=emplacement_ids)
                    )
                else:
                    agent.emplacements_assignes.clear()

                # Acteurs économiques assignés à cet agent
                if acteurs_ids:
                    agent.acteurs_economiques.set(
                        ActeurEconomique.objects.filter(pk__in=acteurs_ids)
                    )
                else:
                    agent.acteurs_economiques.clear()

                # Institutions financières assignées à cet agent
                if institutions_ids:
                    agent.institutions_financieres.set(
                        InstitutionFinanciere.objects.filter(pk__in=institutions_ids)
                    )
                else:
                    agent.institutions_financieres.clear()

                user = agent.user
                user.username = username
                if email:
                    user.email = email
                if password:
                    user.set_password(password)
                user.save()

                messages.success(request, f"Agent {agent.matricule} - {agent.nom_complet} a été modifié.")
        
        url = reverse("liste_agents_collecteurs")
        params = {k: v for k, v in request.GET.items()}
        if params:
            url += "?" + urlencode(params)
        return redirect(url)
    
    agents = (
        AgentCollecteur.objects.select_related("user")
        .prefetch_related("emplacements_assignes", "acteurs_economiques", "institutions_financieres")
        .order_by("-date_creation")
    )
    
    # Filtres
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    if q:
        agents = agents.filter(
            Q(matricule__icontains=q) |
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q) |
            Q(email__icontains=q)
        )
    
    if statut:
        agents = agents.filter(statut=statut)
    
    emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
    acteurs = ActeurEconomique.objects.filter(est_valide_par_mairie=True).order_by("raison_sociale")
    institutions = InstitutionFinanciere.objects.filter(est_valide_par_mairie=True).order_by("nom_institution")

    context = {
        'agents': agents,
        'titre': '👮 Agents Collecteurs',
        'statut_choices': AgentCollecteur.STATUT_CHOICES,
        'emplacements': emplacements,
        'acteurs': acteurs,
        'institutions': institutions,
        'current_filters': {
            'q': q,
            'statut': statut
        }
    }

    return render(request, "admin/liste_agents_collecteurs.html", context)


@login_required
@user_passes_test(is_staff_user)
def ajouter_agent_collecteur(request):
    """Formulaire d'ajout d'un agent collecteur (création User + AgentCollecteur)."""
    from django.contrib.auth import get_user_model
    User = get_user_model()

    if request.method == "POST":
        matricule = (request.POST.get("matricule") or "").strip()
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        telephone = (request.POST.get("telephone") or "").strip()
        email = (request.POST.get("email") or "").strip()
        statut = request.POST.get("statut") or "actif"
        date_embauche_str = request.POST.get("date_embauche") or ""
        notes = (request.POST.get("notes") or "").strip()
        username = (request.POST.get("username") or matricule or "").strip()
        password = request.POST.get("password") or ""
        emplacement_ids = request.POST.getlist("emplacements")

        date_embauche = None
        if date_embauche_str:
            try:
                from datetime import datetime as dt
                date_embauche = dt.strptime(date_embauche_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        erreurs = []
        if not matricule:
            erreurs.append("Le matricule est obligatoire.")
        if AgentCollecteur.objects.filter(matricule=matricule).exists():
            erreurs.append("Un agent avec ce matricule existe déjà.")
        if not nom or not prenom:
            erreurs.append("Le nom et le prénom sont obligatoires.")
        if not telephone:
            erreurs.append("Le téléphone est obligatoire.")
        if not username:
            erreurs.append("L'identifiant de connexion (username) est obligatoire.")
        if User.objects.filter(username=username).exists():
            erreurs.append("Ce nom d'utilisateur existe déjà.")
        if not password or len(password) < 6:
            erreurs.append("Le mot de passe doit contenir au moins 6 caractères.")

        if erreurs:
            for e in erreurs:
                messages.error(request, e)
            emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
            post = {k: request.POST.get(k) for k in ["matricule", "nom", "prenom", "telephone", "email", "statut", "date_embauche", "notes", "username"]}
            post["emplacements"] = [int(x) for x in request.POST.getlist("emplacements") if x.isdigit()]
            return render(request, "admin/ajouter_agent_collecteur.html", {
                "statut_choices": AgentCollecteur.STATUT_CHOICES,
                "emplacements": emplacements,
                "post": post,
            })

        user = User.objects.create_user(
            username=username,
            password=password,
            email=email or username + "@mairie.local",
            is_staff=True,
        )
        agent = AgentCollecteur.objects.create(
            user=user,
            matricule=matricule,
            nom=nom,
            prenom=prenom,
            telephone=telephone,
            email=email or "",
            statut=statut,
            notes=notes,
            date_embauche=date_embauche,
        )
        if emplacement_ids:
            agent.emplacements_assignes.set(EmplacementMarche.objects.filter(pk__in=emplacement_ids))
        messages.success(request, f"Agent {agent.matricule} - {agent.nom_complet} a été créé.")
        return redirect("liste_agents_collecteurs")

    emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
    return render(request, "admin/ajouter_agent_collecteur.html", {
        "statut_choices": AgentCollecteur.STATUT_CHOICES,
        "emplacements": emplacements,
        "post": {"emplacements": []},
    })


@login_required
@user_passes_test(is_staff_user)
def modifier_agent_collecteur(request, agent_id):
    """Formulaire de modification d'un agent collecteur."""
    from django.contrib.auth import get_user_model
    User = get_user_model()

    agent = get_object_or_404(
        AgentCollecteur.objects.prefetch_related("emplacements_assignes"),
        pk=agent_id,
    )

    if request.method == "POST":
        matricule = (request.POST.get("matricule") or "").strip()
        nom = (request.POST.get("nom") or "").strip()
        prenom = (request.POST.get("prenom") or "").strip()
        telephone = (request.POST.get("telephone") or "").strip()
        email = (request.POST.get("email") or "").strip()
        statut = request.POST.get("statut") or "actif"
        date_embauche_str = request.POST.get("date_embauche") or ""
        notes = (request.POST.get("notes") or "").strip()
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        emplacement_ids = request.POST.getlist("emplacements")

        date_embauche = None
        if date_embauche_str:
            try:
                from datetime import datetime as dt
                date_embauche = dt.strptime(date_embauche_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass

        erreurs = []
        if not matricule:
            erreurs.append("Le matricule est obligatoire.")
        if AgentCollecteur.objects.filter(matricule=matricule).exclude(pk=agent.pk).exists():
            erreurs.append("Un autre agent avec ce matricule existe déjà.")
        if not nom or not prenom:
            erreurs.append("Le nom et le prénom sont obligatoires.")
        if not telephone:
            erreurs.append("Le téléphone est obligatoire.")
        if not username:
            erreurs.append("L'identifiant de connexion (username) est obligatoire.")
        if User.objects.filter(username=username).exclude(pk=agent.user_id).exists():
            erreurs.append("Ce nom d'utilisateur est déjà utilisé.")
        if password and len(password) < 6:
            erreurs.append("Le mot de passe doit contenir au moins 6 caractères.")

        if erreurs:
            for e in erreurs:
                messages.error(request, e)
            emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
            return render(request, "admin/modifier_agent_collecteur.html", {
                "agent": agent,
                "statut_choices": AgentCollecteur.STATUT_CHOICES,
                "emplacements": emplacements,
                "post": {
                    "matricule": matricule,
                    "nom": nom,
                    "prenom": prenom,
                    "telephone": telephone,
                    "email": email,
                    "statut": statut,
                    "date_embauche": date_embauche_str,
                    "notes": notes,
                    "username": username,
                    "emplacements": [int(x) for x in emplacement_ids if x.isdigit()],
                },
            })

        agent.matricule = matricule
        agent.nom = nom
        agent.prenom = prenom
        agent.telephone = telephone
        agent.email = email or ""
        agent.statut = statut
        agent.notes = notes
        agent.date_embauche = date_embauche
        agent.save()

        if emplacement_ids:
            agent.emplacements_assignes.set(EmplacementMarche.objects.filter(pk__in=emplacement_ids))
        else:
            agent.emplacements_assignes.clear()

        user = agent.user
        user.username = username
        if email:
            user.email = email
        if password:
            user.set_password(password)
        user.save()

        messages.success(request, f"Agent {agent.matricule} - {agent.nom_complet} a été modifié.")
        return redirect("liste_agents_collecteurs")

    emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
    post = {
        "matricule": agent.matricule,
        "nom": agent.nom,
        "prenom": agent.prenom,
        "telephone": agent.telephone,
        "email": agent.email or "",
        "statut": agent.statut,
        "date_embauche": agent.date_embauche.strftime("%Y-%m-%d") if agent.date_embauche else "",
        "notes": agent.notes or "",
        "username": agent.user.username,
        "emplacements": list(agent.emplacements_assignes.values_list("pk", flat=True)),
    }
    return render(request, "admin/modifier_agent_collecteur.html", {
        "agent": agent,
        "statut_choices": AgentCollecteur.STATUT_CHOICES,
        "emplacements": emplacements,
        "post": post,
    })


def _parse_date(s):
    """Parse une date au format YYYY-MM-DD. Retourne None si invalide."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


@login_required
@user_passes_test(is_staff_user)
def liste_contribuables(request):
    """Liste des contribuables (marchés et places publiques)."""
    
    contribuables = Contribuable.objects.select_related('user').prefetch_related('boutiques_magasins').order_by('-date_creation')
    
    # Filtres
    q = request.GET.get('q', '')
    nationalite = request.GET.get('nationalite', '')
    date_du = request.GET.get('date_du', '')
    date_au = request.GET.get('date_au', '')
    
    if q:
        contribuables = contribuables.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q)
        )
    
    if nationalite:
        contribuables = contribuables.filter(nationalite__icontains=nationalite)
    
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        contribuables = contribuables.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        contribuables = contribuables.filter(date_creation__date__lte=date_au_parsed)
    
    # Ajouter le nombre de boutiques pour chaque contribuable
    for contribuable in contribuables:
        contribuable.nombre_boutiques = contribuable.boutiques_magasins.count()
    
    context = {
        'contribuables': contribuables,
        'titre': '👥 Contribuables (Marchés / Places publiques)',
        'current_filters': {
            'q': q,
            'nationalite': nationalite,
            'date_du': date_du,
            'date_au': date_au,
        }
    }
    
    return render(request, "admin/liste_contribuables.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_boutiques(request):
    """Liste et création des boutiques / magasins de marché depuis le tableau de bord."""

    boutiques = BoutiqueMagasin.objects.select_related("contribuable", "emplacement", "agent_collecteur").order_by("-id")

    # Filtres
    q = request.GET.get("q", "")
    contribuable_id = request.GET.get("contribuable", "")
    agent_collecteur_id = request.GET.get("agent_collecteur", "")
    date_du = request.GET.get("date_du", "")
    date_au = request.GET.get("date_au", "")
    if q:
        boutiques = boutiques.filter(
            Q(matricule__icontains=q)
            | Q(contribuable__nom__icontains=q)
            | Q(contribuable__prenom__icontains=q)
            | Q(emplacement__nom_lieu__icontains=q)
        )
    if contribuable_id:
        try:
            boutiques = boutiques.filter(contribuable_id=int(contribuable_id))
        except (ValueError, TypeError):
            pass
    if agent_collecteur_id:
        try:
            boutiques = boutiques.filter(agent_collecteur_id=int(agent_collecteur_id))
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        boutiques = boutiques.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        boutiques = boutiques.filter(date_creation__date__lte=date_au_parsed)

    from urllib.parse import urlencode
    from django.urls import reverse

    # POST : soit créer un local (sans locataire), soit louer un local existant, soit modifier une boutique
    if request.method == "POST":
        action = request.POST.get("action", "louer")

        if action == "modifier_boutique":
            # Modifier une boutique existante
            boutique_id = request.POST.get("boutique_id")
            matricule = (request.POST.get("matricule") or "").strip()
            emplacement_id = request.POST.get("emplacement")
            type_local = request.POST.get("type_local") or "boutique"
            superficie = request.POST.get("superficie_m2") or ""
            loyer_mensuel = request.POST.get("prix_location_mensuel") or ""
            loyer_annuel = request.POST.get("prix_location_annuel") or ""
            contribuable_id_post = request.POST.get("contribuable")
            agent_collecteur_id_post = request.POST.get("agent_collecteur")
            activite = (request.POST.get("activite_vendue") or "").strip()
            description = (request.POST.get("description") or "").strip()
            est_actif = request.POST.get("est_actif") == "1"
            
            erreurs = []
            try:
                boutique = BoutiqueMagasin.objects.get(pk=boutique_id)
            except (BoutiqueMagasin.DoesNotExist, ValueError, TypeError):
                boutique = None
                erreurs.append("Boutique introuvable.")
            
            if not matricule:
                erreurs.append("La matricule du local est obligatoire.")
            elif boutique and BoutiqueMagasin.objects.filter(matricule=matricule).exclude(pk=boutique_id).exists():
                erreurs.append("Un local avec cette matricule existe déjà.")
            
            try:
                emplacement = EmplacementMarche.objects.get(pk=emplacement_id)
            except (EmplacementMarche.DoesNotExist, ValueError, TypeError):
                emplacement = None
                erreurs.append("Veuillez choisir un emplacement valide.")
            
            try:
                superficie_val = Decimal(str(superficie)) if superficie else Decimal("0")
            except (InvalidOperation, TypeError, ValueError):
                superficie_val = Decimal("0")
            
            try:
                loyer_mensuel_val = Decimal(str(loyer_mensuel)) if loyer_mensuel else Decimal("0")
            except (InvalidOperation, TypeError, ValueError):
                loyer_mensuel_val = Decimal("0")
            
            loyer_annuel_val = None
            if loyer_annuel:
                try:
                    loyer_annuel_val = Decimal(str(loyer_annuel))
                except (InvalidOperation, TypeError, ValueError):
                    pass
            
            contribuable = None
            if contribuable_id_post:
                try:
                    contribuable = Contribuable.objects.get(pk=contribuable_id_post)
                except (Contribuable.DoesNotExist, ValueError, TypeError):
                    erreurs.append("Contribuable sélectionné invalide.")
            
            agent_collecteur = None
            if agent_collecteur_id_post:
                try:
                    agent_collecteur = AgentCollecteur.objects.get(pk=agent_collecteur_id_post)
                except (AgentCollecteur.DoesNotExist, ValueError, TypeError):
                    erreurs.append("Agent collecteur sélectionné invalide.")
            
            if erreurs:
                for e in erreurs:
                    messages.error(request, e)
            elif boutique:
                boutique.matricule = matricule
                boutique.emplacement = emplacement
                boutique.type_local = type_local
                boutique.superficie_m2 = superficie_val
                boutique.prix_location_mensuel = loyer_mensuel_val
                boutique.prix_location_annuel = loyer_annuel_val
                boutique.contribuable = contribuable
                boutique.agent_collecteur = agent_collecteur
                boutique.activite_vendue = activite
                boutique.description = description
                boutique.est_actif = est_actif
                boutique.save()
                messages.success(request, f"Boutique {boutique.matricule} modifiée avec succès.")
                url = reverse("liste_boutiques")
                params = {k: v for k, v in request.GET.items()}
                if params:
                    url += "?" + urlencode(params)
                return redirect(url)

        elif action == "creer_local":
            # Créer un local (boutique, magasin, etc.) sans l’assigner à un locataire
            matricule = (request.POST.get("matricule") or "").strip()
            emplacement_id = request.POST.get("emplacement")
            type_local = request.POST.get("type_local") or "boutique"
            superficie = request.POST.get("superficie_m2") or ""
            loyer_mensuel = request.POST.get("prix_location_mensuel") or ""
            loyer_annuel = request.POST.get("prix_location_annuel") or ""
            description = (request.POST.get("description") or "").strip()
            erreurs = []
            if not matricule:
                erreurs.append("La matricule du local est obligatoire.")
            elif BoutiqueMagasin.objects.filter(matricule=matricule).exists():
                erreurs.append("Un local avec cette matricule existe déjà.")
            try:
                emplacement = EmplacementMarche.objects.get(pk=emplacement_id)
            except (EmplacementMarche.DoesNotExist, ValueError, TypeError):
                emplacement = None
                erreurs.append("Veuillez choisir un emplacement valide.")
            try:
                superficie_val = Decimal(str(superficie)) if superficie else Decimal("0")
            except (InvalidOperation, TypeError, ValueError):
                superficie_val = Decimal("0")
            try:
                loyer_mensuel_val = Decimal(str(loyer_mensuel)) if loyer_mensuel else Decimal("0")
            except (InvalidOperation, TypeError, ValueError):
                loyer_mensuel_val = Decimal("0")
            loyer_annuel_val = None
            if loyer_annuel:
                try:
                    loyer_annuel_val = Decimal(str(loyer_annuel))
                except (InvalidOperation, TypeError, ValueError):
                    pass
            if erreurs:
                for e in erreurs:
                    messages.error(request, e)
            else:
                BoutiqueMagasin.objects.create(
                    matricule=matricule,
                    emplacement=emplacement,
                    type_local=type_local,
                    superficie_m2=superficie_val,
                    prix_location_mensuel=loyer_mensuel_val,
                    prix_location_annuel=loyer_annuel_val,
                    contribuable=None,
                    activite_vendue="",
                    description=description,
                    est_actif=True,
                )
                messages.success(request, f"Local {matricule} créé (non occupé).")
                url = reverse("liste_boutiques")
                params = {k: v for k, v in request.GET.items()}
                if params:
                    url += "?" + urlencode(params)
                return redirect(url)

        else:
            # Louer : assigner un local non occupé à un contribuable
            local_id = request.POST.get("local_id")
            contribuable_id_post = request.POST.get("contribuable")
            agent_collecteur_id_post = request.POST.get("agent_collecteur")
            activite = (request.POST.get("activite_vendue") or "").strip()
            erreurs = []
            try:
                local = BoutiqueMagasin.objects.get(pk=local_id, contribuable__isnull=True)
            except (BoutiqueMagasin.DoesNotExist, ValueError, TypeError):
                local = None
                erreurs.append("Veuillez choisir un local non occupé valide.")
            try:
                contribuable = Contribuable.objects.get(pk=contribuable_id_post)
            except (Contribuable.DoesNotExist, ValueError, TypeError):
                contribuable = None
                erreurs.append("Veuillez choisir un contribuable (locataire) valide.")
            if not activite:
                erreurs.append("Veuillez renseigner l'activité exercée dans ce local.")
            if not agent_collecteur_id_post:
                erreurs.append("Veuillez choisir un agent collecteur assigné à ce local.")
            agent_collecteur = None
            if agent_collecteur_id_post:
                try:
                    agent_collecteur = AgentCollecteur.objects.get(pk=agent_collecteur_id_post)
                except (AgentCollecteur.DoesNotExist, ValueError, TypeError):
                    erreurs.append("Agent collecteur sélectionné invalide.")
            if erreurs:
                for e in erreurs:
                    messages.error(request, e)
            else:
                local.contribuable = contribuable
                local.activite_vendue = activite
                local.agent_collecteur = agent_collecteur
                local.save()
                message = f"Local {local.matricule} loué à {contribuable.nom} {contribuable.prenom}."
                if agent_collecteur:
                    message += f" Agent collecteur assigné : {agent_collecteur.nom} {agent_collecteur.prenom}."
                messages.success(request, message)
                url = reverse("liste_boutiques")
                params = {k: v for k, v in request.GET.items()}
                if params:
                    url += "?" + urlencode(params)
                return redirect(url)

    # Données pour les formulaires
    emplacements = EmplacementMarche.objects.all().order_by("nom_lieu")
    contribuables_all = Contribuable.objects.all().order_by("nom", "prenom")
    agents_collecteurs = AgentCollecteur.objects.filter(statut="actif").order_by("matricule", "nom", "prenom")
    type_local_choices = BoutiqueMagasin.TYPE_LOCAL_CHOICES
    locaux_non_occupes = BoutiqueMagasin.objects.filter(contribuable__isnull=True).select_related(
        "emplacement"
    ).order_by("emplacement__nom_lieu", "matricule")

    context = {
        "boutiques": boutiques,
        "titre": "🏪 Boutiques / Magasins (marchés)",
        "current_filters": {
            "q": q,
            "contribuable": contribuable_id,
            "agent_collecteur": agent_collecteur_id,
            "date_du": date_du,
            "date_au": date_au,
        },
        "emplacements": emplacements,
        "contribuables_all": contribuables_all,
        "agents_collecteurs": agents_collecteurs,
        "type_local_choices": type_local_choices,
        "locaux_non_occupes": locaux_non_occupes,
    }

    return render(request, "admin/liste_boutiques.html", context)


@login_required
@user_passes_test(is_staff_user)
@require_POST
def creer_emplacement_ajax(request):
    """Crée un emplacement (marché/place publique) via AJAX. Retourne JSON avec id et label."""
    from django.http import JsonResponse
    nom_lieu = (request.POST.get("nom_lieu") or "").strip()
    quartier = (request.POST.get("quartier") or "").strip()
    canton = (request.POST.get("canton") or "").strip()
    village = (request.POST.get("village") or "").strip()
    description = (request.POST.get("description") or "").strip()
    if not nom_lieu or not quartier:
        return JsonResponse({"success": False, "error": "Le nom du lieu et le quartier sont obligatoires."}, status=400)
    emplacement = EmplacementMarche.objects.create(
        nom_lieu=nom_lieu,
        quartier=quartier,
        canton=canton,
        village=village,
        description=description,
    )
    label = f"{emplacement.nom_lieu} - {emplacement.quartier}"
    return JsonResponse({"success": True, "id": emplacement.id, "label": label})


@login_required
@user_passes_test(is_staff_user)
@require_http_methods(["POST"])
def sauvegarder_infrastructure_ajax(request):
    """
    Crée ou met à jour une InfrastructureCommune via AJAX.
    Si `infra_id` est fourni, met à jour l'infrastructure existante, sinon en crée une nouvelle.
    """
    infra_id = request.POST.get("infra_id") or ""
    type_infra = (request.POST.get("type_infrastructure") or "").strip()
    nom = (request.POST.get("nom") or "").strip()
    description = (request.POST.get("description") or "").strip()
    adresse = (request.POST.get("adresse") or "").strip()
    lat = (request.POST.get("latitude") or "").strip()
    lng = (request.POST.get("longitude") or "").strip()
    est_active_raw = (request.POST.get("est_active") or "").strip().lower()

    if not type_infra or not nom or not lat or not lng:
        return JsonResponse(
            {
                "success": False,
                "error": "Le type, le nom et les coordonnées GPS (latitude, longitude) sont obligatoires.",
            },
            status=400,
        )

    valid_types = {choice[0] for choice in InfrastructureCommune.TYPE_INFRASTRUCTURE_CHOICES}
    if type_infra not in valid_types:
        return JsonResponse(
            {
                "success": False,
                "error": "Type d'infrastructure invalide.",
            },
            status=400,
        )

    try:
        lat_dec = Decimal(lat)
        lng_dec = Decimal(lng)
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse(
            {"success": False, "error": "Latitude ou longitude invalide."},
            status=400,
        )

    config = ConfigurationMairie.objects.filter(est_active=True).first()
    if not config:
        return JsonResponse(
            {
                "success": False,
                "error": "Aucune configuration de commune active trouvée.",
            },
            status=400,
        )

    try:
        cartographie = CartographieCommune.objects.get(configuration=config)
    except CartographieCommune.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Aucune fiche de cartographie trouvée pour la commune active.",
            },
            status=400,
        )

    est_active = est_active_raw in ("1", "true", "on", "oui", "yes")

    if infra_id:
        try:
            infra_obj = InfrastructureCommune.objects.get(id=int(infra_id), cartographie=cartographie)
        except (ValueError, InfrastructureCommune.DoesNotExist):
            return JsonResponse(
                {"success": False, "error": "Infrastructure introuvable."},
                status=404,
            )
        infra_obj.type_infrastructure = type_infra
        infra_obj.nom = nom
        infra_obj.description = description
        infra_obj.adresse = adresse
        infra_obj.latitude = lat_dec
        infra_obj.longitude = lng_dec
        infra_obj.est_active = est_active
        infra_obj.save()
    else:
        infra_obj = InfrastructureCommune.objects.create(
            cartographie=cartographie,
            type_infrastructure=type_infra,
            nom=nom,
            description=description,
            adresse=adresse,
            latitude=lat_dec,
            longitude=lng_dec,
            est_active=est_active,
        )

    return JsonResponse(
        {
            "success": True,
            "id": infra_obj.id,
            "type_infrastructure": infra_obj.type_infrastructure,
            "nom": infra_obj.nom,
            "description": infra_obj.description,
            "adresse": infra_obj.adresse,
            "latitude": str(infra_obj.latitude),
            "longitude": str(infra_obj.longitude),
            "est_active": infra_obj.est_active,
        }
    )


@login_required
@user_passes_test(is_staff_user)
def liste_contributions(request):
    """Liste des contributions/taxes (cotisations annuelles, paiements mensuels, tickets marché)."""
    
    # Récupération des paramètres de filtrage
    type_contribution = request.GET.get('type', '')
    annee = request.GET.get('annee', '')
    mois = request.GET.get('mois', '')
    agent_collecteur_id = request.GET.get('agent_collecteur', '')
    date_du = request.GET.get('date_du', '')
    date_au = request.GET.get('date_au', '')
    q = request.GET.get('q', '')
    
    cotisations_annuelles = CotisationAnnuelle.objects.select_related(
        'boutique__contribuable', 'boutique__emplacement'
    ).order_by('-annee', '-date_creation')
    
    paiements = PaiementCotisation.objects.select_related(
        'cotisation_annuelle__boutique__contribuable',
        'encaisse_par_agent'
    ).order_by('-date_paiement')
    
    tickets = TicketMarche.objects.select_related(
        'emplacement', 'contribuable', 'encaisse_par_agent'
    ).order_by('-date', '-date_creation')
    
    # Filtres par type
    if type_contribution == 'cotisations':
        paiements = paiements.none()
        tickets = tickets.none()
    elif type_contribution == 'paiements':
        cotisations_annuelles = cotisations_annuelles.none()
        tickets = tickets.none()
    elif type_contribution == 'tickets':
        cotisations_annuelles = cotisations_annuelles.none()
        paiements = paiements.none()
    
    # Filtre par année
    if annee:
        try:
            annee_int = int(annee)
            cotisations_annuelles = cotisations_annuelles.filter(annee=annee_int)
            paiements = paiements.filter(cotisation_annuelle__annee=annee_int)
            tickets = tickets.filter(date__year=annee_int)
        except ValueError:
            pass
    
    # Filtre par mois (paiements: mois 1-12, tickets: date__month)
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements = paiements.filter(mois=mois_int)
                tickets = tickets.filter(date__month=mois_int)
        except ValueError:
            pass
    
    # Filtre par agent collecteur (paiements et tickets)
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements = paiements.filter(encaisse_par_agent_id=agent_id)
            tickets = tickets.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    
    # Filtre par période (date_du, date_au) pour paiements et tickets
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements = paiements.filter(date_paiement__date__gte=date_du_parsed)
        tickets = tickets.filter(date__gte=date_du_parsed)
    if date_au_parsed:
        paiements = paiements.filter(date_paiement__date__lte=date_au_parsed)
        tickets = tickets.filter(date__lte=date_au_parsed)
    
    # Recherche textuelle
    if q:
        cotisations_annuelles = cotisations_annuelles.filter(
            Q(boutique__matricule__icontains=q) |
            Q(boutique__contribuable__nom__icontains=q) |
            Q(boutique__contribuable__prenom__icontains=q)
        )
        paiements = paiements.filter(
            Q(cotisation_annuelle__boutique__matricule__icontains=q) |
            Q(cotisation_annuelle__boutique__contribuable__nom__icontains=q) |
            Q(cotisation_annuelle__boutique__contribuable__prenom__icontains=q)
        )
        tickets = tickets.filter(
            Q(nom_vendeur__icontains=q) |
            Q(contribuable__nom__icontains=q) |
            Q(contribuable__prenom__icontains=q)
        )
    
    # Années disponibles pour le filtre
    annees_cotisations = sorted(
        CotisationAnnuelle.objects.values_list('annee', flat=True).distinct(),
        reverse=True
    )
    annees_tickets = sorted(
        TicketMarche.objects.values_list('date__year', flat=True).distinct(),
        reverse=True
    )
    annees_disponibles = sorted(set(annees_cotisations + annees_tickets), reverse=True)
    agents_collecteurs = AgentCollecteur.objects.filter(statut="actif").order_by("matricule", "nom", "prenom")
    
    context = {
        'cotisations_annuelles': cotisations_annuelles[:100],  # Limiter pour performance
        'paiements': paiements[:100],
        'tickets': tickets[:100],
        'titre': '💰 Contributions / Taxes',
        'annees_disponibles': annees_disponibles,
        'agents_collecteurs': agents_collecteurs,
        'current_filters': {
            'type': type_contribution,
            'annee': annee,
            'mois': mois,
            'agent_collecteur': agent_collecteur_id,
            'date_du': date_du,
            'date_au': date_au,
            'q': q
        }
    }
    
    return render(request, "admin/liste_contributions.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_cotisations_acteurs_institutions(request):
    """Liste des acteurs économiques, institutions financières et leurs cotisations annuelles."""
    
    # Récupération des paramètres de filtrage
    type_contribution = request.GET.get('type', '')  # 'acteurs' ou 'institutions' ou ''
    annee = request.GET.get('annee', '')
    mois = request.GET.get('mois', '')
    agent_collecteur_id = request.GET.get('agent_collecteur', '')
    date_du = request.GET.get('date_du', '')
    date_au = request.GET.get('date_au', '')
    q = request.GET.get('q', '')
    
    # Listes des acteurs et institutions (pour les tableaux principaux)
    acteurs_economiques = ActeurEconomique.objects.all().order_by('raison_sociale')
    institutions_financieres = InstitutionFinanciere.objects.all().order_by('nom_institution')
    
    # Cotisations et paiements
    cotisations_acteurs = CotisationAnnuelleActeur.objects.select_related(
        'acteur'
    ).order_by('-annee', '-date_creation')
    
    cotisations_institutions = CotisationAnnuelleInstitution.objects.select_related(
        'institution'
    ).order_by('-annee', '-date_creation')
    
    paiements_acteurs = PaiementCotisationActeur.objects.select_related(
        'cotisation_annuelle__acteur',
        'encaisse_par_agent'
    ).order_by('-date_paiement')
    
    paiements_institutions = PaiementCotisationInstitution.objects.select_related(
        'cotisation_annuelle__institution',
        'encaisse_par_agent'
    ).order_by('-date_paiement')
    
    # Filtre recherche textuelle sur acteurs et institutions
    if q:
        acteurs_economiques = acteurs_economiques.filter(
            Q(raison_sociale__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(email__icontains=q)
        )
        institutions_financieres = institutions_financieres.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(email__icontains=q)
        )
        cotisations_acteurs = cotisations_acteurs.filter(
            Q(acteur__raison_sociale__icontains=q) |
            Q(acteur__sigle__icontains=q) |
            Q(acteur__nom_responsable__icontains=q)
        )
        cotisations_institutions = cotisations_institutions.filter(
            Q(institution__nom_institution__icontains=q) |
            Q(institution__sigle__icontains=q) |
            Q(institution__nom_responsable__icontains=q)
        )
        paiements_acteurs = paiements_acteurs.filter(
            Q(cotisation_annuelle__acteur__raison_sociale__icontains=q) |
            Q(cotisation_annuelle__acteur__sigle__icontains=q)
        )
        paiements_institutions = paiements_institutions.filter(
            Q(cotisation_annuelle__institution__nom_institution__icontains=q) |
            Q(cotisation_annuelle__institution__sigle__icontains=q)
        )
    
    # Filtres par type (acteurs uniquement / institutions uniquement)
    if type_contribution == 'acteurs':
        institutions_financieres = institutions_financieres.none()
        cotisations_institutions = cotisations_institutions.none()
        paiements_institutions = paiements_institutions.none()
    elif type_contribution == 'institutions':
        acteurs_economiques = acteurs_economiques.none()
        cotisations_acteurs = cotisations_acteurs.none()
        paiements_acteurs = paiements_acteurs.none()
    
    # Filtre par année (pour cotisations et paiements)
    if annee:
        try:
            annee_int = int(annee)
            cotisations_acteurs = cotisations_acteurs.filter(annee=annee_int)
            cotisations_institutions = cotisations_institutions.filter(annee=annee_int)
            paiements_acteurs = paiements_acteurs.filter(cotisation_annuelle__annee=annee_int)
            paiements_institutions = paiements_institutions.filter(cotisation_annuelle__annee=annee_int)
        except ValueError:
            pass
    
    # Filtre par mois (paiements uniquement)
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements_acteurs = paiements_acteurs.filter(date_paiement__month=mois_int)
                paiements_institutions = paiements_institutions.filter(date_paiement__month=mois_int)
        except ValueError:
            pass
    
    # Filtre par agent collecteur (paiements uniquement)
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements_acteurs = paiements_acteurs.filter(encaisse_par_agent_id=agent_id)
            paiements_institutions = paiements_institutions.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    
    # Filtre par période (date_du, date_au) pour paiements
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__gte=date_du_parsed)
        paiements_institutions = paiements_institutions.filter(date_paiement__date__gte=date_du_parsed)
    if date_au_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__lte=date_au_parsed)
        paiements_institutions = paiements_institutions.filter(date_paiement__date__lte=date_au_parsed)
    
    # Années disponibles pour le filtre
    annees_acteurs = sorted(
        CotisationAnnuelleActeur.objects.values_list('annee', flat=True).distinct(),
        reverse=True
    )
    annees_institutions = sorted(
        CotisationAnnuelleInstitution.objects.values_list('annee', flat=True).distinct(),
        reverse=True
    )
    annees_disponibles = sorted(set(annees_acteurs + annees_institutions), reverse=True)
    agents_collecteurs = AgentCollecteur.objects.filter(statut="actif").order_by("matricule", "nom", "prenom")
    
    context = {
        'acteurs_economiques': acteurs_economiques,
        'institutions_financieres': institutions_financieres,
        'cotisations_acteurs': cotisations_acteurs[:100],
        'cotisations_institutions': cotisations_institutions[:100],
        'paiements_acteurs': paiements_acteurs[:100],
        'paiements_institutions': paiements_institutions[:100],
        'titre': '💰 Cotisations Acteurs & Institutions',
        'annees_disponibles': annees_disponibles,
        'agents_collecteurs': agents_collecteurs,
        'current_filters': {
            'type': type_contribution,
            'annee': annee,
            'mois': mois,
            'agent_collecteur': agent_collecteur_id,
            'date_du': date_du,
            'date_au': date_au,
            'q': q
        }
    }
    
    return render(request, "admin/liste_cotisations_acteurs_institutions.html", context)


@login_required
@user_passes_test(is_staff_user)
def definir_taxe_acteur(request, acteur_id):
    """Permet à l'admin de définir ou modifier la taxe (montant annuel dû) pour un acteur économique."""
    acteur = get_object_or_404(ActeurEconomique, id=acteur_id)
    annee_courante = timezone.now().year
    annee_preselect = request.GET.get("annee")
    if annee_preselect:
        try:
            annee_preselect = int(annee_preselect)
        except ValueError:
            annee_preselect = None

    if request.method == "POST":
        annee_str = request.POST.get("annee")
        montant_str = request.POST.get("montant", "").strip()
        if not annee_str:
            messages.error(request, "Veuillez sélectionner une année.")
            return redirect("definir_taxe_acteur", acteur_id=acteur.id)
        try:
            annee = int(annee_str)
        except ValueError:
            messages.error(request, "Année invalide.")
            return redirect("definir_taxe_acteur", acteur_id=acteur.id)
        if not montant_str:
            messages.error(request, "Veuillez saisir le montant annuel dû (FCFA).")
            return redirect("definir_taxe_acteur", acteur_id=acteur.id)
        try:
            montant = Decimal(montant_str)
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, "Montant invalide. Saisissez un nombre.")
            return redirect("definir_taxe_acteur", acteur_id=acteur.id)
        if montant < 0:
            messages.error(request, "Le montant ne peut pas être négatif.")
            return redirect("definir_taxe_acteur", acteur_id=acteur.id)

        cotisation, created = CotisationAnnuelleActeur.objects.update_or_create(
            acteur=acteur,
            annee=annee,
            defaults={"montant_annuel_du": montant},
        )
        if created:
            messages.success(
                request,
                f"Taxe {annee} définie pour {acteur.raison_sociale}: {montant:,.0f} FCFA.",
            )
        else:
            messages.success(
                request,
                f"Taxe {annee} mise à jour pour {acteur.raison_sociale}: {montant:,.0f} FCFA.",
            )
        return redirect("liste_cotisations_acteurs_institutions")

    cotisations_existantes = CotisationAnnuelleActeur.objects.filter(acteur=acteur).order_by("-annee")
    annees_possibles = list(range(annee_courante, annee_courante - 5, -1))

    context = {
        "acteur": acteur,
        "type_entite": "acteur",
        "cotisations_existantes": cotisations_existantes,
        "annees_possibles": annees_possibles,
        "annee_courante": annee_courante,
        "annee_preselect": annee_preselect,
        "titre": "Définir la taxe - Acteur économique",
    }
    return render(request, "admin/definir_taxe.html", context)


@login_required
@user_passes_test(is_staff_user)
def definir_taxe_institution(request, institution_id):
    """Permet à l'admin de définir ou modifier la taxe (montant annuel dû) pour une institution financière."""
    institution = get_object_or_404(InstitutionFinanciere, id=institution_id)
    annee_courante = timezone.now().year
    annee_preselect = request.GET.get("annee")
    if annee_preselect:
        try:
            annee_preselect = int(annee_preselect)
        except ValueError:
            annee_preselect = None

    if request.method == "POST":
        annee_str = request.POST.get("annee")
        montant_str = request.POST.get("montant", "").strip()
        if not annee_str:
            messages.error(request, "Veuillez sélectionner une année.")
            return redirect("definir_taxe_institution", institution_id=institution.id)
        try:
            annee = int(annee_str)
        except ValueError:
            messages.error(request, "Année invalide.")
            return redirect("definir_taxe_institution", institution_id=institution.id)
        if not montant_str:
            messages.error(request, "Veuillez saisir le montant annuel dû (FCFA).")
            return redirect("definir_taxe_institution", institution_id=institution.id)
        try:
            montant = Decimal(montant_str)
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, "Montant invalide. Saisissez un nombre.")
            return redirect("definir_taxe_institution", institution_id=institution.id)
        if montant < 0:
            messages.error(request, "Le montant ne peut pas être négatif.")
            return redirect("definir_taxe_institution", institution_id=institution.id)

        cotisation, created = CotisationAnnuelleInstitution.objects.update_or_create(
            institution=institution,
            annee=annee,
            defaults={"montant_annuel_du": montant},
        )
        if created:
            messages.success(
                request,
                f"Taxe {annee} définie pour {institution.nom_institution}: {montant:,.0f} FCFA.",
            )
        else:
            messages.success(
                request,
                f"Taxe {annee} mise à jour pour {institution.nom_institution}: {montant:,.0f} FCFA.",
            )
        return redirect("liste_cotisations_acteurs_institutions")

    cotisations_existantes = CotisationAnnuelleInstitution.objects.filter(
        institution=institution
    ).order_by("-annee")
    annees_possibles = list(range(annee_courante, annee_courante - 5, -1))

    context = {
        "institution": institution,
        "type_entite": "institution",
        "cotisations_existantes": cotisations_existantes,
        "annees_possibles": annees_possibles,
        "annee_courante": annee_courante,
        "annee_preselect": annee_preselect,
        "titre": "Définir la taxe - Institution financière",
    }
    return render(request, "admin/definir_taxe.html", context)


@login_required
@user_passes_test(is_staff_user)
def detail_suggestion(request, pk):
    """Affiche le détail d'une suggestion."""
    
    suggestion = get_object_or_404(Suggestion, pk=pk)
    
    # Marquer automatiquement comme lue si ce n'est pas déjà fait
    if not suggestion.est_lue:
        suggestion.est_lue = True
        if not suggestion.date_lecture:
            suggestion.date_lecture = timezone.now()
        suggestion.save()
    
    context = {
        'suggestion': suggestion,
    }
    
    return render(request, "admin/detail_suggestion.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_suggestions(request):
    """Liste des suggestions soumises par les visiteurs."""
    
    suggestions = Suggestion.objects.all().order_by('-date_soumission')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    est_lue = request.GET.get('est_lue', '')
    
    # Application des filtres
    if q:
        suggestions = suggestions.filter(
            Q(nom__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q) |
            Q(sujet__icontains=q) |
            Q(message__icontains=q)
        )
    
    if est_lue == 'oui':
        suggestions = suggestions.filter(est_lue=True)
    elif est_lue == 'non':
        suggestions = suggestions.filter(est_lue=False)
    
    context = {
        'suggestions': suggestions,
        'titre': 'Suggestions des Visiteurs',
        'current_filters': {
            'q': q,
            'est_lue': est_lue
        }
    }
    
    return render(request, "admin/liste_suggestions.html", context)


@login_required
@user_passes_test(is_staff_user)
def liste_candidatures(request):
    """Liste des candidatures aux appels d'offres."""
    
    candidatures = Candidature.objects.all().select_related('appel_offre', 'candidat').order_by('-date_soumission')
    
    # Récupération des paramètres de filtrage
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    # Application des filtres
    if q:
        candidatures = candidatures.filter(
            Q(appel_offre__titre__icontains=q) |
            Q(appel_offre__reference__icontains=q) |
            Q(candidat__first_name__icontains=q) |
            Q(candidat__last_name__icontains=q) |
            Q(candidat__email__icontains=q)
        )
    
    if statut:
        candidatures = candidatures.filter(statut=statut)
    
    # Grouper les candidatures par appel d'offres et compter les acceptées
    appels_offres_avec_candidatures = {}
    for candidature in candidatures:
        appel_id = candidature.appel_offre.id
        if appel_id not in appels_offres_avec_candidatures:
            appels_offres_avec_candidatures[appel_id] = {
                'appel_offre': candidature.appel_offre,
                'nb_acceptees': 0
            }
        if candidature.statut == 'acceptee':
            appels_offres_avec_candidatures[appel_id]['nb_acceptees'] += 1
    
    # Créer une liste triée des appels d'offres avec candidatures acceptées pour faciliter l'affichage
    appels_avec_acceptees = [
        (appel_id, info) 
        for appel_id, info in appels_offres_avec_candidatures.items() 
        if info['nb_acceptees'] > 0
    ]
    
    context = {
        'candidatures': candidatures,
        'appels_offres_avec_candidatures': appels_offres_avec_candidatures,
        'appels_avec_acceptees': appels_avec_acceptees,
        'titre': 'Candidatures aux Appels d\'Offres',
        'statut_choices': Candidature.STATUT_CANDIDATURE,
        'current_filters': {
            'q': q,
            'statut': statut
        }
    }
    
    return render(request, "admin/liste_candidatures.html", context)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_candidatures(request, appel_offre_id):
    """
    Génère un PDF des candidatures acceptées pour un appel d'offres spécifique.
    """
    appel_offre = get_object_or_404(AppelOffre, pk=appel_offre_id)
    
    candidatures = Candidature.objects.filter(
        appel_offre=appel_offre,
        statut="acceptee"
    ).select_related("appel_offre", "candidat").order_by("-date_soumission")

    if not candidatures.exists():
        messages.warning(
            request,
            f"Aucun dossier accepté pour l'appel d'offres '{appel_offre.titre}'.",
        )
        return redirect("liste_candidatures")

    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    filename = _make_pdf_filename("candidatures-acceptees", appel_offre.reference or appel_offre.titre)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story = [
        Paragraph("Candidatures acceptées", title_style),
        Paragraph(f"Appel d'offres : {escape(appel_offre.titre)}", styles["Heading2"]),
        Spacer(1, 0.2 * cm),
    ]
    
    if appel_offre.reference:
        story.append(Paragraph(f"Référence : {escape(appel_offre.reference)}", styles["Normal"]))
    
    story.append(Spacer(1, 0.4 * cm))

    # Tableau : Nom de l'entreprise ou Nom & Prénoms du candidat, Email, Date soumission, Téléphone
    data = [["Nom / Raison sociale", "Email", "Date de soumission", "Téléphone"]]
    for candidature in candidatures:
        user = candidature.candidat

        # Nom / Raison sociale
        full_name = user.get_full_name() or user.username
        display_name = full_name

        # Si l'utilisateur est lié à une entreprise ou institution, on affiche la raison sociale
        acteur = getattr(user, "acteur_economique", None)
        institution = getattr(user, "institution_financiere", None)
        profil = getattr(user, "profil_emploi", None)

        if acteur is not None:
            display_name = acteur.raison_sociale
        elif institution is not None:
            display_name = institution.nom_institution
        elif profil is not None:
            display_name = f"{profil.nom} {profil.prenoms}"

        # Numéro de téléphone
        telephone = ""
        if acteur is not None:
            telephone = acteur.telephone1
        elif institution is not None:
            telephone = institution.telephone1
        elif profil is not None:
            telephone = profil.telephone1

        data.append(
            [
                escape(display_name),
                user.email,
                candidature.date_soumission.strftime("%d/%m/%Y %H:%M"),
                telephone,
            ]
        )

    table = Table(data, colWidths=[7 * cm, 7 * cm, 4.5 * cm, 4.5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(canvas, doc):
        _draw_pdf_header(canvas, doc, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
@require_POST
def changer_statut(request, model_name, pk, action):
    """Change le statut d'un objet (accepter/refuser)."""
    
    model_map = {
        'candidature': Candidature,
        'acteur': ActeurEconomique,
        'institution': InstitutionFinanciere,
        'jeune': ProfilEmploi,
        'retraite': ProfilEmploi,
        'diaspora': MembreDiaspora,
        'suggestion': Suggestion,
        'osc': OrganisationSocieteCivile,
    }
    
    ModelClass = model_map.get(model_name)
    if not ModelClass:
        messages.error(request, "Type d'objet invalide.")
        return redirect('tableau_bord')
        
    obj = get_object_or_404(ModelClass, pk=pk)
    
    if model_name == 'candidature':
        if action == 'accepter':
            obj.statut = 'acceptee'
            messages.success(request, f"Candidature de {obj.candidat} acceptée.")
        elif action in ['refuser', 'rejeter']:
            obj.statut = 'refusee'
            messages.warning(request, f"Candidature de {obj.candidat} refusée.")
    elif model_name == 'suggestion':
        if action == 'marquer_lue':
            obj.est_lue = True
            if not obj.date_lecture:
                obj.date_lecture = timezone.now()
            messages.success(request, f"Suggestion de {obj.nom} marquée comme lue.")
    else:
        # Pour les autres modèles, on utilise est_valide_par_mairie
        if action == 'accepter':
            obj.est_valide_par_mairie = True
            messages.success(request, f"{obj} validé avec succès.")
        elif action in ['refuser', 'rejeter']:
            obj.est_valide_par_mairie = False
            messages.warning(request, f"{obj} refusé/invalidé.")
            
    obj.save()
    
    # Vérifier si une redirection personnalisée est demandée
    redirect_to = request.POST.get('redirect_to')
    if redirect_to:
        return redirect(redirect_to)
    
    # Redirection vers la liste appropriée
    redirect_map = {
        'candidature': 'liste_candidatures',
        'acteur': 'liste_acteurs',
        'institution': 'liste_institutions',
        'jeune': 'liste_jeunes',
        'retraite': 'liste_retraites',
        'diaspora': 'liste_diaspora_tableau_bord',
        'suggestion': 'liste_suggestions',
        'osc': 'liste_osc_tableau_bord',
    }
    
    return redirect(redirect_map.get(model_name, 'tableau_bord'))


@login_required
@user_passes_test(is_staff_user)
def export_pdf_acteur_detail(request, pk):
    acteur = get_object_or_404(ActeurEconomique, pk=pk)
    sections = [
        (
            "Informations générales",
            [
                ("Raison sociale", acteur.raison_sociale),
                ("Sigle / Acronyme", acteur.sigle),
                ("Type d'acteur", acteur.get_type_acteur_display()),
                ("Secteur d'activité", acteur.get_secteur_activite_display()),
                ("Statut juridique", acteur.get_statut_juridique_display()),
                ("Description", acteur.description),
            ],
        ),
        (
            "Informations légales et fiscales",
            [
                ("N° RCCM", acteur.rccm),
                ("N° CFE", acteur.cfe),
                ("N° Carte opérateur économique", acteur.numero_carte_operateur),
                ("NIF", acteur.nif),
                ("Date de création", acteur.date_creation),
                (
                    "Capital social",
                    f"{acteur.capital_social} FCFA" if acteur.capital_social is not None else None,
                ),
            ],
        ),
        (
            "Responsable et contacts",
            [
                ("Nom du responsable", acteur.nom_responsable),
                ("Fonction", acteur.fonction_responsable),
                ("Téléphone principal", acteur.telephone1),
                ("Téléphone secondaire", acteur.telephone2),
                ("Email professionnel", acteur.email),
                ("Site web", acteur.site_web),
            ],
        ),
        (
            "Localisation et présence",
            [
                ("Situation", acteur.get_situation_display()),
                ("Quartier", acteur.quartier),
                ("Canton", acteur.canton),
                ("Adresse complète", acteur.adresse_complete),
            ],
        ),
        (
            "Informations complémentaires",
            [
                (
                    "Nombre d'employés",
                    acteur.get_nombre_employes_display() if acteur.nombre_employes else None,
                ),
                (
                    "Chiffre d'affaires",
                    acteur.get_chiffre_affaires_display() if acteur.chiffre_affaires else None,
                ),
                ("Accepte publication publique", acteur.accepte_public),
                ("Certifie les informations", acteur.certifie_information),
                ("Accepte les conditions", acteur.accepte_conditions),
                ("Validé par la mairie", acteur.est_valide_par_mairie),
                ("Date d'enregistrement", acteur.date_enregistrement),
            ],
        ),
    ]

    filename = _make_pdf_filename("acteur", acteur.raison_sociale)
    title = f"Fiche Acteur Économique - {acteur.raison_sociale}"
    return _build_detail_pdf(filename, title, sections)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_osc_detail(request, pk):
    """Génère une fiche PDF détaillée pour une OSC (comme pour un acteur économique)."""
    osc = get_object_or_404(OrganisationSocieteCivile, pk=pk)

    sections = [
        (
            "Informations générales",
            [
                ("Nom de l'OSC", osc.nom_osc),
                ("Sigle", osc.sigle),
                ("Type d'OSC", get_osc_type_display(osc.type_osc)),
                ("Date de création", osc.date_creation),
            ],
        ),
        (
            "Coordonnées",
            [
                ("Adresse", osc.adresse),
                ("Téléphone", osc.telephone),
                ("Email", osc.email),
            ],
        ),
        (
            "Domaines d'intervention",
            [
                (
                    "Domaines d'intervention",
                    osc.domaines_intervention,
                ),
            ],
        ),
        (
            "Membres / Responsables",
            [
                (
                    "Membres / Responsables",
                    osc.membres_responsables,
                ),
            ],
        ),
        (
            "Statut et métadonnées",
            [
                ("Validée par la mairie", osc.est_valide_par_mairie),
                ("Date d'enregistrement", osc.date_enregistrement),
                ("Utilisateur associé", getattr(osc.user, "username", None)),
            ],
        ),
    ]

    filename = _make_pdf_filename("osc", osc.nom_osc)
    title = f"Fiche Organisation de la Société Civile - {osc.nom_osc}"
    return _build_detail_pdf(filename, title, sections)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_institution_detail(request, pk):
    institution = get_object_or_404(InstitutionFinanciere, pk=pk)
    services_text = (
        ", ".join(part.strip().title() for part in institution.services.split(",") if part.strip())
        if institution.services
        else None
    )

    sections = [
        (
            "Informations générales",
            [
                ("Nom de l'institution", institution.nom_institution),
                ("Sigle", institution.sigle),
                ("Type d'institution", institution.get_type_institution_display()),
                ("Année de création", institution.annee_creation),
                ("Numéro d'agrément", institution.numero_agrement),
                ("IFU", institution.ifu),
                ("Description des services", institution.description_services),
                ("Services disponibles", services_text),
            ],
        ),
        (
            "Conditions financières",
            [
                ("Taux crédit", institution.taux_credit),
                ("Taux épargne", institution.taux_epargne),
                ("Conditions d'éligibilité", institution.conditions_eligibilite),
                ("Public cible", institution.public_cible),
            ],
        ),
        (
            "Responsable et contacts",
            [
                ("Nom du responsable", institution.nom_responsable),
                ("Fonction", institution.fonction_responsable),
                ("Téléphone principal", institution.telephone1),
                ("Téléphone secondaire", institution.telephone2),
                ("WhatsApp", institution.whatsapp),
                ("Email", institution.email),
                ("Site web", institution.site_web),
                ("Page Facebook", institution.facebook),
            ],
        ),
        (
            "Localisation et présence",
            [
                ("Situation", institution.get_situation_display()),
                ("Quartier", institution.quartier),
                ("Canton", institution.canton),
                ("Adresse complète", institution.adresse_complete),
                ("Nombre d'agences dans Kloto 1", institution.nombre_agences),
                ("Horaires d'ouverture", institution.horaires),
            ],
        ),
        (
            "Engagements et statut",
            [
                ("Certifie les informations", institution.certifie_info),
                ("Accepte la publication publique", institution.accepte_public),
                ("Accepte d'être contacté", institution.accepte_contact),
                ("Engagement pris", institution.engagement),
                ("Validée par la mairie", institution.est_valide_par_mairie),
                ("Date d'enregistrement", institution.date_enregistrement),
            ],
        ),
    ]

    filename = _make_pdf_filename("institution", institution.nom_institution)
    title = f"Fiche Institution Financière - {institution.nom_institution}"
    return _build_detail_pdf(filename, title, sections)


def _export_pdf_profil_detail(pk, profil_type):
    profil = get_object_or_404(ProfilEmploi, pk=pk, type_profil=profil_type)

    sections = [
        (
            "Identité",
            [
                ("Type de profil", profil.get_type_profil_display()),
                ("Nom", profil.nom),
                ("Prénoms", profil.prenoms),
                ("Sexe", profil.get_sexe_display()),
                ("Date de naissance", profil.date_naissance),
                ("Nationalité", profil.nationalite),
                ("Résident Kloto 1", profil.est_resident_kloto),
            ],
        ),
        (
            "Coordonnées",
            [
                ("Téléphone principal", profil.telephone1),
                ("Téléphone secondaire", profil.telephone2),
                ("Email", profil.email),
                ("Quartier", profil.quartier),
                ("Canton", profil.canton),
                ("Adresse complète", profil.adresse_complete),
            ],
        ),
        (
            "Formation et compétences",
            [
                ("Niveau d'étude", profil.get_niveau_etude_display() if profil.niveau_etude else None),
                ("Diplôme principal", profil.diplome_principal),
                ("Domaine de compétence", profil.domaine_competence),
                ("Expériences", profil.experiences),
                ("Dernier poste occupé", profil.dernier_poste),
                ("Années d'expérience", profil.annees_experience),
            ],
        ),
        (
            "Situation professionnelle",
            [
                ("Situation actuelle", profil.get_situation_actuelle_display()),
                ("Employeur actuel", profil.employeur_actuel),
                ("Disponibilité", profil.get_disponibilite_display()),
                (
                    "Type de contrat souhaité",
                    profil.get_type_contrat_souhaite_display() if profil.type_contrat_souhaite else None,
                ),
                ("Salaire souhaité", profil.salaire_souhaite),
                ("Caisse de retraite / régime", profil.caisse_retraite),
            ],
        ),
        (
            "Consentements et statut",
            [
                ("Accepte le traitement des données", profil.accepte_rgpd),
                ("Accepte d'être contacté", profil.accepte_contact),
                ("Validé par la mairie", profil.est_valide_par_mairie),
                ("Date d'inscription", profil.date_inscription),
            ],
        ),
    ]

    filename = _make_pdf_filename(profil_type, f"{profil.nom}-{profil.prenoms}")
    title = f"Fiche {profil.get_type_profil_display()} - {profil.nom} {profil.prenoms}"
    return _build_detail_pdf(filename, title, sections)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_jeune_detail(request, pk):
    return _export_pdf_profil_detail(pk, "jeune")


@login_required
@user_passes_test(is_staff_user)
def export_pdf_retraite_detail(request, pk):
    return _export_pdf_profil_detail(pk, "retraite")


@login_required
@user_passes_test(is_staff_user)
def export_pdf_diaspora_detail(request, pk):
    """Génère un PDF détaillé pour un membre de la diaspora."""
    membre = get_object_or_404(MembreDiaspora, pk=pk)
    
    # Récupérer les appuis financiers
    appuis_financiers = []
    if membre.appui_investissement_projets:
        appuis_financiers.append("Investissement dans des projets communaux")
    if membre.appui_financement_infrastructures:
        appuis_financiers.append("Financement d'infrastructures")
    if membre.appui_parrainage_communautaire:
        appuis_financiers.append("Parrainage de projets communautaires")
    if membre.appui_jeunes_femmes_entrepreneurs:
        appuis_financiers.append("Appui aux jeunes et femmes entrepreneurs")
    
    # Récupérer les compétences techniques
    competences_techniques = []
    if membre.transfert_competences:
        competences_techniques.append("Transfert de compétences")
    if membre.formation_jeunes:
        competences_techniques.append("Formation des jeunes")
    if membre.appui_digitalisation:
        competences_techniques.append("Appui à la digitalisation")
    if membre.conseils_techniques:
        competences_techniques.append("Conseils techniques / expertise")
    if membre.encadrement_mentorat:
        competences_techniques.append("Encadrement à distance (mentorat)")
    
    # Création d'emplois
    creation_emplois = []
    if membre.creation_entreprise_locale:
        creation_emplois.append("Création d'entreprise locale")
    if membre.appui_pme_locales:
        creation_emplois.append("Appui aux PME locales")
    if membre.recrutement_jeunes_commune:
        creation_emplois.append("Recrutement de jeunes de la commune")
    
    # Partenariats
    partenariats = []
    if membre.mise_relation_ong:
        partenariats.append("Mise en relation avec ONG")
    if membre.cooperation_decentralisee:
        partenariats.append("Coopération décentralisée")
    if membre.recherche_financements_internationaux:
        partenariats.append("Recherche de financements internationaux")
    if membre.promotion_commune_international:
        partenariats.append("Promotion de la commune à l'international")
    
    # Engagement citoyen
    engagement_citoyen = []
    if membre.participation_activites_communales:
        engagement_citoyen.append("Participation aux activités communales")
    if membre.participation_reunions_diaspora:
        engagement_citoyen.append("Participation aux réunions de la diaspora")
    if membre.appui_actions_sociales_culturelles:
        engagement_citoyen.append("Appui aux actions sociales et culturelles")
    
    sections = [
        (
            "Informations d'identification",
            [
                ("Nom", membre.nom),
                ("Prénoms", membre.prenoms),
                ("Sexe", membre.get_sexe_display()),
                ("Date de naissance", membre.date_naissance),
                ("Nationalité(s)", membre.nationalites),
                ("Numéro de pièce d'identité", membre.numero_piece_identite),
            ],
        ),
        (
            "Résidence actuelle",
            [
                ("Pays de résidence", membre.pays_residence_actuelle),
                ("Ville de résidence", membre.ville_residence_actuelle),
                ("Adresse complète à l'étranger", membre.adresse_complete_etranger),
            ],
        ),
        (
            "Lien avec la commune",
            [
                ("Commune d'origine", membre.commune_origine),
                ("Quartier / Village d'origine", membre.quartier_village_origine),
                ("Nom du parent/tuteur originaire", membre.nom_parent_tuteur_originaire),
                ("Année de départ du pays", membre.annee_depart_pays),
                ("Fréquence de retour au pays", membre.get_frequence_retour_pays_display()),
            ],
        ),
        (
            "Informations de contact",
            [
                ("Téléphone (WhatsApp)", membre.telephone_whatsapp),
                ("Email", membre.email),
                ("Réseaux sociaux", membre.reseaux_sociaux or "Non renseigné"),
                ("Contact au pays - Nom", membre.contact_au_pays_nom),
                ("Contact au pays - Téléphone", membre.contact_au_pays_telephone),
            ],
        ),
        (
            "Situation professionnelle",
            [
                ("Niveau d'études", membre.get_niveau_etudes_display()),
                ("Domaine de formation", membre.domaine_formation),
                ("Profession actuelle", membre.profession_actuelle),
                ("Secteur d'activité", membre.get_secteur_activite_display()),
                ("Secteur d'activité (autre)", membre.secteur_activite_autre or "Non renseigné"),
                ("Années d'expérience", membre.annees_experience),
                ("Statut professionnel", membre.get_statut_professionnel_display()),
                ("Type de titre de séjour", membre.type_titre_sejour or "Non renseigné"),
            ],
        ),
        (
            "Appui financier proposé",
            [
                ("Types d'appui", ", ".join(appuis_financiers) if appuis_financiers else "Aucun"),
            ],
        ),
        (
            "Appui technique & compétences",
            [
                ("Compétences proposées", ", ".join(competences_techniques) if competences_techniques else "Aucune"),
            ],
        ),
        (
            "Création d'emplois",
            [
                ("Actions proposées", ", ".join(creation_emplois) if creation_emplois else "Aucune"),
            ],
        ),
        (
            "Partenariats & relations internationales",
            [
                ("Actions proposées", ", ".join(partenariats) if partenariats else "Aucune"),
            ],
        ),
        (
            "Engagement citoyen",
            [
                ("Actions proposées", ", ".join(engagement_citoyen) if engagement_citoyen else "Aucune"),
            ],
        ),
        (
            "Questions clés",
            [
                ("Comment souhaitez-vous contribuer ?", membre.comment_contribuer),
                ("Disposition à participer", membre.get_disposition_participation_display()),
                ("Domaine d'intervention prioritaire", membre.domaine_intervention_prioritaire),
            ],
        ),
        (
            "Validation et métadonnées",
            [
                ("Accepte RGPD", membre.accepte_rgpd),
                ("Accepte d'être contacté", membre.accepte_contact),
                ("Validé par la mairie", membre.est_valide_par_mairie),
                ("Date d'inscription", membre.date_inscription),
                ("Date de modification", membre.date_modification),
            ],
        ),
    ]
    
    filename = _make_pdf_filename("diaspora", f"{membre.nom}-{membre.prenoms}")
    title = f"Fiche Membre de la Diaspora - {membre.nom} {membre.prenoms}"
    return _build_detail_pdf(filename, title, sections)


@login_required
@user_passes_test(is_staff_user)
def export_pdf_acteurs(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    type_acteur = request.GET.get("type") or ""
    secteur = request.GET.get("secteur") or ""

    # Uniquement les acteurs validés par la mairie
    qs = ActeurEconomique.objects.filter(est_valide_par_mairie=True)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if type_acteur:
        qs = qs.filter(type_acteur=type_acteur)
    if secteur:
        qs = qs.filter(secteur_activite=secteur)

    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="acteurs_economiques_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Acteurs Économiques", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Raison sociale", "Type", "Secteur", "Responsable", "Téléphone"]]
    for a in qs.order_by("-date_enregistrement")[:1000]:
        data.append([
            a.raison_sociale,
            a.get_type_acteur_display(),
            a.get_secteur_activite_display(),
            a.nom_responsable,
            a.telephone1,
        ])
    table = Table(data, colWidths=[7*cm, 4*cm, 5*cm, 6*cm, 4*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_diaspora(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    pays = request.GET.get("pays") or ""

    # Uniquement les membres validés par la mairie
    qs = MembreDiaspora.objects.filter(est_valide_par_mairie=True)
    if pays:
        qs = qs.filter(pays_residence_actuelle__icontains=pays)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="diaspora_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Membres de la Diaspora", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Pays de résidence", "Ville", "Téléphone", "Email", "Profession"]]
    for m in qs.order_by("-date_inscription")[:1000]:
        data.append([
            m.nom,
            m.prenoms,
            m.pays_residence_actuelle[:25] if m.pays_residence_actuelle else "",
            m.ville_residence_actuelle[:20] if m.ville_residence_actuelle else "",
            m.telephone_whatsapp[:15] if m.telephone_whatsapp else "",
            m.email[:30] if m.email else "",
            m.profession_actuelle[:30] if m.profession_actuelle else "",
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3.5*cm, 5*cm, 4.5*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_osc(request):
    """Export PDF des Organisations de la Société Civile (OSC) validées, par type et/ou période."""

    start = request.GET.get("start")
    end = request.GET.get("end")
    type_osc = request.GET.get("type", "").strip()

    qs = OrganisationSocieteCivile.objects.filter(est_valide_par_mairie=True)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()

    if type_osc:
        qs = qs.filter(type_osc=type_osc)

    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__lte=ed)
        except ValueError:
            pass

    type_label = get_osc_type_display(type_osc) if type_osc else ""
    filename = "osc_valides.pdf"
    if type_label:
        filename = _make_pdf_filename("osc", type_label)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Organisations de la Société Civile (OSC)", title_style))
    if type_label:
        story.append(Paragraph(f"{type_label} validées", styles["Normal"]))
    else:
        story.append(Paragraph("(Uniquement les OSC validées par la mairie)", styles["Normal"]))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))

    data = [["Nom de l'OSC", "Sigle", "Type", "Téléphone", "Email"]]
    for o in qs.order_by("-date_enregistrement")[:1000]:
        data.append(
            [
                o.nom_osc,
                o.sigle or "",
                get_osc_type_display(o.type_osc),
                o.telephone or "",
                o.email or "",
            ]
        )

    table = Table(data, colWidths=[7 * cm, 3 * cm, 6 * cm, 4 * cm, 6 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def notifications_candidats(request):
    """Liste des appels d'offres avec candidats acceptés pour envoyer des notifications."""
    
    # Récupérer tous les appels d'offres qui ont au moins une candidature acceptée
    appels_offres = AppelOffre.objects.filter(
        candidatures__statut='acceptee'
    ).distinct().annotate(
        nb_candidats_acceptes=Count('candidatures', filter=Q(candidatures__statut='acceptee'))
    ).order_by('-date_debut')
    
    context = {
        'appels_offres': appels_offres,
    }
    
    return render(request, "admin/notifications_candidats.html", context)


@login_required
@user_passes_test(is_staff_user)
def envoyer_notifications_candidats(request, appel_offre_id):
    """Affiche le formulaire et traite l'envoi de notifications aux candidats acceptés."""
    
    appel_offre = get_object_or_404(AppelOffre, pk=appel_offre_id)
    
    # Récupérer uniquement les candidats acceptés pour cet appel d'offres
    candidats_acceptes = Candidature.objects.filter(
        appel_offre=appel_offre,
        statut='acceptee'
    ).select_related('candidat')
    
    if not candidats_acceptes.exists():
        messages.warning(request, "Aucun candidat accepté trouvé pour cet appel d'offres.")
        return redirect('notifications_candidats')
    
    if request.method == 'POST':
        titre = request.POST.get('titre', '')
        message = request.POST.get('message', '')
        type_notification = request.POST.get('type', Notification.TYPE_INFO)
        
        if not titre or not message:
            messages.error(request, "Le titre et le message sont obligatoires.")
        else:
            # Créer une notification pour chaque candidat accepté
            notifications_creees = 0
            for candidature in candidats_acceptes:
                Notification.objects.create(
                    recipient=candidature.candidat,
                    title=titre,
                    message=message,
                    type=type_notification,
                    created_by=request.user,
                )
                notifications_creees += 1
            
            messages.success(
                request, 
                f"Notifications envoyées avec succès à {notifications_creees} candidat(s) accepté(s)."
            )
            return redirect('notifications_candidats')
    
    context = {
        'appel_offre': appel_offre,
        'candidats_acceptes': candidats_acceptes,
        'nb_candidats': candidats_acceptes.count(),
        'type_choices': Notification.TYPE_CHOICES,
    }
    
    return render(request, "admin/envoyer_notifications_candidats.html", context)


# ========== FONCTIONS D'EXPORT EXCEL ==========

def _format_excel_value(value):
    """Formate une valeur pour l'export Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else value.strftime("%d/%m/%Y")
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value if item)
    return str(value)


def _style_excel_header(ws, row_num):
    """Applique un style au header Excel."""
    header_fill = PatternFill(start_color="006233", end_color="006233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


@login_required
@user_passes_test(is_staff_user)
def export_excel_acteurs(request):
    """Exporte tous les acteurs économiques en Excel avec tous les champs."""
    acteurs = ActeurEconomique.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres comme dans la vue liste
    q = request.GET.get('q', '')
    type_acteur = request.GET.get('type', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        acteurs = acteurs.filter(
            Q(raison_sociale__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_acteur:
        acteurs = acteurs.filter(type_acteur=type_acteur)
    if secteur:
        acteurs = acteurs.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Acteurs Economiques"
    
    # En-têtes
    headers = [
        "ID", "Raison sociale", "Sigle", "Type d'acteur", "Secteur d'activité", "Statut juridique",
        "Description", "RCCM", "CFE", "N° Carte opérateur", "NIF", "Date de création",
        "Capital social (FCFA)", "Nom responsable", "Fonction responsable", "Téléphone 1",
        "Téléphone 2", "Email", "Site web", "Quartier", "Canton", "Adresse complète",
        "Situation", "Latitude", "Longitude", "Nombre d'employés", "Chiffre d'affaires",
        "Accepte publication", "Certifie informations", "Accepte conditions",
        "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for acteur in acteurs:
        row = [
            acteur.pk,
            acteur.raison_sociale,
            acteur.sigle or "",
            acteur.get_type_acteur_display(),
            acteur.get_secteur_activite_display(),
            acteur.get_statut_juridique_display(),
            acteur.description,
            acteur.rccm or "",
            acteur.cfe or "",
            acteur.numero_carte_operateur or "",
            acteur.nif or "",
            _format_excel_value(acteur.date_creation),
            acteur.capital_social if acteur.capital_social else "",
            acteur.nom_responsable,
            acteur.fonction_responsable,
            acteur.telephone1,
            acteur.telephone2 or "",
            acteur.email,
            acteur.site_web or "",
            acteur.quartier,
            acteur.canton or "",
            acteur.adresse_complete,
            acteur.get_situation_display(),
            acteur.latitude if acteur.latitude else "",
            acteur.longitude if acteur.longitude else "",
            acteur.get_nombre_employes_display() if acteur.nombre_employes else "",
            acteur.get_chiffre_affaires_display() if acteur.chiffre_affaires else "",
            "Oui" if acteur.accepte_public else "Non",
            "Oui" if acteur.certifie_information else "Non",
            "Oui" if acteur.accepte_conditions else "Non",
            "Oui" if acteur.est_valide_par_mairie else "Non",
            _format_excel_value(acteur.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="acteurs_economiques.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_institutions(request):
    """Exporte toutes les institutions financières en Excel avec tous les champs."""
    institutions = InstitutionFinanciere.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    type_inst = request.GET.get('type', '')
    
    if q:
        institutions = institutions.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_inst:
        institutions = institutions.filter(type_institution=type_inst)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Institutions Financieres"
    
    # En-têtes
    headers = [
        "ID", "Nom institution", "Sigle", "Type institution", "Année création",
        "N° Agrément", "IFU", "Description services", "Services disponibles",
        "Taux crédit", "Taux épargne", "Conditions éligibilité", "Public cible",
        "Nom responsable", "Fonction responsable", "Téléphone 1", "Téléphone 2",
        "WhatsApp", "Email", "Site web", "Facebook", "Quartier", "Canton",
        "Adresse complète", "Situation", "Latitude", "Longitude", "Nombre agences",
        "Horaires", "Certifie informations", "Accepte publication", "Accepte contact",
        "Engagement", "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for inst in institutions:
        services_text = ", ".join(part.strip().title() for part in inst.services.split(",") if part.strip()) if inst.services else ""
        row = [
            inst.pk,
            inst.nom_institution,
            inst.sigle or "",
            inst.get_type_institution_display(),
            inst.annee_creation if inst.annee_creation else "",
            inst.numero_agrement or "",
            inst.ifu or "",
            inst.description_services,
            services_text,
            inst.taux_credit or "",
            inst.taux_epargne or "",
            inst.conditions_eligibilite or "",
            inst.public_cible or "",
            inst.nom_responsable,
            inst.fonction_responsable,
            inst.telephone1,
            inst.telephone2 or "",
            inst.whatsapp or "",
            inst.email,
            inst.site_web or "",
            inst.facebook or "",
            inst.quartier,
            inst.canton or "",
            inst.adresse_complete,
            inst.get_situation_display(),
            inst.latitude if inst.latitude else "",
            inst.longitude if inst.longitude else "",
            inst.nombre_agences if inst.nombre_agences else "",
            inst.horaires,
            "Oui" if inst.certifie_info else "Non",
            "Oui" if inst.accepte_public else "Non",
            "Oui" if inst.accepte_contact else "Non",
            "Oui" if inst.engagement else "Non",
            "Oui" if inst.est_valide_par_mairie else "Non",
            _format_excel_value(inst.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="institutions_financieres.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_jeunes(request):
    """Exporte tous les jeunes demandeurs d'emploi en Excel avec tous les champs."""
    jeunes = ProfilEmploi.objects.filter(type_profil='jeune').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        jeunes = jeunes.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        jeunes = jeunes.filter(niveau_etude=niveau)
    if dispo:
        jeunes = jeunes.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jeunes Demandeurs Emploi"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Service citoyen obligatoire", "Accepte RGPD", "Accepte contact",
        "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for jeune in jeunes:
        row = [
            jeune.pk,
            jeune.nom,
            jeune.prenoms,
            jeune.get_sexe_display(),
            _format_excel_value(jeune.date_naissance),
            jeune.nationalite or "",
            jeune.telephone1,
            jeune.telephone2 or "",
            jeune.email,
            jeune.quartier,
            jeune.canton or "",
            jeune.adresse_complete,
            "Oui" if jeune.est_resident_kloto else "Non",
            jeune.get_niveau_etude_display() if jeune.niveau_etude else "",
            jeune.diplome_principal or "",
            jeune.domaine_competence,
            jeune.experiences or "",
            jeune.get_situation_actuelle_display(),
            jeune.employeur_actuel or "",
            jeune.get_disponibilite_display(),
            jeune.get_type_contrat_souhaite_display() if jeune.type_contrat_souhaite else "",
            jeune.salaire_souhaite or "",
            "Oui" if jeune.service_citoyen_obligatoire else "Non",
            "Oui" if jeune.accepte_rgpd else "Non",
            "Oui" if jeune.accepte_contact else "Non",
            "Oui" if jeune.est_valide_par_mairie else "Non",
            _format_excel_value(jeune.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="jeunes_demandeurs_emploi.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_retraites(request):
    """Exporte tous les retraités actifs en Excel avec tous les champs."""
    retraites = ProfilEmploi.objects.filter(type_profil='retraite').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        retraites = retraites.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        retraites = retraites.filter(niveau_etude=niveau)
    if dispo:
        retraites = retraites.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Retraites Actifs"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Caisse retraite", "Dernier poste", "Années expérience",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for retraite in retraites:
        row = [
            retraite.pk,
            retraite.nom,
            retraite.prenoms,
            retraite.get_sexe_display(),
            _format_excel_value(retraite.date_naissance),
            retraite.nationalite or "",
            retraite.telephone1,
            retraite.telephone2 or "",
            retraite.email,
            retraite.quartier,
            retraite.canton or "",
            retraite.adresse_complete,
            "Oui" if retraite.est_resident_kloto else "Non",
            retraite.get_niveau_etude_display() if retraite.niveau_etude else "",
            retraite.diplome_principal or "",
            retraite.domaine_competence,
            retraite.experiences or "",
            retraite.get_situation_actuelle_display(),
            retraite.employeur_actuel or "",
            retraite.get_disponibilite_display(),
            retraite.get_type_contrat_souhaite_display() if retraite.type_contrat_souhaite else "",
            retraite.salaire_souhaite or "",
            retraite.caisse_retraite or "",
            retraite.dernier_poste or "",
            retraite.annees_experience if retraite.annees_experience else "",
            "Oui" if retraite.accepte_rgpd else "Non",
            "Oui" if retraite.accepte_contact else "Non",
            "Oui" if retraite.est_valide_par_mairie else "Non",
            _format_excel_value(retraite.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="retraites_actifs.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_diaspora(request):
    """Exporte tous les membres de la diaspora en Excel avec tous les champs."""
    membres = MembreDiaspora.objects.all().order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    pays = request.GET.get('pays', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        membres = membres.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_whatsapp__icontains=q) |
            Q(profession_actuelle__icontains=q) |
            Q(domaine_formation__icontains=q)
        )
    if pays:
        membres = membres.filter(pays_residence_actuelle__icontains=pays)
    if secteur:
        membres = membres.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres Diaspora"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité(s)",
        "N° Pièce identité", "Pays résidence", "Ville résidence", "Adresse étranger",
        "Commune origine", "Quartier/Village origine", "Nom parent/tuteur", "Année départ",
        "Fréquence retour", "Téléphone WhatsApp", "Email", "Réseaux sociaux",
        "Contact pays - Nom", "Contact pays - Téléphone", "Niveau études",
        "Domaine formation", "Profession actuelle", "Secteur activité",
        "Secteur activité (autre)", "Années expérience", "Statut professionnel",
        "Type titre séjour", "Appui investissement projets", "Appui financement infrastructures",
        "Appui parrainage communautaire", "Appui jeunes/femmes entrepreneurs",
        "Transfert compétences", "Formation jeunes", "Appui digitalisation",
        "Conseils techniques", "Encadrement mentorat", "Création entreprise locale",
        "Appui PME locales", "Recrutement jeunes commune", "Mise relation ONG",
        "Coopération décentralisée", "Recherche financements internationaux",
        "Promotion commune international", "Participation activités communales",
        "Participation réunions diaspora", "Appui actions sociales/culturelles",
        "Comment contribuer", "Disposition participation", "Domaine intervention prioritaire",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription", "Date modification"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for membre in membres:
        row = [
            membre.pk,
            membre.nom,
            membre.prenoms,
            membre.get_sexe_display(),
            _format_excel_value(membre.date_naissance),
            membre.nationalites,
            membre.numero_piece_identite,
            membre.pays_residence_actuelle,
            membre.ville_residence_actuelle,
            membre.adresse_complete_etranger,
            membre.commune_origine,
            membre.quartier_village_origine,
            membre.nom_parent_tuteur_originaire,
            membre.annee_depart_pays,
            membre.get_frequence_retour_pays_display(),
            membre.telephone_whatsapp,
            membre.email,
            membre.reseaux_sociaux or "",
            membre.contact_au_pays_nom,
            membre.contact_au_pays_telephone,
            membre.get_niveau_etudes_display(),
            membre.domaine_formation,
            membre.profession_actuelle,
            membre.get_secteur_activite_display(),
            membre.secteur_activite_autre or "",
            membre.annees_experience,
            membre.get_statut_professionnel_display(),
            membre.type_titre_sejour or "",
            "Oui" if membre.appui_investissement_projets else "Non",
            "Oui" if membre.appui_financement_infrastructures else "Non",
            "Oui" if membre.appui_parrainage_communautaire else "Non",
            "Oui" if membre.appui_jeunes_femmes_entrepreneurs else "Non",
            "Oui" if membre.transfert_competences else "Non",
            "Oui" if membre.formation_jeunes else "Non",
            "Oui" if membre.appui_digitalisation else "Non",
            "Oui" if membre.conseils_techniques else "Non",
            "Oui" if membre.encadrement_mentorat else "Non",
            "Oui" if membre.creation_entreprise_locale else "Non",
            "Oui" if membre.appui_pme_locales else "Non",
            "Oui" if membre.recrutement_jeunes_commune else "Non",
            "Oui" if membre.mise_relation_ong else "Non",
            "Oui" if membre.cooperation_decentralisee else "Non",
            "Oui" if membre.recherche_financements_internationaux else "Non",
            "Oui" if membre.promotion_commune_international else "Non",
            "Oui" if membre.participation_activites_communales else "Non",
            "Oui" if membre.participation_reunions_diaspora else "Non",
            "Oui" if membre.appui_actions_sociales_culturelles else "Non",
            membre.comment_contribuer,
            membre.get_disposition_participation_display(),
            membre.domaine_intervention_prioritaire,
            "Oui" if membre.accepte_rgpd else "Non",
            "Oui" if membre.accepte_contact else "Non",
            "Oui" if membre.est_valide_par_mairie else "Non",
            _format_excel_value(membre.date_inscription),
            _format_excel_value(membre.date_modification),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="membres_diaspora.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_osc(request):
    """Exporte toutes les OSC en Excel avec les principaux champs."""
    osc_qs = OrganisationSocieteCivile.objects.all().order_by("-date_enregistrement")

    # Filtres simples
    q = request.GET.get("q", "") or ""
    type_osc = request.GET.get("type", "") or ""

    if q:
        osc_qs = osc_qs.filter(
            Q(nom_osc__icontains=q)
            | Q(sigle__icontains=q)
            | Q(email__icontains=q)
            | Q(telephone__icontains=q)
        )
    if type_osc:
        osc_qs = osc_qs.filter(type_osc=type_osc)

    wb = Workbook()
    ws = wb.active
    ws.title = "OSC"

    headers = [
        "ID",
        "Nom de l'OSC",
        "Sigle",
        "Type d'OSC",
        "Date de création",
        "Adresse",
        "Téléphone",
        "Email",
        "Domaines d'intervention",
        "Membres / Responsables",
        "Validé par mairie",
        "Date d'enregistrement",
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)

    for o in osc_qs:
        row = [
            o.pk,
            o.nom_osc,
            o.sigle or "",
            get_osc_type_display(o.type_osc),
            _format_excel_value(o.date_creation),
            o.adresse or "",
            o.telephone or "",
            o.email or "",
            (o.domaines_intervention or "").replace("\n", " / "),
            (o.membres_responsables or "").replace("\n", " / "),
            "Oui" if o.est_valide_par_mairie else "Non",
            _format_excel_value(o.date_enregistrement),
        ]
        ws.append(row)

    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="osc.xlsx"'
    wb.save(response)
    return response


# ========== EXPORTS AGENTS COLLECTEURS, CONTRIBUABLES, BOUTIQUES, CONTRIBUTIONS, COTISATIONS ==========

@login_required
@user_passes_test(is_staff_user)
def export_pdf_agents_collecteurs(request):
    """Export PDF des agents collecteurs (avec filtres q, statut)."""
    q = request.GET.get("q", "").strip()
    statut = request.GET.get("statut", "").strip()
    qs = AgentCollecteur.objects.select_related("user").prefetch_related(
        "emplacements_assignes", "acteurs_economiques", "institutions_financieres"
    ).order_by("-date_creation")
    if q:
        qs = qs.filter(
            Q(matricule__icontains=q)
            | Q(nom__icontains=q)
            | Q(prenom__icontains=q)
            | Q(telephone__icontains=q)
            | Q(email__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="agents_collecteurs.pdf"'
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Agents Collecteurs", title_style))
    if q or statut:
        parts = []
        if q:
            parts.append(f"Recherche: {q}")
        if statut:
            parts.append(f"Statut: {statut}")
        story.append(Paragraph(" | ".join(parts), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [
        [
            "Matricule",
            "Nom",
            "Prénom",
            "Téléphone",
            "Email",
            "Statut",
            "Date embauche",
        ]
    ]
    for a in qs[:1000]:
        data.append(
            [
                a.matricule or "",
                a.nom or "",
                a.prenom or "",
                a.telephone or "",
                (a.email or "")[:30],
                a.get_statut_display(),
                a.date_embauche.strftime("%d/%m/%Y") if a.date_embauche else "",
            ]
        )
    col_widths = [3 * cm, 4 * cm, 4 * cm, 3.5 * cm, 5 * cm, 2.5 * cm, 3 * cm]
    table = Table(data, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_agents_collecteurs(request):
    """Export Excel des agents collecteurs (avec filtres q, statut)."""
    q = request.GET.get("q", "").strip()
    statut = request.GET.get("statut", "").strip()
    qs = AgentCollecteur.objects.select_related("user").prefetch_related(
        "emplacements_assignes", "acteurs_economiques", "institutions_financieres"
    ).order_by("-date_creation")
    if q:
        qs = qs.filter(
            Q(matricule__icontains=q)
            | Q(nom__icontains=q)
            | Q(prenom__icontains=q)
            | Q(telephone__icontains=q)
            | Q(email__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    wb = Workbook()
    ws = wb.active
    ws.title = "Agents Collecteurs"
    headers = [
        "ID",
        "Matricule",
        "Nom",
        "Prénom",
        "Téléphone",
        "Email",
        "Statut",
        "Date embauche",
        "Notes",
        "Date création",
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    for a in qs:
        ws.append(
            [
                a.pk,
                a.matricule or "",
                a.nom or "",
                a.prenom or "",
                a.telephone or "",
                a.email or "",
                a.get_statut_display(),
                _format_excel_value(a.date_embauche),
                (a.notes or "")[:500],
                _format_excel_value(a.date_creation),
            ]
        )
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="agents_collecteurs.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_contribuables(request):
    """Export PDF des contribuables (avec filtres q, nationalite, date_du, date_au)."""
    q = request.GET.get("q", "").strip()
    nationalite = request.GET.get("nationalite", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    qs = Contribuable.objects.select_related("user").prefetch_related("boutiques_magasins").order_by("-date_creation")
    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(telephone__icontains=q)
        )
    if nationalite:
        qs = qs.filter(nationalite__icontains=nationalite)
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        qs = qs.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        qs = qs.filter(date_creation__date__lte=date_au_parsed)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="contribuables.pdf"'
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Contribuables (Marchés / Places publiques)", title_style))
    if q or nationalite or date_du or date_au:
        parts = []
        if q:
            parts.append(f"Recherche: {q}")
        if nationalite:
            parts.append(f"Nationalité: {nationalite}")
        if date_du:
            parts.append(f"Du: {date_du}")
        if date_au:
            parts.append(f"Au: {date_au}")
        story.append(Paragraph(" | ".join(parts), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénom", "Téléphone", "Nationalité", "Nb boutiques"]]
    for c in qs[:1000]:
        nb = c.boutiques_magasins.count()
        data.append([c.nom or "", c.prenom or "", c.telephone or "", c.nationalite or "", str(nb)])
    if len(data) > 1:
        data.append(["", "", "", "TOTAL", str(qs.count()) + " contribuable(s)"])
    col_widths = [4 * cm, 4 * cm, 4 * cm, 4 * cm, 2.5 * cm]
    table = Table(data, colWidths=col_widths)
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if len(data) > 1:
        table_style.extend([
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
        ])
    table.setStyle(TableStyle(table_style))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_contribuables(request):
    """Export Excel des contribuables (avec filtres q, nationalite, date_du, date_au)."""
    q = request.GET.get("q", "").strip()
    nationalite = request.GET.get("nationalite", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    qs = Contribuable.objects.select_related("user").prefetch_related("boutiques_magasins").order_by("-date_creation")
    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(telephone__icontains=q)
        )
    if nationalite:
        qs = qs.filter(nationalite__icontains=nationalite)
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        qs = qs.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        qs = qs.filter(date_creation__date__lte=date_au_parsed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contribuables"
    headers = [
        "ID",
        "Nom",
        "Prénom",
        "Téléphone",
        "Date naissance",
        "Lieu naissance",
        "Nationalité",
        "Nb boutiques",
        "Date création",
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    for c in qs:
        ws.append(
            [
                c.pk,
                c.nom or "",
                c.prenom or "",
                c.telephone or "",
                _format_excel_value(c.date_naissance),
                c.lieu_naissance or "",
                c.nationalite or "",
                c.boutiques_magasins.count(),
                _format_excel_value(c.date_creation),
            ]
        )
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contribuables.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_boutiques(request):
    """Export PDF des boutiques / magasins (filtres q, contribuable, agent_collecteur, date_du, date_au)."""
    q = request.GET.get("q", "").strip()
    contribuable_id = request.GET.get("contribuable", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    qs = BoutiqueMagasin.objects.select_related(
        "contribuable", "emplacement", "agent_collecteur"
    ).order_by("-id")
    if q:
        qs = qs.filter(
            Q(matricule__icontains=q)
            | Q(contribuable__nom__icontains=q)
            | Q(contribuable__prenom__icontains=q)
            | Q(emplacement__nom_lieu__icontains=q)
        )
    if contribuable_id:
        try:
            qs = qs.filter(contribuable_id=int(contribuable_id))
        except (ValueError, TypeError):
            pass
    if agent_collecteur_id:
        try:
            qs = qs.filter(agent_collecteur_id=int(agent_collecteur_id))
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        qs = qs.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        qs = qs.filter(date_creation__date__lte=date_au_parsed)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="boutiques_magasins.pdf"'
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Boutiques / Magasins (marchés)", title_style))
    if q or contribuable_id or agent_collecteur_id or date_du or date_au:
        parts = []
        if q:
            parts.append(f"Recherche: {q}")
        if contribuable_id:
            parts.append(f"Contribuable ID: {contribuable_id}")
        if agent_collecteur_id:
            parts.append(f"Agent collecteur ID: {agent_collecteur_id}")
        if date_du:
            parts.append(f"Du: {date_du}")
        if date_au:
            parts.append(f"Au: {date_au}")
        story.append(Paragraph(" | ".join(parts), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    agg_bout = qs.aggregate(
        total_loyer=Sum("prix_location_mensuel"),
        total_superficie=Sum("superficie_m2"),
    )
    total_loyer_bout = agg_bout.get("total_loyer") or Decimal("0")
    total_superficie_bout = agg_bout.get("total_superficie") or Decimal("0")
    data = [
        [
            "Matricule",
            "Emplacement",
            "Type",
            "Contribuable",
            "Superficie (m²)",
            "Loyer mensuel",
            "Activité",
            "Agent",
        ]
    ]
    for b in qs[:1000]:
        contrib = b.contribuable.nom_complet if b.contribuable else "—"
        agent = f"{b.agent_collecteur.nom} {b.agent_collecteur.prenom}" if b.agent_collecteur else "—"
        data.append(
            [
                b.matricule or "",
                (b.emplacement.nom_lieu if b.emplacement else "")[:20],
                b.get_type_local_display(),
                contrib[:25],
                str(b.superficie_m2) if b.superficie_m2 else "",
                str(b.prix_location_mensuel) if b.prix_location_mensuel else "",
                (b.activite_vendue or "")[:25],
                agent[:20],
            ]
        )
    if len(data) > 1:
        data.append([
            f"TOTAL ({qs.count()} boutiques)",
            "",
            "",
            "",
            str(total_superficie_bout),
            str(total_loyer_bout) + " FCFA",
            "",
            "",
        ])
    col_widths = [3 * cm, 4 * cm, 2.5 * cm, 5 * cm, 2 * cm, 2.5 * cm, 4 * cm, 4 * cm]
    table = Table(data, colWidths=col_widths)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
        ])
    )
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_boutiques(request):
    """Export Excel des boutiques / magasins (filtres q, contribuable, agent_collecteur, date_du, date_au)."""
    q = request.GET.get("q", "").strip()
    contribuable_id = request.GET.get("contribuable", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    qs = BoutiqueMagasin.objects.select_related(
        "contribuable", "emplacement", "agent_collecteur"
    ).order_by("-id")
    if q:
        qs = qs.filter(
            Q(matricule__icontains=q)
            | Q(contribuable__nom__icontains=q)
            | Q(contribuable__prenom__icontains=q)
            | Q(emplacement__nom_lieu__icontains=q)
        )
    if contribuable_id:
        try:
            qs = qs.filter(contribuable_id=int(contribuable_id))
        except (ValueError, TypeError):
            pass
    if agent_collecteur_id:
        try:
            qs = qs.filter(agent_collecteur_id=int(agent_collecteur_id))
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        qs = qs.filter(date_creation__date__gte=date_du_parsed)
    if date_au_parsed:
        qs = qs.filter(date_creation__date__lte=date_au_parsed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Boutiques Magasins"
    headers = [
        "ID",
        "Matricule",
        "Emplacement",
        "Type local",
        "Superficie (m²)",
        "Loyer mensuel",
        "Loyer annuel",
        "Contribuable",
        "Activité",
        "Agent collecteur",
        "Actif",
        "Date création",
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    for b in qs:
        contrib = b.contribuable.nom_complet if b.contribuable else ""
        agent = f"{b.agent_collecteur.nom} {b.agent_collecteur.prenom}" if b.agent_collecteur else ""
        ws.append(
            [
                b.pk,
                b.matricule or "",
                b.emplacement.nom_lieu if b.emplacement else "",
                b.get_type_local_display(),
                b.superficie_m2,
                b.prix_location_mensuel,
                b.prix_location_annuel or "",
                contrib,
                b.activite_vendue or "",
                agent,
                "Oui" if b.est_actif else "Non",
                _format_excel_value(b.date_creation) if hasattr(b, "date_creation") else "",
            ]
        )
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="boutiques_magasins.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_contributions(request):
    """Export PDF des contributions (filtres type, annee, mois, agent_collecteur, date_du, date_au, q)."""
    type_contribution = request.GET.get("type", "").strip()
    annee = request.GET.get("annee", "").strip()
    mois = request.GET.get("mois", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    q = request.GET.get("q", "").strip()
    cotisations = CotisationAnnuelle.objects.select_related(
        "boutique__contribuable", "boutique__emplacement"
    ).order_by("-annee", "-date_creation")
    paiements = PaiementCotisation.objects.select_related(
        "cotisation_annuelle__boutique__contribuable",
        "encaisse_par_agent",
    ).order_by("-date_paiement")
    tickets = TicketMarche.objects.select_related(
        "emplacement", "contribuable", "encaisse_par_agent"
    ).order_by("-date", "-date_creation")
    if type_contribution == "paiements":
        cotisations = cotisations.none()
        tickets = tickets.none()
    elif type_contribution == "tickets":
        cotisations = cotisations.none()
        paiements = paiements.none()
    elif type_contribution == "cotisations":
        paiements = paiements.none()
        tickets = tickets.none()
    if annee:
        try:
            annee_int = int(annee)
            cotisations = cotisations.filter(annee=annee_int)
            paiements = paiements.filter(cotisation_annuelle__annee=annee_int)
            tickets = tickets.filter(date__year=annee_int)
        except ValueError:
            pass
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements = paiements.filter(mois=mois_int)
                tickets = tickets.filter(date__month=mois_int)
        except ValueError:
            pass
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements = paiements.filter(encaisse_par_agent_id=agent_id)
            tickets = tickets.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements = paiements.filter(date_paiement__date__gte=date_du_parsed)
        tickets = tickets.filter(date__gte=date_du_parsed)
    if date_au_parsed:
        paiements = paiements.filter(date_paiement__date__lte=date_au_parsed)
        tickets = tickets.filter(date__lte=date_au_parsed)
    if q:
        cotisations = cotisations.filter(
            Q(boutique__matricule__icontains=q)
            | Q(boutique__contribuable__nom__icontains=q)
            | Q(boutique__contribuable__prenom__icontains=q)
        )
        paiements = paiements.filter(
            Q(cotisation_annuelle__boutique__matricule__icontains=q)
            | Q(cotisation_annuelle__boutique__contribuable__nom__icontains=q)
            | Q(cotisation_annuelle__boutique__contribuable__prenom__icontains=q)
        )
        tickets = tickets.filter(
            Q(nom_vendeur__icontains=q)
            | Q(contribuable__nom__icontains=q)
            | Q(contribuable__prenom__icontains=q)
        )
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="contributions.pdf"'
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Contributions / Taxes (marchés)", title_style))
    parts_filtres = []
    if type_contribution:
        type_lib = {"cotisations": "Cotisations", "paiements": "Paiements", "tickets": "Tickets"}.get(
            type_contribution, type_contribution
        )
        parts_filtres.append(f"Type: {type_lib}")
    if annee:
        parts_filtres.append(f"Année: {annee}")
    if mois:
        mois_noms = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        try:
            m = int(mois)
            if 1 <= m <= 12:
                parts_filtres.append(f"Mois: {mois_noms[m]}")
        except ValueError:
            parts_filtres.append(f"Mois: {mois}")
    if agent_collecteur_id:
        parts_filtres.append(f"Agent: {agent_collecteur_id}")
    if date_du:
        parts_filtres.append(f"Du: {date_du}")
    if date_au:
        parts_filtres.append(f"Au: {date_au}")
    if q:
        parts_filtres.append(f"Recherche: {q}")
    if parts_filtres:
        story.append(Paragraph(" | ".join(parts_filtres), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data_cot, data_pay, data_tick = [[""]], [[""]], [[""]]
    total_du_cot = total_paye_cot = total_reste_cot = Decimal("0")
    total_recettes_pay = total_recettes_tick = Decimal("0")
    if not type_contribution or type_contribution == "cotisations":
        data_cot = [["Boutique", "Année", "Montant dû", "Montant payé", "Reste"]]
        total_du_cot = Decimal("0")
        total_paye_cot = Decimal("0")
        total_reste_cot = Decimal("0")
        for c in cotisations[:500]:
            mp = c.montant_paye() if hasattr(c, "montant_paye") and callable(c.montant_paye) else Decimal("0")
            try:
                reste_val = c.montant_annuel_du - mp
            except Exception:
                reste_val = Decimal("0")
            total_du_cot += c.montant_annuel_du
            total_paye_cot += mp
            total_reste_cot += reste_val
            data_cot.append(
                [
                    (c.boutique.matricule if c.boutique else "")[:15],
                    str(c.annee),
                    str(c.montant_annuel_du),
                    str(mp),
                    str(reste_val),
                ]
            )
        if len(data_cot) > 1:
            data_cot.append(["TOTAL", "", str(total_du_cot), str(total_paye_cot), str(total_reste_cot)])
            table_cot = Table(data_cot, colWidths=[3 * cm, 1.5 * cm, 3 * cm, 3 * cm, 3 * cm])
            table_cot.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                    ]
                )
            )
            story.append(Paragraph("Cotisations annuelles", styles["Heading2"]))
            story.append(table_cot)
            story.append(Spacer(1, 0.3 * cm))
    if not type_contribution or type_contribution == "paiements":
        data_pay = [["Boutique", "Année", "Montant", "Date", "Agent"]]
        total_recettes_pay = Decimal("0")
        for p in paiements[:500]:
            montant_val = Decimal(str(p.montant_paye)) if p.montant_paye is not None else Decimal("0")
            total_recettes_pay += montant_val
            data_pay.append(
                [
                    (p.cotisation_annuelle.boutique.matricule if p.cotisation_annuelle and p.cotisation_annuelle.boutique else "")[:15],
                    str(getattr(p.cotisation_annuelle, "annee", "")),
                    str(p.montant_paye),
                    p.date_paiement.strftime("%d/%m/%Y") if hasattr(p.date_paiement, "strftime") else "",
                    p.encaisse_par_agent.nom_complet if p.encaisse_par_agent else "—",
                ]
            )
        if len(data_pay) > 1:
            data_pay.append(["TOTAL RECETTES", "", str(total_recettes_pay), "", ""])
            table_pay = Table(data_pay, colWidths=[3 * cm, 1.5 * cm, 3 * cm, 3 * cm, 5 * cm])
            table_pay.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                    ]
                )
            )
            story.append(Paragraph("Paiements", styles["Heading2"]))
            story.append(table_pay)
            story.append(Spacer(1, 0.3 * cm))
    if not type_contribution or type_contribution == "tickets":
        data_tick = [["Emplacement", "Vendeur", "Montant", "Date", "Agent"]]
        total_recettes_tick = Decimal("0")
        for t in tickets[:500]:
            montant_val = Decimal(str(t.montant)) if t.montant is not None else Decimal("0")
            total_recettes_tick += montant_val
            data_tick.append(
                [
                    (t.emplacement.nom_lieu if t.emplacement else "")[:15],
                    (t.nom_vendeur or (t.contribuable.nom_complet if t.contribuable else ""))[:20],
                    str(t.montant),
                    t.date.strftime("%d/%m/%Y") if hasattr(t.date, "strftime") else "",
                    t.encaisse_par_agent.nom_complet if t.encaisse_par_agent else "—",
                ]
            )
        if len(data_tick) > 1:
            data_tick.append(["TOTAL RECETTES (tickets marché)", "", str(total_recettes_tick), "", ""])
            table_tick = Table(data_tick, colWidths=[4 * cm, 5 * cm, 3 * cm, 3 * cm, 5 * cm])
            table_tick.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                    ]
                )
            )
            story.append(Paragraph("Tickets marché", styles["Heading2"]))
            story.append(table_tick)
    # Récapitulatif global des totaux
    recap_parts = []
    if (not type_contribution or type_contribution == "cotisations") and len(data_cot) > 1:
        recap_parts.append(f"Cotisations: Dû {total_du_cot} | Payé {total_paye_cot} | Reste à payer {total_reste_cot} FCFA")
    if (not type_contribution or type_contribution == "paiements") and len(data_pay) > 1:
        recap_parts.append(f"Recettes paiements: {total_recettes_pay} FCFA")
    if (not type_contribution or type_contribution == "tickets") and len(data_tick) > 1:
        recap_parts.append(f"Recettes tickets: {total_recettes_tick} FCFA")
    if recap_parts:
        story.append(Spacer(1, 0.4 * cm))
        recap_style = ParagraphStyle(
            "Recap",
            parent=styles["Normal"],
            fontSize=11,
            textColor=colors.HexColor("#006233"),
            fontName="Helvetica-Bold",
            spaceBefore=6,
            spaceAfter=6,
        )
        story.append(Paragraph("RÉCAPITULATIF DES TOTAUX", recap_style))
        for part in recap_parts:
            story.append(Paragraph(f"• {part}", styles["Normal"]))
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_contributions(request):
    """Export Excel des contributions (filtres type, annee, mois, agent_collecteur, date_du, date_au, q)."""
    type_contribution = request.GET.get("type", "").strip()
    annee = request.GET.get("annee", "").strip()
    mois = request.GET.get("mois", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    q = request.GET.get("q", "").strip()
    cotisations = CotisationAnnuelle.objects.select_related(
        "boutique__contribuable", "boutique__emplacement"
    ).order_by("-annee", "-date_creation")
    paiements = PaiementCotisation.objects.select_related(
        "cotisation_annuelle__boutique__contribuable",
        "encaisse_par_agent",
    ).order_by("-date_paiement")
    tickets = TicketMarche.objects.select_related(
        "emplacement", "contribuable", "encaisse_par_agent"
    ).order_by("-date", "-date_creation")
    if type_contribution == "paiements":
        cotisations = cotisations.none()
        tickets = tickets.none()
    elif type_contribution == "tickets":
        cotisations = cotisations.none()
        paiements = paiements.none()
    elif type_contribution == "cotisations":
        paiements = paiements.none()
        tickets = tickets.none()
    if annee:
        try:
            annee_int = int(annee)
            cotisations = cotisations.filter(annee=annee_int)
            paiements = paiements.filter(cotisation_annuelle__annee=annee_int)
            tickets = tickets.filter(date__year=annee_int)
        except ValueError:
            pass
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements = paiements.filter(mois=mois_int)
                tickets = tickets.filter(date__month=mois_int)
        except ValueError:
            pass
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements = paiements.filter(encaisse_par_agent_id=agent_id)
            tickets = tickets.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements = paiements.filter(date_paiement__date__gte=date_du_parsed)
        tickets = tickets.filter(date__gte=date_du_parsed)
    if date_au_parsed:
        paiements = paiements.filter(date_paiement__date__lte=date_au_parsed)
        tickets = tickets.filter(date__lte=date_au_parsed)
    if q:
        cotisations = cotisations.filter(
            Q(boutique__matricule__icontains=q)
            | Q(boutique__contribuable__nom__icontains=q)
            | Q(boutique__contribuable__prenom__icontains=q)
        )
        paiements = paiements.filter(
            Q(cotisation_annuelle__boutique__matricule__icontains=q)
            | Q(cotisation_annuelle__boutique__contribuable__nom__icontains=q)
            | Q(cotisation_annuelle__boutique__contribuable__prenom__icontains=q)
        )
        tickets = tickets.filter(
            Q(nom_vendeur__icontains=q)
            | Q(contribuable__nom__icontains=q)
            | Q(contribuable__prenom__icontains=q)
        )
    wb = Workbook()
    # Feuille Cotisations
    ws_cot = wb.active
    ws_cot.title = "Cotisations"
    h_cot = ["Boutique", "Emplacement", "Contribuable", "Année", "Montant dû", "Montant payé", "Reste"]
    ws_cot.append(h_cot)
    _style_excel_header(ws_cot, 1)
    for c in cotisations:
        mp = c.montant_paye() if callable(c.montant_paye) else getattr(c, "montant_paye", 0)
        reste = c.montant_annuel_du - mp
        contrib = c.boutique.contribuable.nom_complet if c.boutique and c.boutique.contribuable else ""
        ws_cot.append(
            [
                c.boutique.matricule if c.boutique else "",
                c.boutique.emplacement.nom_lieu if c.boutique and c.boutique.emplacement else "",
                contrib,
                c.annee,
                c.montant_annuel_du,
                mp,
                reste,
            ]
        )
    # Feuille Paiements
    ws_pay = wb.create_sheet("Paiements")
    h_pay = ["Boutique", "Année", "Montant", "Date paiement", "Agent"]
    ws_pay.append(h_pay)
    _style_excel_header(ws_pay, 1)
    for p in paiements:
        ws_pay.append(
            [
                p.cotisation_annuelle.boutique.matricule if p.cotisation_annuelle and p.cotisation_annuelle.boutique else "",
                getattr(p.cotisation_annuelle, "annee", ""),
                p.montant_paye,
                _format_excel_value(p.date_paiement),
                p.encaisse_par_agent.nom_complet if p.encaisse_par_agent else "",
            ]
        )
    # Feuille Tickets
    ws_tick = wb.create_sheet("Tickets")
    h_tick = ["Emplacement", "Vendeur", "Contribuable", "Montant", "Date", "Agent"]
    ws_tick.append(h_tick)
    _style_excel_header(ws_tick, 1)
    for t in tickets:
        ws_tick.append(
            [
                t.emplacement.nom_lieu if t.emplacement else "",
                t.nom_vendeur or "",
                t.contribuable.nom_complet if t.contribuable else "",
                t.montant,
                _format_excel_value(t.date),
                t.encaisse_par_agent.nom_complet if t.encaisse_par_agent else "",
            ]
        )
    for sheet in [ws_cot, ws_pay, ws_tick]:
        for idx in range(1, 15):
            sheet.column_dimensions[get_column_letter(idx)].width = 18
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="contributions.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_cotisations_acteurs_institutions(request):
    """Export PDF des cotisations acteurs / institutions (filtres type, annee, mois, agent, date_du, date_au, q)."""
    type_contribution = request.GET.get("type", "").strip()
    annee = request.GET.get("annee", "").strip()
    mois = request.GET.get("mois", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    q = request.GET.get("q", "").strip()
    cot_acteurs = CotisationAnnuelleActeur.objects.select_related("acteur").order_by("-annee", "-date_creation")
    cot_inst = CotisationAnnuelleInstitution.objects.select_related("institution").order_by("-annee", "-date_creation")
    paiements_acteurs = PaiementCotisationActeur.objects.select_related(
        "cotisation_annuelle__acteur", "encaisse_par_agent"
    ).order_by("-date_paiement")
    paiements_inst = PaiementCotisationInstitution.objects.select_related(
        "cotisation_annuelle__institution", "encaisse_par_agent"
    ).order_by("-date_paiement")
    if type_contribution == "institutions":
        cot_acteurs = cot_acteurs.none()
        paiements_acteurs = paiements_acteurs.none()
    elif type_contribution == "acteurs":
        cot_inst = cot_inst.none()
        paiements_inst = paiements_inst.none()
    if annee:
        try:
            annee_int = int(annee)
            cot_acteurs = cot_acteurs.filter(annee=annee_int)
            cot_inst = cot_inst.filter(annee=annee_int)
            paiements_acteurs = paiements_acteurs.filter(cotisation_annuelle__annee=annee_int)
            paiements_inst = paiements_inst.filter(cotisation_annuelle__annee=annee_int)
        except ValueError:
            pass
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements_acteurs = paiements_acteurs.filter(date_paiement__month=mois_int)
                paiements_inst = paiements_inst.filter(date_paiement__month=mois_int)
        except ValueError:
            pass
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements_acteurs = paiements_acteurs.filter(encaisse_par_agent_id=agent_id)
            paiements_inst = paiements_inst.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__gte=date_du_parsed)
        paiements_inst = paiements_inst.filter(date_paiement__date__gte=date_du_parsed)
    if date_au_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__lte=date_au_parsed)
        paiements_inst = paiements_inst.filter(date_paiement__date__lte=date_au_parsed)
    if q:
        cot_acteurs = cot_acteurs.filter(
            Q(acteur__raison_sociale__icontains=q)
            | Q(acteur__sigle__icontains=q)
            | Q(acteur__nom_responsable__icontains=q)
        )
        cot_inst = cot_inst.filter(
            Q(institution__nom_institution__icontains=q)
            | Q(institution__sigle__icontains=q)
            | Q(institution__nom_responsable__icontains=q)
        )
        paiements_acteurs = paiements_acteurs.filter(
            Q(cotisation_annuelle__acteur__raison_sociale__icontains=q)
            | Q(cotisation_annuelle__acteur__sigle__icontains=q)
        )
        paiements_inst = paiements_inst.filter(
            Q(cotisation_annuelle__institution__nom_institution__icontains=q)
            | Q(cotisation_annuelle__institution__sigle__icontains=q)
        )
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="cotisations_acteurs_institutions.pdf"'
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        topMargin=PDF_HEADER_HEIGHT_CM * cm,
        bottomMargin=1.5 * cm,
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Cotisations Acteurs & Institutions", title_style))
    parts_filtres = []
    if type_contribution:
        parts_filtres.append(f"Type: {'Acteurs économiques' if type_contribution == 'acteurs' else 'Institutions financières'}")
    if annee:
        parts_filtres.append(f"Année: {annee}")
    if mois:
        mois_noms = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        try:
            m = int(mois)
            if 1 <= m <= 12:
                parts_filtres.append(f"Mois: {mois_noms[m]}")
        except ValueError:
            parts_filtres.append(f"Mois: {mois}")
    if agent_collecteur_id:
        parts_filtres.append(f"Agent: {agent_collecteur_id}")
    if date_du:
        parts_filtres.append(f"Du: {date_du}")
    if date_au:
        parts_filtres.append(f"Au: {date_au}")
    if q:
        parts_filtres.append(f"Recherche: {q}")
    if parts_filtres:
        story.append(Paragraph(" | ".join(parts_filtres), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    if not type_contribution or type_contribution == "acteurs":
        data_act = [["Acteur", "Montant dû", "Montant payé", "Reste"]]
        total_du_act = total_paye_act = total_reste_act = Decimal("0")
        for c in cot_acteurs[:500]:
            mp = c.montant_paye() if callable(c.montant_paye) else Decimal("0")
            reste = c.montant_annuel_du - mp
            total_du_act += c.montant_annuel_du
            total_paye_act += mp
            total_reste_act += reste
            data_act.append(
                [
                    (c.acteur.raison_sociale if c.acteur else "")[:30],
                    str(c.montant_annuel_du),
                    str(mp),
                    str(reste),
                ]
            )
        if len(data_act) > 1:
            data_act.append(["TOTAL", str(total_du_act), str(total_paye_act), str(total_reste_act)])
            table_act = Table(data_act, colWidths=[8 * cm, 3 * cm, 3 * cm, 3 * cm])
            table_act.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                ])
            )
            story.append(Paragraph("Acteurs économiques", styles["Heading2"]))
            story.append(table_act)
            story.append(Spacer(1, 0.3 * cm))
    if not type_contribution or type_contribution == "institutions":
        data_inst = [["Institution", "Montant dû", "Montant payé", "Reste"]]
        total_du_inst = total_paye_inst = total_reste_inst = Decimal("0")
        for c in cot_inst[:500]:
            mp = c.montant_paye() if callable(c.montant_paye) else Decimal("0")
            reste = c.montant_annuel_du - mp
            total_du_inst += c.montant_annuel_du
            total_paye_inst += mp
            total_reste_inst += reste
            data_inst.append(
                [
                    (c.institution.nom_institution if c.institution else "")[:30],
                    str(c.montant_annuel_du),
                    str(mp),
                    str(reste),
                ]
            )
        if len(data_inst) > 1:
            data_inst.append(["TOTAL", str(total_du_inst), str(total_paye_inst), str(total_reste_inst)])
            table_inst = Table(data_inst, colWidths=[8 * cm, 3 * cm, 3 * cm, 3 * cm])
            table_inst.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                ])
            )
            story.append(Paragraph("Institutions financières", styles["Heading2"]))
            story.append(table_inst)
    # Paiements acteurs
    if not type_contribution or type_contribution == "acteurs":
        agg_pay_act = paiements_acteurs.aggregate(total=Sum("montant_paye"))
        total_recettes_pay_act = agg_pay_act.get("total") or Decimal("0")
        data_pay_act = [["Acteur", "Année", "Montant", "Date paiement", "Agent"]]
        for p in paiements_acteurs[:300]:
            data_pay_act.append(
                [
                    (p.cotisation_annuelle.acteur.raison_sociale if p.cotisation_annuelle and p.cotisation_annuelle.acteur else "")[:25],
                    str(getattr(p.cotisation_annuelle, "annee", "")),
                    str(p.montant_paye),
                    p.date_paiement.strftime("%d/%m/%Y") if hasattr(p.date_paiement, "strftime") else "",
                    (p.encaisse_par_agent.nom_complet or "—")[:20] if p.encaisse_par_agent else "—",
                ]
            )
        if len(data_pay_act) > 1:
            data_pay_act.append(["TOTAL RECETTES", "", str(total_recettes_pay_act), "", ""])
            story.append(Paragraph("Paiements Acteurs économiques", styles["Heading2"]))
            table_pay_act = Table(data_pay_act, colWidths=[5 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 4 * cm])
            table_pay_act.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                ])
            )
            story.append(table_pay_act)
            story.append(Spacer(1, 0.3 * cm))
    # Paiements institutions
    if not type_contribution or type_contribution == "institutions":
        agg_pay_inst = paiements_inst.aggregate(total=Sum("montant_paye"))
        total_recettes_pay_inst = agg_pay_inst.get("total") or Decimal("0")
        data_pay_inst = [["Institution", "Année", "Montant", "Date paiement", "Agent"]]
        for p in paiements_inst[:300]:
            data_pay_inst.append(
                [
                    (p.cotisation_annuelle.institution.nom_institution if p.cotisation_annuelle and p.cotisation_annuelle.institution else "")[:25],
                    str(getattr(p.cotisation_annuelle, "annee", "")),
                    str(p.montant_paye),
                    p.date_paiement.strftime("%d/%m/%Y") if hasattr(p.date_paiement, "strftime") else "",
                    (p.encaisse_par_agent.nom_complet or "—")[:20] if p.encaisse_par_agent else "—",
                ]
            )
        if len(data_pay_inst) > 1:
            data_pay_inst.append(["TOTAL RECETTES", "", str(total_recettes_pay_inst), "", ""])
            story.append(Paragraph("Paiements Institutions financières", styles["Heading2"]))
            table_pay_inst = Table(data_pay_inst, colWidths=[5 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 4 * cm])
            table_pay_inst.setStyle(
                TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
                ])
            )
            story.append(table_pay_inst)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_cotisations_acteurs_institutions(request):
    """Export Excel des cotisations acteurs / institutions (filtres type, annee, mois, agent, date_du, date_au, q)."""
    type_contribution = request.GET.get("type", "").strip()
    annee = request.GET.get("annee", "").strip()
    mois = request.GET.get("mois", "").strip()
    agent_collecteur_id = request.GET.get("agent_collecteur", "").strip()
    date_du = request.GET.get("date_du", "").strip()
    date_au = request.GET.get("date_au", "").strip()
    q = request.GET.get("q", "").strip()
    cot_acteurs = CotisationAnnuelleActeur.objects.select_related("acteur").order_by("-annee", "-date_creation")
    cot_inst = CotisationAnnuelleInstitution.objects.select_related("institution").order_by("-annee", "-date_creation")
    paiements_acteurs = PaiementCotisationActeur.objects.select_related(
        "cotisation_annuelle__acteur", "encaisse_par_agent"
    ).order_by("-date_paiement")
    paiements_inst = PaiementCotisationInstitution.objects.select_related(
        "cotisation_annuelle__institution", "encaisse_par_agent"
    ).order_by("-date_paiement")
    if type_contribution == "institutions":
        cot_acteurs = cot_acteurs.none()
        paiements_acteurs = paiements_acteurs.none()
    elif type_contribution == "acteurs":
        cot_inst = cot_inst.none()
        paiements_inst = paiements_inst.none()
    if annee:
        try:
            annee_int = int(annee)
            cot_acteurs = cot_acteurs.filter(annee=annee_int)
            cot_inst = cot_inst.filter(annee=annee_int)
            paiements_acteurs = paiements_acteurs.filter(cotisation_annuelle__annee=annee_int)
            paiements_inst = paiements_inst.filter(cotisation_annuelle__annee=annee_int)
        except ValueError:
            pass
    if mois:
        try:
            mois_int = int(mois)
            if 1 <= mois_int <= 12:
                paiements_acteurs = paiements_acteurs.filter(date_paiement__month=mois_int)
                paiements_inst = paiements_inst.filter(date_paiement__month=mois_int)
        except ValueError:
            pass
    if agent_collecteur_id:
        try:
            agent_id = int(agent_collecteur_id)
            paiements_acteurs = paiements_acteurs.filter(encaisse_par_agent_id=agent_id)
            paiements_inst = paiements_inst.filter(encaisse_par_agent_id=agent_id)
        except (ValueError, TypeError):
            pass
    date_du_parsed = _parse_date(date_du)
    date_au_parsed = _parse_date(date_au)
    if date_du_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__gte=date_du_parsed)
        paiements_inst = paiements_inst.filter(date_paiement__date__gte=date_du_parsed)
    if date_au_parsed:
        paiements_acteurs = paiements_acteurs.filter(date_paiement__date__lte=date_au_parsed)
        paiements_inst = paiements_inst.filter(date_paiement__date__lte=date_au_parsed)
    if q:
        cot_acteurs = cot_acteurs.filter(
            Q(acteur__raison_sociale__icontains=q)
            | Q(acteur__sigle__icontains=q)
            | Q(acteur__nom_responsable__icontains=q)
        )
        cot_inst = cot_inst.filter(
            Q(institution__nom_institution__icontains=q)
            | Q(institution__sigle__icontains=q)
            | Q(institution__nom_responsable__icontains=q)
        )
        paiements_acteurs = paiements_acteurs.filter(
            Q(cotisation_annuelle__acteur__raison_sociale__icontains=q)
            | Q(cotisation_annuelle__acteur__sigle__icontains=q)
        )
        paiements_inst = paiements_inst.filter(
            Q(cotisation_annuelle__institution__nom_institution__icontains=q)
            | Q(cotisation_annuelle__institution__sigle__icontains=q)
        )
    wb = Workbook()
    ws_act = wb.active
    ws_act.title = "Cotisations Acteurs"
    h_act = ["Acteur", "Sigle", "Année", "Montant dû", "Montant payé", "Reste"]
    ws_act.append(h_act)
    _style_excel_header(ws_act, 1)
    for c in cot_acteurs:
        mp = c.montant_paye() if callable(c.montant_paye) else Decimal("0")
        reste = c.montant_annuel_du - mp
        ws_act.append(
            [
                c.acteur.raison_sociale if c.acteur else "",
                c.acteur.sigle if c.acteur else "",
                c.annee,
                c.montant_annuel_du,
                mp,
                reste,
            ]
        )
    ws_inst = wb.create_sheet("Cotisations Institutions")
    h_inst = ["Institution", "Sigle", "Année", "Montant dû", "Montant payé", "Reste"]
    ws_inst.append(h_inst)
    _style_excel_header(ws_inst, 1)
    for c in cot_inst:
        mp = c.montant_paye() if callable(c.montant_paye) else Decimal("0")
        reste = c.montant_annuel_du - mp
        ws_inst.append(
            [
                c.institution.nom_institution if c.institution else "",
                c.institution.sigle if c.institution else "",
                c.annee,
                c.montant_annuel_du,
                mp,
                reste,
            ]
        )
    ws_pay_act = wb.create_sheet("Paiements Acteurs")
    h_pay_act = ["Acteur", "Année", "Montant", "Date paiement", "Agent"]
    ws_pay_act.append(h_pay_act)
    _style_excel_header(ws_pay_act, 1)
    for p in paiements_acteurs:
        ws_pay_act.append(
            [
                p.cotisation_annuelle.acteur.raison_sociale if p.cotisation_annuelle and p.cotisation_annuelle.acteur else "",
                getattr(p.cotisation_annuelle, "annee", ""),
                p.montant_paye,
                _format_excel_value(p.date_paiement),
                p.encaisse_par_agent.nom_complet if p.encaisse_par_agent else "",
            ]
        )
    ws_pay_inst = wb.create_sheet("Paiements Institutions")
    h_pay_inst = ["Institution", "Année", "Montant", "Date paiement", "Agent"]
    ws_pay_inst.append(h_pay_inst)
    _style_excel_header(ws_pay_inst, 1)
    for p in paiements_inst:
        ws_pay_inst.append(
            [
                p.cotisation_annuelle.institution.nom_institution if p.cotisation_annuelle and p.cotisation_annuelle.institution else "",
                getattr(p.cotisation_annuelle, "annee", ""),
                p.montant_paye,
                _format_excel_value(p.date_paiement),
                p.encaisse_par_agent.nom_complet if p.encaisse_par_agent else "",
            ]
        )
    for sheet in [ws_act, ws_inst, ws_pay_act, ws_pay_inst]:
        for idx in range(1, 10):
            sheet.column_dimensions[get_column_letter(idx)].width = 20
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cotisations_acteurs_institutions.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_candidatures(request):
    """Exporte toutes les candidatures en Excel avec tous les champs."""
    candidatures = Candidature.objects.all().select_related('appel_offre', 'candidat').order_by('-date_soumission')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    if q:
        candidatures = candidatures.filter(
            Q(appel_offre__titre__icontains=q) |
            Q(appel_offre__reference__icontains=q) |
            Q(candidat__first_name__icontains=q) |
            Q(candidat__last_name__icontains=q) |
            Q(candidat__email__icontains=q)
        )
    if statut:
        candidatures = candidatures.filter(statut=statut)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    
    # En-têtes
    headers = [
        "ID", "Appel d'offres - Titre", "Appel d'offres - Référence", "Appel d'offres - Description",
        "Appel d'offres - Public cible", "Appel d'offres - Date début", "Appel d'offres - Date fin",
        "Appel d'offres - Budget estimé", "Candidat - Username", "Candidat - Nom", "Candidat - Prénom",
        "Candidat - Email", "Statut candidature", "Message accompagnement", "Date soumission"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for candidature in candidatures:
        row = [
            candidature.pk,
            candidature.appel_offre.titre,
            candidature.appel_offre.reference or "",
            candidature.appel_offre.description,
            candidature.appel_offre.get_public_cible_display(),
            _format_excel_value(candidature.appel_offre.date_debut),
            _format_excel_value(candidature.appel_offre.date_fin),
            candidature.appel_offre.budget_estime if candidature.appel_offre.budget_estime else "",
            candidature.candidat.username,
            candidature.candidat.last_name or "",
            candidature.candidat.first_name or "",
            candidature.candidat.email,
            candidature.get_statut_display(),
            candidature.message_accompagnement or "",
            _format_excel_value(candidature.date_soumission),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidatures.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_entreprises(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    secteur = request.GET.get("secteur") or ""

    # Uniquement les entreprises validées
    qs = ActeurEconomique.objects.filter(type_acteur="entreprise", est_valide_par_mairie=True)
    if secteur:
        qs = qs.filter(secteur_activite=secteur)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="entreprises_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Entreprises", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Raison sociale", "Secteur", "Responsable", "Téléphone"]]
    for a in qs.order_by("-date_enregistrement")[:1000]:
        data.append([
            a.raison_sociale,
            a.get_secteur_activite_display(),
            a.nom_responsable,
            a.telephone1,
        ])
    table = Table(data, colWidths=[8*cm, 7*cm, 7*cm, 5*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_institutions(request):
    start = request.GET.get("start")
    end = request.GET.get("end")
    type_inst = request.GET.get("type") or ""

    # Uniquement les institutions validées
    qs = InstitutionFinanciere.objects.filter(est_valide_par_mairie=True)
    if type_inst:
        qs = qs.filter(type_institution=type_inst)

    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_enregistrement__date__lte=ed)
        except ValueError:
            pass

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="institutions_financieres_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#006233"),
        alignment=1,
        spaceAfter=12,
    )
    story.append(Paragraph("Institutions Financières Validées", title_style))
    if start or end:
        story.append(
            Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"])
        )
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom de l'institution", "Type", "Responsable", "Téléphone", "Quartier"]]
    for inst in qs.order_by("-date_enregistrement")[:1000]:
        data.append(
            [
                inst.nom_institution,
                inst.get_type_institution_display(),
                inst.nom_responsable,
                inst.telephone1,
                inst.quartier,
            ]
        )
    table = Table(data, colWidths=[8 * cm, 6 * cm, 6 * cm, 4 * cm, 5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F5E9")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(
        Paragraph(
            "Date et Signature : ________________________________", styles["Normal"]
        )
    )

    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_diaspora(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    pays = request.GET.get("pays") or ""

    # Uniquement les membres validés par la mairie
    qs = MembreDiaspora.objects.filter(est_valide_par_mairie=True)
    if pays:
        qs = qs.filter(pays_residence_actuelle__icontains=pays)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="diaspora_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Membres de la Diaspora", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Pays de résidence", "Ville", "Téléphone", "Email", "Profession"]]
    for m in qs.order_by("-date_inscription")[:1000]:
        data.append([
            m.nom,
            m.prenoms,
            m.pays_residence_actuelle[:25] if m.pays_residence_actuelle else "",
            m.ville_residence_actuelle[:20] if m.ville_residence_actuelle else "",
            m.telephone_whatsapp[:15] if m.telephone_whatsapp else "",
            m.email[:30] if m.email else "",
            m.profession_actuelle[:30] if m.profession_actuelle else "",
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3.5*cm, 5*cm, 4.5*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def notifications_candidats(request):
    """Liste des appels d'offres avec candidats acceptés pour envoyer des notifications."""
    
    # Récupérer tous les appels d'offres qui ont au moins une candidature acceptée
    appels_offres = AppelOffre.objects.filter(
        candidatures__statut='acceptee'
    ).distinct().annotate(
        nb_candidats_acceptes=Count('candidatures', filter=Q(candidatures__statut='acceptee'))
    ).order_by('-date_debut')
    
    context = {
        'appels_offres': appels_offres,
    }
    
    return render(request, "admin/notifications_candidats.html", context)


@login_required
@user_passes_test(is_staff_user)
def envoyer_notifications_candidats(request, appel_offre_id):
    """Affiche le formulaire et traite l'envoi de notifications aux candidats acceptés."""
    
    appel_offre = get_object_or_404(AppelOffre, pk=appel_offre_id)
    
    # Récupérer uniquement les candidats acceptés pour cet appel d'offres
    candidats_acceptes = Candidature.objects.filter(
        appel_offre=appel_offre,
        statut='acceptee'
    ).select_related('candidat')
    
    if not candidats_acceptes.exists():
        messages.warning(request, "Aucun candidat accepté trouvé pour cet appel d'offres.")
        return redirect('notifications_candidats')
    
    if request.method == 'POST':
        titre = request.POST.get('titre', '')
        message = request.POST.get('message', '')
        type_notification = request.POST.get('type', Notification.TYPE_INFO)
        
        if not titre or not message:
            messages.error(request, "Le titre et le message sont obligatoires.")
        else:
            # Créer une notification pour chaque candidat accepté
            notifications_creees = 0
            for candidature in candidats_acceptes:
                Notification.objects.create(
                    recipient=candidature.candidat,
                    title=titre,
                    message=message,
                    type=type_notification,
                    created_by=request.user,
                )
                notifications_creees += 1
            
            messages.success(
                request, 
                f"Notifications envoyées avec succès à {notifications_creees} candidat(s) accepté(s)."
            )
            return redirect('notifications_candidats')
    
    context = {
        'appel_offre': appel_offre,
        'candidats_acceptes': candidats_acceptes,
        'nb_candidats': candidats_acceptes.count(),
        'type_choices': Notification.TYPE_CHOICES,
    }
    
    return render(request, "admin/envoyer_notifications_candidats.html", context)


# ========== FONCTIONS D'EXPORT EXCEL ==========

def _format_excel_value(value):
    """Formate une valeur pour l'export Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else value.strftime("%d/%m/%Y")
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value if item)
    return str(value)


def _style_excel_header(ws, row_num):
    """Applique un style au header Excel."""
    header_fill = PatternFill(start_color="006233", end_color="006233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


@login_required
@user_passes_test(is_staff_user)
def export_excel_acteurs(request):
    """Exporte tous les acteurs économiques en Excel avec tous les champs."""
    acteurs = ActeurEconomique.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres comme dans la vue liste
    q = request.GET.get('q', '')
    type_acteur = request.GET.get('type', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        acteurs = acteurs.filter(
            Q(raison_sociale__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_acteur:
        acteurs = acteurs.filter(type_acteur=type_acteur)
    if secteur:
        acteurs = acteurs.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Acteurs Economiques"
    
    # En-têtes
    headers = [
        "ID", "Raison sociale", "Sigle", "Type d'acteur", "Secteur d'activité", "Statut juridique",
        "Description", "RCCM", "CFE", "N° Carte opérateur", "NIF", "Date de création",
        "Capital social (FCFA)", "Nom responsable", "Fonction responsable", "Téléphone 1",
        "Téléphone 2", "Email", "Site web", "Quartier", "Canton", "Adresse complète",
        "Situation", "Latitude", "Longitude", "Nombre d'employés", "Chiffre d'affaires",
        "Accepte publication", "Certifie informations", "Accepte conditions",
        "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for acteur in acteurs:
        row = [
            acteur.pk,
            acteur.raison_sociale,
            acteur.sigle or "",
            acteur.get_type_acteur_display(),
            acteur.get_secteur_activite_display(),
            acteur.get_statut_juridique_display(),
            acteur.description,
            acteur.rccm or "",
            acteur.cfe or "",
            acteur.numero_carte_operateur or "",
            acteur.nif or "",
            _format_excel_value(acteur.date_creation),
            acteur.capital_social if acteur.capital_social else "",
            acteur.nom_responsable,
            acteur.fonction_responsable,
            acteur.telephone1,
            acteur.telephone2 or "",
            acteur.email,
            acteur.site_web or "",
            acteur.quartier,
            acteur.canton or "",
            acteur.adresse_complete,
            acteur.get_situation_display(),
            acteur.latitude if acteur.latitude else "",
            acteur.longitude if acteur.longitude else "",
            acteur.get_nombre_employes_display() if acteur.nombre_employes else "",
            acteur.get_chiffre_affaires_display() if acteur.chiffre_affaires else "",
            "Oui" if acteur.accepte_public else "Non",
            "Oui" if acteur.certifie_information else "Non",
            "Oui" if acteur.accepte_conditions else "Non",
            "Oui" if acteur.est_valide_par_mairie else "Non",
            _format_excel_value(acteur.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="acteurs_economiques.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_institutions(request):
    """Exporte toutes les institutions financières en Excel avec tous les champs."""
    institutions = InstitutionFinanciere.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    type_inst = request.GET.get('type', '')
    
    if q:
        institutions = institutions.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_inst:
        institutions = institutions.filter(type_institution=type_inst)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Institutions Financieres"
    
    # En-têtes
    headers = [
        "ID", "Nom institution", "Sigle", "Type institution", "Année création",
        "N° Agrément", "IFU", "Description services", "Services disponibles",
        "Taux crédit", "Taux épargne", "Conditions éligibilité", "Public cible",
        "Nom responsable", "Fonction responsable", "Téléphone 1", "Téléphone 2",
        "WhatsApp", "Email", "Site web", "Facebook", "Quartier", "Canton",
        "Adresse complète", "Situation", "Latitude", "Longitude", "Nombre agences",
        "Horaires", "Certifie informations", "Accepte publication", "Accepte contact",
        "Engagement", "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for inst in institutions:
        services_text = ", ".join(part.strip().title() for part in inst.services.split(",") if part.strip()) if inst.services else ""
        row = [
            inst.pk,
            inst.nom_institution,
            inst.sigle or "",
            inst.get_type_institution_display(),
            inst.annee_creation if inst.annee_creation else "",
            inst.numero_agrement or "",
            inst.ifu or "",
            inst.description_services,
            services_text,
            inst.taux_credit or "",
            inst.taux_epargne or "",
            inst.conditions_eligibilite or "",
            inst.public_cible or "",
            inst.nom_responsable,
            inst.fonction_responsable,
            inst.telephone1,
            inst.telephone2 or "",
            inst.whatsapp or "",
            inst.email,
            inst.site_web or "",
            inst.facebook or "",
            inst.quartier,
            inst.canton or "",
            inst.adresse_complete,
            inst.get_situation_display(),
            inst.latitude if inst.latitude else "",
            inst.longitude if inst.longitude else "",
            inst.nombre_agences if inst.nombre_agences else "",
            inst.horaires,
            "Oui" if inst.certifie_info else "Non",
            "Oui" if inst.accepte_public else "Non",
            "Oui" if inst.accepte_contact else "Non",
            "Oui" if inst.engagement else "Non",
            "Oui" if inst.est_valide_par_mairie else "Non",
            _format_excel_value(inst.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="institutions_financieres.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_jeunes(request):
    """Exporte tous les jeunes demandeurs d'emploi en Excel avec tous les champs."""
    jeunes = ProfilEmploi.objects.filter(type_profil='jeune').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        jeunes = jeunes.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        jeunes = jeunes.filter(niveau_etude=niveau)
    if dispo:
        jeunes = jeunes.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jeunes Demandeurs Emploi"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Service citoyen obligatoire", "Accepte RGPD", "Accepte contact",
        "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for jeune in jeunes:
        row = [
            jeune.pk,
            jeune.nom,
            jeune.prenoms,
            jeune.get_sexe_display(),
            _format_excel_value(jeune.date_naissance),
            jeune.nationalite or "",
            jeune.telephone1,
            jeune.telephone2 or "",
            jeune.email,
            jeune.quartier,
            jeune.canton or "",
            jeune.adresse_complete,
            "Oui" if jeune.est_resident_kloto else "Non",
            jeune.get_niveau_etude_display() if jeune.niveau_etude else "",
            jeune.diplome_principal or "",
            jeune.domaine_competence,
            jeune.experiences or "",
            jeune.get_situation_actuelle_display(),
            jeune.employeur_actuel or "",
            jeune.get_disponibilite_display(),
            jeune.get_type_contrat_souhaite_display() if jeune.type_contrat_souhaite else "",
            jeune.salaire_souhaite or "",
            "Oui" if jeune.service_citoyen_obligatoire else "Non",
            "Oui" if jeune.accepte_rgpd else "Non",
            "Oui" if jeune.accepte_contact else "Non",
            "Oui" if jeune.est_valide_par_mairie else "Non",
            _format_excel_value(jeune.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="jeunes_demandeurs_emploi.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_retraites(request):
    """Exporte tous les retraités actifs en Excel avec tous les champs."""
    retraites = ProfilEmploi.objects.filter(type_profil='retraite').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        retraites = retraites.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        retraites = retraites.filter(niveau_etude=niveau)
    if dispo:
        retraites = retraites.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Retraites Actifs"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Caisse retraite", "Dernier poste", "Années expérience",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for retraite in retraites:
        row = [
            retraite.pk,
            retraite.nom,
            retraite.prenoms,
            retraite.get_sexe_display(),
            _format_excel_value(retraite.date_naissance),
            retraite.nationalite or "",
            retraite.telephone1,
            retraite.telephone2 or "",
            retraite.email,
            retraite.quartier,
            retraite.canton or "",
            retraite.adresse_complete,
            "Oui" if retraite.est_resident_kloto else "Non",
            retraite.get_niveau_etude_display() if retraite.niveau_etude else "",
            retraite.diplome_principal or "",
            retraite.domaine_competence,
            retraite.experiences or "",
            retraite.get_situation_actuelle_display(),
            retraite.employeur_actuel or "",
            retraite.get_disponibilite_display(),
            retraite.get_type_contrat_souhaite_display() if retraite.type_contrat_souhaite else "",
            retraite.salaire_souhaite or "",
            retraite.caisse_retraite or "",
            retraite.dernier_poste or "",
            retraite.annees_experience if retraite.annees_experience else "",
            "Oui" if retraite.accepte_rgpd else "Non",
            "Oui" if retraite.accepte_contact else "Non",
            "Oui" if retraite.est_valide_par_mairie else "Non",
            _format_excel_value(retraite.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="retraites_actifs.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_diaspora(request):
    """Exporte tous les membres de la diaspora en Excel avec tous les champs."""
    membres = MembreDiaspora.objects.all().order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    pays = request.GET.get('pays', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        membres = membres.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_whatsapp__icontains=q) |
            Q(profession_actuelle__icontains=q) |
            Q(domaine_formation__icontains=q)
        )
    if pays:
        membres = membres.filter(pays_residence_actuelle__icontains=pays)
    if secteur:
        membres = membres.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres Diaspora"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité(s)",
        "N° Pièce identité", "Pays résidence", "Ville résidence", "Adresse étranger",
        "Commune origine", "Quartier/Village origine", "Nom parent/tuteur", "Année départ",
        "Fréquence retour", "Téléphone WhatsApp", "Email", "Réseaux sociaux",
        "Contact pays - Nom", "Contact pays - Téléphone", "Niveau études",
        "Domaine formation", "Profession actuelle", "Secteur activité",
        "Secteur activité (autre)", "Années expérience", "Statut professionnel",
        "Type titre séjour", "Appui investissement projets", "Appui financement infrastructures",
        "Appui parrainage communautaire", "Appui jeunes/femmes entrepreneurs",
        "Transfert compétences", "Formation jeunes", "Appui digitalisation",
        "Conseils techniques", "Encadrement mentorat", "Création entreprise locale",
        "Appui PME locales", "Recrutement jeunes commune", "Mise relation ONG",
        "Coopération décentralisée", "Recherche financements internationaux",
        "Promotion commune international", "Participation activités communales",
        "Participation réunions diaspora", "Appui actions sociales/culturelles",
        "Comment contribuer", "Disposition participation", "Domaine intervention prioritaire",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription", "Date modification"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for membre in membres:
        row = [
            membre.pk,
            membre.nom,
            membre.prenoms,
            membre.get_sexe_display(),
            _format_excel_value(membre.date_naissance),
            membre.nationalites,
            membre.numero_piece_identite,
            membre.pays_residence_actuelle,
            membre.ville_residence_actuelle,
            membre.adresse_complete_etranger,
            membre.commune_origine,
            membre.quartier_village_origine,
            membre.nom_parent_tuteur_originaire,
            membre.annee_depart_pays,
            membre.get_frequence_retour_pays_display(),
            membre.telephone_whatsapp,
            membre.email,
            membre.reseaux_sociaux or "",
            membre.contact_au_pays_nom,
            membre.contact_au_pays_telephone,
            membre.get_niveau_etudes_display(),
            membre.domaine_formation,
            membre.profession_actuelle,
            membre.get_secteur_activite_display(),
            membre.secteur_activite_autre or "",
            membre.annees_experience,
            membre.get_statut_professionnel_display(),
            membre.type_titre_sejour or "",
            "Oui" if membre.appui_investissement_projets else "Non",
            "Oui" if membre.appui_financement_infrastructures else "Non",
            "Oui" if membre.appui_parrainage_communautaire else "Non",
            "Oui" if membre.appui_jeunes_femmes_entrepreneurs else "Non",
            "Oui" if membre.transfert_competences else "Non",
            "Oui" if membre.formation_jeunes else "Non",
            "Oui" if membre.appui_digitalisation else "Non",
            "Oui" if membre.conseils_techniques else "Non",
            "Oui" if membre.encadrement_mentorat else "Non",
            "Oui" if membre.creation_entreprise_locale else "Non",
            "Oui" if membre.appui_pme_locales else "Non",
            "Oui" if membre.recrutement_jeunes_commune else "Non",
            "Oui" if membre.mise_relation_ong else "Non",
            "Oui" if membre.cooperation_decentralisee else "Non",
            "Oui" if membre.recherche_financements_internationaux else "Non",
            "Oui" if membre.promotion_commune_international else "Non",
            "Oui" if membre.participation_activites_communales else "Non",
            "Oui" if membre.participation_reunions_diaspora else "Non",
            "Oui" if membre.appui_actions_sociales_culturelles else "Non",
            membre.comment_contribuer,
            membre.get_disposition_participation_display(),
            membre.domaine_intervention_prioritaire,
            "Oui" if membre.accepte_rgpd else "Non",
            "Oui" if membre.accepte_contact else "Non",
            "Oui" if membre.est_valide_par_mairie else "Non",
            _format_excel_value(membre.date_inscription),
            _format_excel_value(membre.date_modification),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="membres_diaspora.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_candidatures(request):
    """Exporte toutes les candidatures en Excel avec tous les champs."""
    candidatures = Candidature.objects.all().select_related('appel_offre', 'candidat').order_by('-date_soumission')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    if q:
        candidatures = candidatures.filter(
            Q(appel_offre__titre__icontains=q) |
            Q(appel_offre__reference__icontains=q) |
            Q(candidat__first_name__icontains=q) |
            Q(candidat__last_name__icontains=q) |
            Q(candidat__email__icontains=q)
        )
    if statut:
        candidatures = candidatures.filter(statut=statut)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    
    # En-têtes
    headers = [
        "ID", "Appel d'offres - Titre", "Appel d'offres - Référence", "Appel d'offres - Description",
        "Appel d'offres - Public cible", "Appel d'offres - Date début", "Appel d'offres - Date fin",
        "Appel d'offres - Budget estimé", "Candidat - Username", "Candidat - Nom", "Candidat - Prénom",
        "Candidat - Email", "Statut candidature", "Message accompagnement", "Date soumission"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for candidature in candidatures:
        row = [
            candidature.pk,
            candidature.appel_offre.titre,
            candidature.appel_offre.reference or "",
            candidature.appel_offre.description,
            candidature.appel_offre.get_public_cible_display(),
            _format_excel_value(candidature.appel_offre.date_debut),
            _format_excel_value(candidature.appel_offre.date_fin),
            candidature.appel_offre.budget_estime if candidature.appel_offre.budget_estime else "",
            candidature.candidat.username,
            candidature.candidat.last_name or "",
            candidature.candidat.first_name or "",
            candidature.candidat.email,
            candidature.get_statut_display(),
            candidature.message_accompagnement or "",
            _format_excel_value(candidature.date_soumission),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidatures.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_jeunes(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    niveau = request.GET.get("niveau") or ""

    # Uniquement les profils validés par la mairie
    qs = ProfilEmploi.objects.filter(type_profil="jeune", est_valide_par_mairie=True)
    if niveau:
        qs = qs.filter(niveau_etude=niveau)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="jeunes_demandeurs_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Jeunes Demandeurs d'Emploi", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Diplôme", "Téléphone", "Quartier", "Compétences"]]
    for p in qs.order_by("-date_inscription")[:1000]:
        data.append([
            p.nom,
            p.prenoms,
            (p.diplome_principal or "")[:30],
            p.telephone1,
            p.quartier,
            (p.domaine_competence or "")[:60],
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3*cm, 8*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        _draw_pdf_header(c, d, conf)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_diaspora(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    pays = request.GET.get("pays") or ""

    # Uniquement les membres validés par la mairie
    qs = MembreDiaspora.objects.filter(est_valide_par_mairie=True)
    if pays:
        qs = qs.filter(pays_residence_actuelle__icontains=pays)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="diaspora_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Membres de la Diaspora", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Pays de résidence", "Ville", "Téléphone", "Email", "Profession"]]
    for m in qs.order_by("-date_inscription")[:1000]:
        data.append([
            m.nom,
            m.prenoms,
            m.pays_residence_actuelle[:25] if m.pays_residence_actuelle else "",
            m.ville_residence_actuelle[:20] if m.ville_residence_actuelle else "",
            m.telephone_whatsapp[:15] if m.telephone_whatsapp else "",
            m.email[:30] if m.email else "",
            m.profession_actuelle[:30] if m.profession_actuelle else "",
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3.5*cm, 5*cm, 4.5*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        width, height = d.pagesize
        y = height - 40
        c.saveState()
        c.translate(width/2, height/2)
        c.rotate(45)
        c.setFont("Helvetica-Bold", 36)
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.drawCentredString(0, 0, (conf.nom_commune if conf else "Mairie de Kloto 1").upper())
        c.restoreState()
        if conf and getattr(conf, "logo", None) and getattr(conf.logo, "path", None):
            try:
                c.drawImage(conf.logo.path, 40, y-30, width=40, height=40, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        c.setFont("Helvetica-Bold", 12)
        c.drawString(100, y, f"République Togolaise – {conf.nom_commune if conf else 'Mairie de Kloto 1'}")
        
        c.setFont("Helvetica", 9)
        c.drawString(40, 30, timezone.now().strftime("Édité le %d/%m/%Y %H:%M"))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def notifications_candidats(request):
    """Liste des appels d'offres avec candidats acceptés pour envoyer des notifications."""
    
    # Récupérer tous les appels d'offres qui ont au moins une candidature acceptée
    appels_offres = AppelOffre.objects.filter(
        candidatures__statut='acceptee'
    ).distinct().annotate(
        nb_candidats_acceptes=Count('candidatures', filter=Q(candidatures__statut='acceptee'))
    ).order_by('-date_debut')
    
    context = {
        'appels_offres': appels_offres,
    }
    
    return render(request, "admin/notifications_candidats.html", context)


@login_required
@user_passes_test(is_staff_user)
def envoyer_notifications_candidats(request, appel_offre_id):
    """Affiche le formulaire et traite l'envoi de notifications aux candidats acceptés."""
    
    appel_offre = get_object_or_404(AppelOffre, pk=appel_offre_id)
    
    # Récupérer uniquement les candidats acceptés pour cet appel d'offres
    candidats_acceptes = Candidature.objects.filter(
        appel_offre=appel_offre,
        statut='acceptee'
    ).select_related('candidat')
    
    if not candidats_acceptes.exists():
        messages.warning(request, "Aucun candidat accepté trouvé pour cet appel d'offres.")
        return redirect('notifications_candidats')
    
    if request.method == 'POST':
        titre = request.POST.get('titre', '')
        message = request.POST.get('message', '')
        type_notification = request.POST.get('type', Notification.TYPE_INFO)
        
        if not titre or not message:
            messages.error(request, "Le titre et le message sont obligatoires.")
        else:
            # Créer une notification pour chaque candidat accepté
            notifications_creees = 0
            for candidature in candidats_acceptes:
                Notification.objects.create(
                    recipient=candidature.candidat,
                    title=titre,
                    message=message,
                    type=type_notification,
                    created_by=request.user,
                )
                notifications_creees += 1
            
            messages.success(
                request, 
                f"Notifications envoyées avec succès à {notifications_creees} candidat(s) accepté(s)."
            )
            return redirect('notifications_candidats')
    
    context = {
        'appel_offre': appel_offre,
        'candidats_acceptes': candidats_acceptes,
        'nb_candidats': candidats_acceptes.count(),
        'type_choices': Notification.TYPE_CHOICES,
    }
    
    return render(request, "admin/envoyer_notifications_candidats.html", context)


# ========== FONCTIONS D'EXPORT EXCEL ==========

def _format_excel_value(value):
    """Formate une valeur pour l'export Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else value.strftime("%d/%m/%Y")
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value if item)
    return str(value)


def _style_excel_header(ws, row_num):
    """Applique un style au header Excel."""
    header_fill = PatternFill(start_color="006233", end_color="006233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


@login_required
@user_passes_test(is_staff_user)
def export_excel_acteurs(request):
    """Exporte tous les acteurs économiques en Excel avec tous les champs."""
    acteurs = ActeurEconomique.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres comme dans la vue liste
    q = request.GET.get('q', '')
    type_acteur = request.GET.get('type', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        acteurs = acteurs.filter(
            Q(raison_sociale__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_acteur:
        acteurs = acteurs.filter(type_acteur=type_acteur)
    if secteur:
        acteurs = acteurs.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Acteurs Economiques"
    
    # En-têtes
    headers = [
        "ID", "Raison sociale", "Sigle", "Type d'acteur", "Secteur d'activité", "Statut juridique",
        "Description", "RCCM", "CFE", "N° Carte opérateur", "NIF", "Date de création",
        "Capital social (FCFA)", "Nom responsable", "Fonction responsable", "Téléphone 1",
        "Téléphone 2", "Email", "Site web", "Quartier", "Canton", "Adresse complète",
        "Situation", "Latitude", "Longitude", "Nombre d'employés", "Chiffre d'affaires",
        "Accepte publication", "Certifie informations", "Accepte conditions",
        "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for acteur in acteurs:
        row = [
            acteur.pk,
            acteur.raison_sociale,
            acteur.sigle or "",
            acteur.get_type_acteur_display(),
            acteur.get_secteur_activite_display(),
            acteur.get_statut_juridique_display(),
            acteur.description,
            acteur.rccm or "",
            acteur.cfe or "",
            acteur.numero_carte_operateur or "",
            acteur.nif or "",
            _format_excel_value(acteur.date_creation),
            acteur.capital_social if acteur.capital_social else "",
            acteur.nom_responsable,
            acteur.fonction_responsable,
            acteur.telephone1,
            acteur.telephone2 or "",
            acteur.email,
            acteur.site_web or "",
            acteur.quartier,
            acteur.canton or "",
            acteur.adresse_complete,
            acteur.get_situation_display(),
            acteur.latitude if acteur.latitude else "",
            acteur.longitude if acteur.longitude else "",
            acteur.get_nombre_employes_display() if acteur.nombre_employes else "",
            acteur.get_chiffre_affaires_display() if acteur.chiffre_affaires else "",
            "Oui" if acteur.accepte_public else "Non",
            "Oui" if acteur.certifie_information else "Non",
            "Oui" if acteur.accepte_conditions else "Non",
            "Oui" if acteur.est_valide_par_mairie else "Non",
            _format_excel_value(acteur.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="acteurs_economiques.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_institutions(request):
    """Exporte toutes les institutions financières en Excel avec tous les champs."""
    institutions = InstitutionFinanciere.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    type_inst = request.GET.get('type', '')
    
    if q:
        institutions = institutions.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_inst:
        institutions = institutions.filter(type_institution=type_inst)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Institutions Financieres"
    
    # En-têtes
    headers = [
        "ID", "Nom institution", "Sigle", "Type institution", "Année création",
        "N° Agrément", "IFU", "Description services", "Services disponibles",
        "Taux crédit", "Taux épargne", "Conditions éligibilité", "Public cible",
        "Nom responsable", "Fonction responsable", "Téléphone 1", "Téléphone 2",
        "WhatsApp", "Email", "Site web", "Facebook", "Quartier", "Canton",
        "Adresse complète", "Situation", "Latitude", "Longitude", "Nombre agences",
        "Horaires", "Certifie informations", "Accepte publication", "Accepte contact",
        "Engagement", "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for inst in institutions:
        services_text = ", ".join(part.strip().title() for part in inst.services.split(",") if part.strip()) if inst.services else ""
        row = [
            inst.pk,
            inst.nom_institution,
            inst.sigle or "",
            inst.get_type_institution_display(),
            inst.annee_creation if inst.annee_creation else "",
            inst.numero_agrement or "",
            inst.ifu or "",
            inst.description_services,
            services_text,
            inst.taux_credit or "",
            inst.taux_epargne or "",
            inst.conditions_eligibilite or "",
            inst.public_cible or "",
            inst.nom_responsable,
            inst.fonction_responsable,
            inst.telephone1,
            inst.telephone2 or "",
            inst.whatsapp or "",
            inst.email,
            inst.site_web or "",
            inst.facebook or "",
            inst.quartier,
            inst.canton or "",
            inst.adresse_complete,
            inst.get_situation_display(),
            inst.latitude if inst.latitude else "",
            inst.longitude if inst.longitude else "",
            inst.nombre_agences if inst.nombre_agences else "",
            inst.horaires,
            "Oui" if inst.certifie_info else "Non",
            "Oui" if inst.accepte_public else "Non",
            "Oui" if inst.accepte_contact else "Non",
            "Oui" if inst.engagement else "Non",
            "Oui" if inst.est_valide_par_mairie else "Non",
            _format_excel_value(inst.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="institutions_financieres.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_jeunes(request):
    """Exporte tous les jeunes demandeurs d'emploi en Excel avec tous les champs."""
    jeunes = ProfilEmploi.objects.filter(type_profil='jeune').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        jeunes = jeunes.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        jeunes = jeunes.filter(niveau_etude=niveau)
    if dispo:
        jeunes = jeunes.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jeunes Demandeurs Emploi"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Service citoyen obligatoire", "Accepte RGPD", "Accepte contact",
        "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for jeune in jeunes:
        row = [
            jeune.pk,
            jeune.nom,
            jeune.prenoms,
            jeune.get_sexe_display(),
            _format_excel_value(jeune.date_naissance),
            jeune.nationalite or "",
            jeune.telephone1,
            jeune.telephone2 or "",
            jeune.email,
            jeune.quartier,
            jeune.canton or "",
            jeune.adresse_complete,
            "Oui" if jeune.est_resident_kloto else "Non",
            jeune.get_niveau_etude_display() if jeune.niveau_etude else "",
            jeune.diplome_principal or "",
            jeune.domaine_competence,
            jeune.experiences or "",
            jeune.get_situation_actuelle_display(),
            jeune.employeur_actuel or "",
            jeune.get_disponibilite_display(),
            jeune.get_type_contrat_souhaite_display() if jeune.type_contrat_souhaite else "",
            jeune.salaire_souhaite or "",
            "Oui" if jeune.service_citoyen_obligatoire else "Non",
            "Oui" if jeune.accepte_rgpd else "Non",
            "Oui" if jeune.accepte_contact else "Non",
            "Oui" if jeune.est_valide_par_mairie else "Non",
            _format_excel_value(jeune.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="jeunes_demandeurs_emploi.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_retraites(request):
    """Exporte tous les retraités actifs en Excel avec tous les champs."""
    retraites = ProfilEmploi.objects.filter(type_profil='retraite').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        retraites = retraites.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        retraites = retraites.filter(niveau_etude=niveau)
    if dispo:
        retraites = retraites.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Retraites Actifs"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Caisse retraite", "Dernier poste", "Années expérience",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for retraite in retraites:
        row = [
            retraite.pk,
            retraite.nom,
            retraite.prenoms,
            retraite.get_sexe_display(),
            _format_excel_value(retraite.date_naissance),
            retraite.nationalite or "",
            retraite.telephone1,
            retraite.telephone2 or "",
            retraite.email,
            retraite.quartier,
            retraite.canton or "",
            retraite.adresse_complete,
            "Oui" if retraite.est_resident_kloto else "Non",
            retraite.get_niveau_etude_display() if retraite.niveau_etude else "",
            retraite.diplome_principal or "",
            retraite.domaine_competence,
            retraite.experiences or "",
            retraite.get_situation_actuelle_display(),
            retraite.employeur_actuel or "",
            retraite.get_disponibilite_display(),
            retraite.get_type_contrat_souhaite_display() if retraite.type_contrat_souhaite else "",
            retraite.salaire_souhaite or "",
            retraite.caisse_retraite or "",
            retraite.dernier_poste or "",
            retraite.annees_experience if retraite.annees_experience else "",
            "Oui" if retraite.accepte_rgpd else "Non",
            "Oui" if retraite.accepte_contact else "Non",
            "Oui" if retraite.est_valide_par_mairie else "Non",
            _format_excel_value(retraite.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="retraites_actifs.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_diaspora(request):
    """Exporte tous les membres de la diaspora en Excel avec tous les champs."""
    membres = MembreDiaspora.objects.all().order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    pays = request.GET.get('pays', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        membres = membres.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_whatsapp__icontains=q) |
            Q(profession_actuelle__icontains=q) |
            Q(domaine_formation__icontains=q)
        )
    if pays:
        membres = membres.filter(pays_residence_actuelle__icontains=pays)
    if secteur:
        membres = membres.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres Diaspora"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité(s)",
        "N° Pièce identité", "Pays résidence", "Ville résidence", "Adresse étranger",
        "Commune origine", "Quartier/Village origine", "Nom parent/tuteur", "Année départ",
        "Fréquence retour", "Téléphone WhatsApp", "Email", "Réseaux sociaux",
        "Contact pays - Nom", "Contact pays - Téléphone", "Niveau études",
        "Domaine formation", "Profession actuelle", "Secteur activité",
        "Secteur activité (autre)", "Années expérience", "Statut professionnel",
        "Type titre séjour", "Appui investissement projets", "Appui financement infrastructures",
        "Appui parrainage communautaire", "Appui jeunes/femmes entrepreneurs",
        "Transfert compétences", "Formation jeunes", "Appui digitalisation",
        "Conseils techniques", "Encadrement mentorat", "Création entreprise locale",
        "Appui PME locales", "Recrutement jeunes commune", "Mise relation ONG",
        "Coopération décentralisée", "Recherche financements internationaux",
        "Promotion commune international", "Participation activités communales",
        "Participation réunions diaspora", "Appui actions sociales/culturelles",
        "Comment contribuer", "Disposition participation", "Domaine intervention prioritaire",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription", "Date modification"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for membre in membres:
        row = [
            membre.pk,
            membre.nom,
            membre.prenoms,
            membre.get_sexe_display(),
            _format_excel_value(membre.date_naissance),
            membre.nationalites,
            membre.numero_piece_identite,
            membre.pays_residence_actuelle,
            membre.ville_residence_actuelle,
            membre.adresse_complete_etranger,
            membre.commune_origine,
            membre.quartier_village_origine,
            membre.nom_parent_tuteur_originaire,
            membre.annee_depart_pays,
            membre.get_frequence_retour_pays_display(),
            membre.telephone_whatsapp,
            membre.email,
            membre.reseaux_sociaux or "",
            membre.contact_au_pays_nom,
            membre.contact_au_pays_telephone,
            membre.get_niveau_etudes_display(),
            membre.domaine_formation,
            membre.profession_actuelle,
            membre.get_secteur_activite_display(),
            membre.secteur_activite_autre or "",
            membre.annees_experience,
            membre.get_statut_professionnel_display(),
            membre.type_titre_sejour or "",
            "Oui" if membre.appui_investissement_projets else "Non",
            "Oui" if membre.appui_financement_infrastructures else "Non",
            "Oui" if membre.appui_parrainage_communautaire else "Non",
            "Oui" if membre.appui_jeunes_femmes_entrepreneurs else "Non",
            "Oui" if membre.transfert_competences else "Non",
            "Oui" if membre.formation_jeunes else "Non",
            "Oui" if membre.appui_digitalisation else "Non",
            "Oui" if membre.conseils_techniques else "Non",
            "Oui" if membre.encadrement_mentorat else "Non",
            "Oui" if membre.creation_entreprise_locale else "Non",
            "Oui" if membre.appui_pme_locales else "Non",
            "Oui" if membre.recrutement_jeunes_commune else "Non",
            "Oui" if membre.mise_relation_ong else "Non",
            "Oui" if membre.cooperation_decentralisee else "Non",
            "Oui" if membre.recherche_financements_internationaux else "Non",
            "Oui" if membre.promotion_commune_international else "Non",
            "Oui" if membre.participation_activites_communales else "Non",
            "Oui" if membre.participation_reunions_diaspora else "Non",
            "Oui" if membre.appui_actions_sociales_culturelles else "Non",
            membre.comment_contribuer,
            membre.get_disposition_participation_display(),
            membre.domaine_intervention_prioritaire,
            "Oui" if membre.accepte_rgpd else "Non",
            "Oui" if membre.accepte_contact else "Non",
            "Oui" if membre.est_valide_par_mairie else "Non",
            _format_excel_value(membre.date_inscription),
            _format_excel_value(membre.date_modification),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="membres_diaspora.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_candidatures(request):
    """Exporte toutes les candidatures en Excel avec tous les champs."""
    candidatures = Candidature.objects.all().select_related('appel_offre', 'candidat').order_by('-date_soumission')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    if q:
        candidatures = candidatures.filter(
            Q(appel_offre__titre__icontains=q) |
            Q(appel_offre__reference__icontains=q) |
            Q(candidat__first_name__icontains=q) |
            Q(candidat__last_name__icontains=q) |
            Q(candidat__email__icontains=q)
        )
    if statut:
        candidatures = candidatures.filter(statut=statut)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    
    # En-têtes
    headers = [
        "ID", "Appel d'offres - Titre", "Appel d'offres - Référence", "Appel d'offres - Description",
        "Appel d'offres - Public cible", "Appel d'offres - Date début", "Appel d'offres - Date fin",
        "Appel d'offres - Budget estimé", "Candidat - Username", "Candidat - Nom", "Candidat - Prénom",
        "Candidat - Email", "Statut candidature", "Message accompagnement", "Date soumission"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for candidature in candidatures:
        row = [
            candidature.pk,
            candidature.appel_offre.titre,
            candidature.appel_offre.reference or "",
            candidature.appel_offre.description,
            candidature.appel_offre.get_public_cible_display(),
            _format_excel_value(candidature.appel_offre.date_debut),
            _format_excel_value(candidature.appel_offre.date_fin),
            candidature.appel_offre.budget_estime if candidature.appel_offre.budget_estime else "",
            candidature.candidat.username,
            candidature.candidat.last_name or "",
            candidature.candidat.first_name or "",
            candidature.candidat.email,
            candidature.get_statut_display(),
            candidature.message_accompagnement or "",
            _format_excel_value(candidature.date_soumission),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidatures.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_retraites(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    niveau = request.GET.get("niveau") or ""

    # Uniquement les profils validés par la mairie
    qs = ProfilEmploi.objects.filter(type_profil="retraite", est_valide_par_mairie=True)
    if niveau:
        qs = qs.filter(niveau_etude=niveau)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="retraites_actifs_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=PDF_HEADER_HEIGHT_CM * cm, bottomMargin=1.5 * cm)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Retraités Actifs", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Diplôme", "Téléphone", "Quartier", "Dernier poste"]]
    for p in qs.order_by("-date_inscription")[:1000]:
        data.append([
            p.nom,
            p.prenoms,
            (p.diplome_principal or "")[:30],
            p.telephone1,
            p.quartier,
            (p.dernier_poste or "")[:60],
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3*cm, 8*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        width, height = d.pagesize
        y = height - 40
        c.saveState()
        c.translate(width/2, height/2)
        c.rotate(45)
        c.setFont("Helvetica-Bold", 36)
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.drawCentredString(0, 0, (conf.nom_commune if conf else "Mairie de Kloto 1").upper())
        c.restoreState()
        if conf and getattr(conf, "logo", None) and getattr(conf.logo, "path", None):
            try:
                c.drawImage(conf.logo.path, 40, y-30, width=40, height=40, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        c.setFont("Helvetica-Bold", 12)
        c.drawString(100, y, f"République Togolaise – {conf.nom_commune if conf else 'Mairie de Kloto 1'}")
        
        c.setFont("Helvetica", 9)
        c.drawString(40, 30, timezone.now().strftime("Édité le %d/%m/%Y %H:%M"))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_pdf_diaspora(request):
    start = request.GET.get('start')
    end = request.GET.get('end')
    pays = request.GET.get("pays") or ""

    # Uniquement les membres validés par la mairie
    qs = MembreDiaspora.objects.filter(est_valide_par_mairie=True)
    if pays:
        qs = qs.filter(pays_residence_actuelle__icontains=pays)
    conf = ConfigurationMairie.objects.filter(est_active=True).first()
    if start:
        try:
            sd = datetime.strptime(start, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__gte=sd)
        except ValueError:
            pass
    if end:
        try:
            ed = datetime.strptime(end, "%Y-%m-%d").date()
            qs = qs.filter(date_inscription__date__lte=ed)
        except ValueError:
            pass
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="diaspora_valides.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#006233"), alignment=1, spaceAfter=12)
    story.append(Paragraph("Membres de la Diaspora", title_style))
    if start or end:
        story.append(Paragraph(f"Période: {start or '...'} au {end or '...'}", styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))
    data = [["Nom", "Prénoms", "Pays de résidence", "Ville", "Téléphone", "Email", "Profession"]]
    for m in qs.order_by("-date_inscription")[:1000]:
        data.append([
            m.nom,
            m.prenoms,
            m.pays_residence_actuelle[:25] if m.pays_residence_actuelle else "",
            m.ville_residence_actuelle[:20] if m.ville_residence_actuelle else "",
            m.telephone_whatsapp[:15] if m.telephone_whatsapp else "",
            m.email[:30] if m.email else "",
            m.profession_actuelle[:30] if m.profession_actuelle else "",
        ])
    table = Table(data, colWidths=[3.5*cm, 4*cm, 4*cm, 3.5*cm, 3.5*cm, 5*cm, 4.5*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E8F5E9")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Date et Signature : ________________________________", styles["Normal"]))
    
    def on_page(c, d):
        width, height = d.pagesize
        y = height - 40
        c.saveState()
        c.translate(width/2, height/2)
        c.rotate(45)
        c.setFont("Helvetica-Bold", 36)
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.drawCentredString(0, 0, (conf.nom_commune if conf else "Mairie de Kloto 1").upper())
        c.restoreState()
        if conf and getattr(conf, "logo", None) and getattr(conf.logo, "path", None):
            try:
                c.drawImage(conf.logo.path, 40, y-30, width=40, height=40, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        c.setFont("Helvetica-Bold", 12)
        c.drawString(100, y, f"République Togolaise – {conf.nom_commune if conf else 'Mairie de Kloto 1'}")
        
        c.setFont("Helvetica", 9)
        c.drawString(40, 30, timezone.now().strftime("Édité le %d/%m/%Y %H:%M"))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page, canvasmaker=NumberedCanvas)
    return response


@login_required
@user_passes_test(is_staff_user)
def notifications_candidats(request):
    """Liste des appels d'offres avec candidats acceptés pour envoyer des notifications."""
    
    # Récupérer tous les appels d'offres qui ont au moins une candidature acceptée
    appels_offres = AppelOffre.objects.filter(
        candidatures__statut='acceptee'
    ).distinct().annotate(
        nb_candidats_acceptes=Count('candidatures', filter=Q(candidatures__statut='acceptee'))
    ).order_by('-date_debut')
    
    context = {
        'appels_offres': appels_offres,
    }
    
    return render(request, "admin/notifications_candidats.html", context)


@login_required
@user_passes_test(is_staff_user)
def envoyer_notifications_candidats(request, appel_offre_id):
    """Affiche le formulaire et traite l'envoi de notifications aux candidats acceptés."""
    
    appel_offre = get_object_or_404(AppelOffre, pk=appel_offre_id)
    
    # Récupérer uniquement les candidats acceptés pour cet appel d'offres
    candidats_acceptes = Candidature.objects.filter(
        appel_offre=appel_offre,
        statut='acceptee'
    ).select_related('candidat')
    
    if not candidats_acceptes.exists():
        messages.warning(request, "Aucun candidat accepté trouvé pour cet appel d'offres.")
        return redirect('notifications_candidats')
    
    if request.method == 'POST':
        titre = request.POST.get('titre', '')
        message = request.POST.get('message', '')
        type_notification = request.POST.get('type', Notification.TYPE_INFO)
        
        if not titre or not message:
            messages.error(request, "Le titre et le message sont obligatoires.")
        else:
            # Créer une notification pour chaque candidat accepté
            notifications_creees = 0
            for candidature in candidats_acceptes:
                Notification.objects.create(
                    recipient=candidature.candidat,
                    title=titre,
                    message=message,
                    type=type_notification,
                    created_by=request.user,
                )
                notifications_creees += 1
            
            messages.success(
                request, 
                f"Notifications envoyées avec succès à {notifications_creees} candidat(s) accepté(s)."
            )
            return redirect('notifications_candidats')
    
    context = {
        'appel_offre': appel_offre,
        'candidats_acceptes': candidats_acceptes,
        'nb_candidats': candidats_acceptes.count(),
        'type_choices': Notification.TYPE_CHOICES,
    }
    
    return render(request, "admin/envoyer_notifications_candidats.html", context)


# ========== FONCTIONS D'EXPORT EXCEL ==========

def _format_excel_value(value):
    """Formate une valeur pour l'export Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else value.strftime("%d/%m/%Y")
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value if item)
    return str(value)


def _style_excel_header(ws, row_num):
    """Applique un style au header Excel."""
    header_fill = PatternFill(start_color="006233", end_color="006233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


@login_required
@user_passes_test(is_staff_user)
def export_excel_acteurs(request):
    """Exporte tous les acteurs économiques en Excel avec tous les champs."""
    acteurs = ActeurEconomique.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres comme dans la vue liste
    q = request.GET.get('q', '')
    type_acteur = request.GET.get('type', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        acteurs = acteurs.filter(
            Q(raison_sociale__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_acteur:
        acteurs = acteurs.filter(type_acteur=type_acteur)
    if secteur:
        acteurs = acteurs.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Acteurs Economiques"
    
    # En-têtes
    headers = [
        "ID", "Raison sociale", "Sigle", "Type d'acteur", "Secteur d'activité", "Statut juridique",
        "Description", "RCCM", "CFE", "N° Carte opérateur", "NIF", "Date de création",
        "Capital social (FCFA)", "Nom responsable", "Fonction responsable", "Téléphone 1",
        "Téléphone 2", "Email", "Site web", "Quartier", "Canton", "Adresse complète",
        "Situation", "Latitude", "Longitude", "Nombre d'employés", "Chiffre d'affaires",
        "Accepte publication", "Certifie informations", "Accepte conditions",
        "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for acteur in acteurs:
        row = [
            acteur.pk,
            acteur.raison_sociale,
            acteur.sigle or "",
            acteur.get_type_acteur_display(),
            acteur.get_secteur_activite_display(),
            acteur.get_statut_juridique_display(),
            acteur.description,
            acteur.rccm or "",
            acteur.cfe or "",
            acteur.numero_carte_operateur or "",
            acteur.nif or "",
            _format_excel_value(acteur.date_creation),
            acteur.capital_social if acteur.capital_social else "",
            acteur.nom_responsable,
            acteur.fonction_responsable,
            acteur.telephone1,
            acteur.telephone2 or "",
            acteur.email,
            acteur.site_web or "",
            acteur.quartier,
            acteur.canton or "",
            acteur.adresse_complete,
            acteur.get_situation_display(),
            acteur.latitude if acteur.latitude else "",
            acteur.longitude if acteur.longitude else "",
            acteur.get_nombre_employes_display() if acteur.nombre_employes else "",
            acteur.get_chiffre_affaires_display() if acteur.chiffre_affaires else "",
            "Oui" if acteur.accepte_public else "Non",
            "Oui" if acteur.certifie_information else "Non",
            "Oui" if acteur.accepte_conditions else "Non",
            "Oui" if acteur.est_valide_par_mairie else "Non",
            _format_excel_value(acteur.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="acteurs_economiques.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_institutions(request):
    """Exporte toutes les institutions financières en Excel avec tous les champs."""
    institutions = InstitutionFinanciere.objects.all().order_by('-date_enregistrement')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    type_inst = request.GET.get('type', '')
    
    if q:
        institutions = institutions.filter(
            Q(nom_institution__icontains=q) |
            Q(sigle__icontains=q) |
            Q(nom_responsable__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q)
        )
    if type_inst:
        institutions = institutions.filter(type_institution=type_inst)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Institutions Financieres"
    
    # En-têtes
    headers = [
        "ID", "Nom institution", "Sigle", "Type institution", "Année création",
        "N° Agrément", "IFU", "Description services", "Services disponibles",
        "Taux crédit", "Taux épargne", "Conditions éligibilité", "Public cible",
        "Nom responsable", "Fonction responsable", "Téléphone 1", "Téléphone 2",
        "WhatsApp", "Email", "Site web", "Facebook", "Quartier", "Canton",
        "Adresse complète", "Situation", "Latitude", "Longitude", "Nombre agences",
        "Horaires", "Certifie informations", "Accepte publication", "Accepte contact",
        "Engagement", "Validé par mairie", "Date d'enregistrement"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for inst in institutions:
        services_text = ", ".join(part.strip().title() for part in inst.services.split(",") if part.strip()) if inst.services else ""
        row = [
            inst.pk,
            inst.nom_institution,
            inst.sigle or "",
            inst.get_type_institution_display(),
            inst.annee_creation if inst.annee_creation else "",
            inst.numero_agrement or "",
            inst.ifu or "",
            inst.description_services,
            services_text,
            inst.taux_credit or "",
            inst.taux_epargne or "",
            inst.conditions_eligibilite or "",
            inst.public_cible or "",
            inst.nom_responsable,
            inst.fonction_responsable,
            inst.telephone1,
            inst.telephone2 or "",
            inst.whatsapp or "",
            inst.email,
            inst.site_web or "",
            inst.facebook or "",
            inst.quartier,
            inst.canton or "",
            inst.adresse_complete,
            inst.get_situation_display(),
            inst.latitude if inst.latitude else "",
            inst.longitude if inst.longitude else "",
            inst.nombre_agences if inst.nombre_agences else "",
            inst.horaires,
            "Oui" if inst.certifie_info else "Non",
            "Oui" if inst.accepte_public else "Non",
            "Oui" if inst.accepte_contact else "Non",
            "Oui" if inst.engagement else "Non",
            "Oui" if inst.est_valide_par_mairie else "Non",
            _format_excel_value(inst.date_enregistrement),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="institutions_financieres.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_jeunes(request):
    """Exporte tous les jeunes demandeurs d'emploi en Excel avec tous les champs."""
    jeunes = ProfilEmploi.objects.filter(type_profil='jeune').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        jeunes = jeunes.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        jeunes = jeunes.filter(niveau_etude=niveau)
    if dispo:
        jeunes = jeunes.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jeunes Demandeurs Emploi"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Service citoyen obligatoire", "Accepte RGPD", "Accepte contact",
        "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for jeune in jeunes:
        row = [
            jeune.pk,
            jeune.nom,
            jeune.prenoms,
            jeune.get_sexe_display(),
            _format_excel_value(jeune.date_naissance),
            jeune.nationalite or "",
            jeune.telephone1,
            jeune.telephone2 or "",
            jeune.email,
            jeune.quartier,
            jeune.canton or "",
            jeune.adresse_complete,
            "Oui" if jeune.est_resident_kloto else "Non",
            jeune.get_niveau_etude_display() if jeune.niveau_etude else "",
            jeune.diplome_principal or "",
            jeune.domaine_competence,
            jeune.experiences or "",
            jeune.get_situation_actuelle_display(),
            jeune.employeur_actuel or "",
            jeune.get_disponibilite_display(),
            jeune.get_type_contrat_souhaite_display() if jeune.type_contrat_souhaite else "",
            jeune.salaire_souhaite or "",
            "Oui" if jeune.service_citoyen_obligatoire else "Non",
            "Oui" if jeune.accepte_rgpd else "Non",
            "Oui" if jeune.accepte_contact else "Non",
            "Oui" if jeune.est_valide_par_mairie else "Non",
            _format_excel_value(jeune.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="jeunes_demandeurs_emploi.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_retraites(request):
    """Exporte tous les retraités actifs en Excel avec tous les champs."""
    retraites = ProfilEmploi.objects.filter(type_profil='retraite').order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    niveau = request.GET.get('niveau', '')
    dispo = request.GET.get('dispo', '')
    
    if q:
        retraites = retraites.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone1__icontains=q) |
            Q(domaine_competence__icontains=q)
        )
    if niveau:
        retraites = retraites.filter(niveau_etude=niveau)
    if dispo:
        retraites = retraites.filter(disponibilite=dispo)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Retraites Actifs"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité",
        "Téléphone 1", "Téléphone 2", "Email", "Quartier", "Canton",
        "Adresse complète", "Résident Kloto 1", "Niveau étude", "Diplôme principal",
        "Domaine compétence", "Expériences", "Situation actuelle", "Employeur actuel",
        "Disponibilité", "Type contrat souhaité", "Salaire souhaité",
        "Caisse retraite", "Dernier poste", "Années expérience",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for retraite in retraites:
        row = [
            retraite.pk,
            retraite.nom,
            retraite.prenoms,
            retraite.get_sexe_display(),
            _format_excel_value(retraite.date_naissance),
            retraite.nationalite or "",
            retraite.telephone1,
            retraite.telephone2 or "",
            retraite.email,
            retraite.quartier,
            retraite.canton or "",
            retraite.adresse_complete,
            "Oui" if retraite.est_resident_kloto else "Non",
            retraite.get_niveau_etude_display() if retraite.niveau_etude else "",
            retraite.diplome_principal or "",
            retraite.domaine_competence,
            retraite.experiences or "",
            retraite.get_situation_actuelle_display(),
            retraite.employeur_actuel or "",
            retraite.get_disponibilite_display(),
            retraite.get_type_contrat_souhaite_display() if retraite.type_contrat_souhaite else "",
            retraite.salaire_souhaite or "",
            retraite.caisse_retraite or "",
            retraite.dernier_poste or "",
            retraite.annees_experience if retraite.annees_experience else "",
            "Oui" if retraite.accepte_rgpd else "Non",
            "Oui" if retraite.accepte_contact else "Non",
            "Oui" if retraite.est_valide_par_mairie else "Non",
            _format_excel_value(retraite.date_inscription),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="retraites_actifs.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_diaspora(request):
    """Exporte tous les membres de la diaspora en Excel avec tous les champs."""
    membres = MembreDiaspora.objects.all().order_by('-date_inscription')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    pays = request.GET.get('pays', '')
    secteur = request.GET.get('secteur', '')
    
    if q:
        membres = membres.filter(
            Q(nom__icontains=q) |
            Q(prenoms__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_whatsapp__icontains=q) |
            Q(profession_actuelle__icontains=q) |
            Q(domaine_formation__icontains=q)
        )
    if pays:
        membres = membres.filter(pays_residence_actuelle__icontains=pays)
    if secteur:
        membres = membres.filter(secteur_activite=secteur)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres Diaspora"
    
    # En-têtes
    headers = [
        "ID", "Nom", "Prénoms", "Sexe", "Date naissance", "Nationalité(s)",
        "N° Pièce identité", "Pays résidence", "Ville résidence", "Adresse étranger",
        "Commune origine", "Quartier/Village origine", "Nom parent/tuteur", "Année départ",
        "Fréquence retour", "Téléphone WhatsApp", "Email", "Réseaux sociaux",
        "Contact pays - Nom", "Contact pays - Téléphone", "Niveau études",
        "Domaine formation", "Profession actuelle", "Secteur activité",
        "Secteur activité (autre)", "Années expérience", "Statut professionnel",
        "Type titre séjour", "Appui investissement projets", "Appui financement infrastructures",
        "Appui parrainage communautaire", "Appui jeunes/femmes entrepreneurs",
        "Transfert compétences", "Formation jeunes", "Appui digitalisation",
        "Conseils techniques", "Encadrement mentorat", "Création entreprise locale",
        "Appui PME locales", "Recrutement jeunes commune", "Mise relation ONG",
        "Coopération décentralisée", "Recherche financements internationaux",
        "Promotion commune international", "Participation activités communales",
        "Participation réunions diaspora", "Appui actions sociales/culturelles",
        "Comment contribuer", "Disposition participation", "Domaine intervention prioritaire",
        "Accepte RGPD", "Accepte contact", "Validé par mairie", "Date inscription", "Date modification"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for membre in membres:
        row = [
            membre.pk,
            membre.nom,
            membre.prenoms,
            membre.get_sexe_display(),
            _format_excel_value(membre.date_naissance),
            membre.nationalites,
            membre.numero_piece_identite,
            membre.pays_residence_actuelle,
            membre.ville_residence_actuelle,
            membre.adresse_complete_etranger,
            membre.commune_origine,
            membre.quartier_village_origine,
            membre.nom_parent_tuteur_originaire,
            membre.annee_depart_pays,
            membre.get_frequence_retour_pays_display(),
            membre.telephone_whatsapp,
            membre.email,
            membre.reseaux_sociaux or "",
            membre.contact_au_pays_nom,
            membre.contact_au_pays_telephone,
            membre.get_niveau_etudes_display(),
            membre.domaine_formation,
            membre.profession_actuelle,
            membre.get_secteur_activite_display(),
            membre.secteur_activite_autre or "",
            membre.annees_experience,
            membre.get_statut_professionnel_display(),
            membre.type_titre_sejour or "",
            "Oui" if membre.appui_investissement_projets else "Non",
            "Oui" if membre.appui_financement_infrastructures else "Non",
            "Oui" if membre.appui_parrainage_communautaire else "Non",
            "Oui" if membre.appui_jeunes_femmes_entrepreneurs else "Non",
            "Oui" if membre.transfert_competences else "Non",
            "Oui" if membre.formation_jeunes else "Non",
            "Oui" if membre.appui_digitalisation else "Non",
            "Oui" if membre.conseils_techniques else "Non",
            "Oui" if membre.encadrement_mentorat else "Non",
            "Oui" if membre.creation_entreprise_locale else "Non",
            "Oui" if membre.appui_pme_locales else "Non",
            "Oui" if membre.recrutement_jeunes_commune else "Non",
            "Oui" if membre.mise_relation_ong else "Non",
            "Oui" if membre.cooperation_decentralisee else "Non",
            "Oui" if membre.recherche_financements_internationaux else "Non",
            "Oui" if membre.promotion_commune_international else "Non",
            "Oui" if membre.participation_activites_communales else "Non",
            "Oui" if membre.participation_reunions_diaspora else "Non",
            "Oui" if membre.appui_actions_sociales_culturelles else "Non",
            membre.comment_contribuer,
            membre.get_disposition_participation_display(),
            membre.domaine_intervention_prioritaire,
            "Oui" if membre.accepte_rgpd else "Non",
            "Oui" if membre.accepte_contact else "Non",
            "Oui" if membre.est_valide_par_mairie else "Non",
            _format_excel_value(membre.date_inscription),
            _format_excel_value(membre.date_modification),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="membres_diaspora.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_staff_user)
def export_excel_candidatures(request):
    """Exporte toutes les candidatures en Excel avec tous les champs."""
    candidatures = Candidature.objects.all().select_related('appel_offre', 'candidat').order_by('-date_soumission')
    
    # Appliquer les filtres
    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    
    if q:
        candidatures = candidatures.filter(
            Q(appel_offre__titre__icontains=q) |
            Q(appel_offre__reference__icontains=q) |
            Q(candidat__first_name__icontains=q) |
            Q(candidat__last_name__icontains=q) |
            Q(candidat__email__icontains=q)
        )
    if statut:
        candidatures = candidatures.filter(statut=statut)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    
    # En-têtes
    headers = [
        "ID", "Appel d'offres - Titre", "Appel d'offres - Référence", "Appel d'offres - Description",
        "Appel d'offres - Public cible", "Appel d'offres - Date début", "Appel d'offres - Date fin",
        "Appel d'offres - Budget estimé", "Candidat - Username", "Candidat - Nom", "Candidat - Prénom",
        "Candidat - Email", "Statut candidature", "Message accompagnement", "Date soumission"
    ]
    ws.append(headers)
    _style_excel_header(ws, 1)
    
    # Données
    for candidature in candidatures:
        row = [
            candidature.pk,
            candidature.appel_offre.titre,
            candidature.appel_offre.reference or "",
            candidature.appel_offre.description,
            candidature.appel_offre.get_public_cible_display(),
            _format_excel_value(candidature.appel_offre.date_debut),
            _format_excel_value(candidature.appel_offre.date_fin),
            candidature.appel_offre.budget_estime if candidature.appel_offre.budget_estime else "",
            candidature.candidat.username,
            candidature.candidat.last_name or "",
            candidature.candidat.first_name or "",
            candidature.candidat.email,
            candidature.get_statut_display(),
            candidature.message_accompagnement or "",
            _format_excel_value(candidature.date_soumission),
        ]
        ws.append(row)
    
    # Ajuster la largeur des colonnes
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="candidatures.xlsx"'
    wb.save(response)
    return response
